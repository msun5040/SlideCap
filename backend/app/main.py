from contextlib import asynccontextmanager
from pathlib import Path
from fastapi import FastAPI, HTTPException, Query, Depends, UploadFile, File, Request
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
from io import BytesIO
import csv
import io
import os
import json
import shutil
import subprocess
import zipfile
import queue
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
from datetime import datetime
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import text as sa_text, func

from .config import settings
from .db import init_db, get_db, get_session, Case, Slide, Tag, Project, Cohort, CohortFlag, CohortPatient, CohortPatientCase, Analysis, AnalysisJob, JobSlide, RequestSheet, RequestRow, Study, StudyGroup, init_lock, get_lock, Patient, ExternalMapping, generate_slidecap_id
from .services import SlideHasher, SlideIndexer, ClusterService, JobStatusPoller
from .auth import AuthMiddleware, create_challenge, verify_challenge, create_token, verify_token, cleanup_expired_challenges
from . import demo as demo_mod


def _is_demo() -> bool:
    return settings.APP_MODE == "demo"


# Demo cluster connection state — only used when APP_MODE=demo
_demo_cluster_connected: bool = False

# Demo staging scan cache (so /staging/sort knows what /staging/scan returned)
_demo_staging_files: list[dict] = []


# Request models for bulk operations
class BulkTagRequest(BaseModel):
    slide_hashes: List[str]
    tags: List[str]
    color: Optional[str] = None  # Color for new tags


# Global instances (initialized on startup)
hasher = None
indexer = None
db_lock = None
cluster_service: Optional[ClusterService] = None
job_poller: Optional[JobStatusPoller] = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    global hasher, indexer, db_lock, cluster_service, job_poller

    print("=" * 60)
    print("Starting Slide Organizer API")
    print("=" * 60)

    # Validate configuration
    if not os.path.exists(settings.NETWORK_ROOT):
        print(f"ERROR: Network root does not exist: {settings.NETWORK_ROOT}")
        print("Please update NETWORK_ROOT in config.py or set via environment variable")
        raise RuntimeError(f"Network root not found: {settings.NETWORK_ROOT}")

    if not settings.slides_path.exists():
        print(f"Creating slides directory: {settings.slides_path}")
        settings.slides_path.mkdir(parents=True, exist_ok=True)

    print(f"Network root: {settings.NETWORK_ROOT}")
    print(f"Slides path: {settings.slides_path}")
    print(f"Database: {settings.db_path}")

    # Initialize database lock for multi-user safety
    print("Initializing database lock...")
    db_lock = init_lock(settings.app_data_path)

    # Initialize database (creates session factory)
    print("Initializing database...")
    init_db(settings.db_path)

    # Initialize hasher
    print("Initializing hasher...")
    hasher = SlideHasher(settings.salt_path)

    # Initialize indexer with the configured parser patterns (or built-in default)
    from .services.filename_parser import FilenameParser, patterns_from_config
    parser_patterns = patterns_from_config(settings.PARSER_PATTERNS)
    print(f"Initializing indexer with {len(parser_patterns)} parser pattern(s): "
          f"{', '.join(p.name for p in parser_patterns)}")
    indexer = SlideIndexer(hasher, str(settings.slides_path),
                           parser=FilenameParser(parser_patterns))

    # Build path cache for fast lookups
    print("Building path cache...")
    cache_count = indexer.build_path_cache()
    print(f"Cached {cache_count} slide paths")

    # Clean up any expired auth challenges
    cleanup_expired_challenges()

    # Auto-run incremental indexing to catch new files
    # Use a dedicated session for startup operations
    print("Running incremental index...")
    startup_db = get_session()
    try:
        index_stats = indexer.build_incremental_index(startup_db)
        startup_db.commit()
        if index_stats['new_slides_indexed'] > 0:
            print(f"Indexed {index_stats['new_slides_indexed']} new slides")
        else:
            print("No new slides found")
        # Pull newly-onboarded slides into any case-following cohorts.
        added = reconcile_auto_cohorts(startup_db)
        if added:
            print(f"Auto-added {added} slide(s) to case-following cohorts")
    except Exception:
        startup_db.rollback()
        raise
    finally:
        startup_db.close()

    # Reset any slides that were left in 'transferring' with no cluster_job_id —
    # these are orphans from a previous backend crash mid-rsync. The tmux session
    # was never started so they can never self-recover; mark them failed so the
    # user can resubmit.
    orphan_db = get_session()
    try:
        orphaned = (
            orphan_db.query(JobSlide)
            .filter(
                JobSlide.status == "transferring",
                JobSlide.cluster_job_id.is_(None),
            )
            .all()
        )
        if orphaned:
            affected_jobs = set()
            for js in orphaned:
                js.status = "failed"
                js.error_message = "Transfer interrupted by server restart"
                js.completed_at = datetime.utcnow()
                affected_jobs.add(js.job_id)
            orphan_db.flush()
            for jid in affected_jobs:
                job = orphan_db.query(AnalysisJob).options(
                    joinedload(AnalysisJob.slides)
                ).filter_by(id=jid).first()
                if job:
                    _recompute_job_status(job)
            orphan_db.commit()
            print(f"[Startup] Reset {len(orphaned)} orphaned 'transferring' slide(s) to failed")
    except Exception as e:
        orphan_db.rollback()
        print(f"[Startup] Failed to reset orphaned slides: {e}")
    finally:
        orphan_db.close()

    # Initialize cluster service (connection is per-session via UI)
    cluster_service = ClusterService(
        host=settings.CLUSTER_HOST,
        port=settings.CLUSTER_PORT,
    )
    job_poller = JobStatusPoller(
        cluster_service, interval=15,
        indexer=indexer, analyses_path=settings.analyses_path,
    )
    job_poller.start()
    print("Cluster service initialized (connect via UI to enable job submission)")

    # Demo mode: auto-register CellViT so the demo flow works without manual setup
    if _is_demo():
        print(f"[Demo] APP_MODE=demo — seeding demo analyses, mock cluster active")
        try:
            demo_mod.seed_demo_analyses()
        except Exception as e:
            print(f"[Demo] Failed to seed analyses: {e}")
        # Pre-warm the heavy artifacts for anchor slides so the first click is
        # instant. Runs in a daemon thread so startup isn't blocked — if it
        # takes ~5s for a UMAP fit, that's hidden from the user.
        try:
            import threading
            threading.Thread(
                target=demo_mod.prewarm_anchor_caches,
                name="demo-anchor-prewarm",
                daemon=True,
            ).start()
            print("[Demo] anchor cache pre-warm scheduled")
        except Exception as e:
            print(f"[Demo] Failed to schedule pre-warm: {e}")

    print("=" * 60)
    print(f"API ready at http://{settings.HOST}:{settings.PORT}")
    print("=" * 60)

    yield  # Application runs here

    # Cleanup
    if job_poller:
        job_poller.stop()
    if cluster_service:
        cluster_service.disconnect()
    print("Shutting down...")


# Create FastAPI app
app = FastAPI(
    title="Slide Organizer API",
    description="Backend API for organizing and searching pathology slides",
    version="0.1.0",
    lifespan=lifespan
)

# CORS middleware (allows frontend to connect)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, restrict to your frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Auth middleware (must be added AFTER CORS so CORS handles preflight first)
app.add_middleware(AuthMiddleware)


# ============================================================
# Authentication Endpoints
# ============================================================

class VerifyRequest(BaseModel):
    code: str


@app.post("/auth/challenge")
def auth_challenge():
    """Create a network drive challenge for authentication."""
    try:
        code, file_path = create_challenge()
    except OSError as e:
        raise HTTPException(
            status_code=503,
            detail=f"Cannot write to network drive. Is it mounted? ({e})"
        )
    return {
        "message": "A 6-digit verification code has been written to the network drive.",
        "file_path": file_path,
        "expires_in_minutes": settings.AUTH_CHALLENGE_EXPIRY_MINUTES,
    }


@app.post("/auth/verify")
def auth_verify(body: VerifyRequest):
    """Verify the challenge code and issue a session token."""
    if verify_challenge(body.code):
        token = create_token()
        return {"token": token, "expires_in_days": settings.AUTH_TOKEN_EXPIRY_DAYS}
    raise HTTPException(status_code=401, detail="Invalid or expired code")


@app.get("/auth/status")
def auth_status(request: Request):
    """Check if current token is valid."""
    auth_header = request.headers.get("Authorization", "")
    token = auth_header.replace("Bearer ", "") if auth_header.startswith("Bearer ") else ""
    return {"authenticated": verify_token(token) if token else False}


# ============================================================
# Health & Status Endpoints
# ============================================================

@app.get("/")
def root():
    """API root - basic info."""
    return {
        "name": "Slide Organizer API",
        "version": "0.1.0",
        "status": "running"
    }


@app.get("/health")
def health():
    """Health check endpoint."""
    return {
        "status": "ok",
        "network_root": settings.NETWORK_ROOT,
        "network_accessible": os.path.exists(settings.NETWORK_ROOT),
        "app_mode": settings.APP_MODE,
    }


# ── Parser settings (read-only + test) ────────────────────────────────


class ParserTestRequest(BaseModel):
    filename: str
    regex: Optional[str] = None  # optional ad-hoc regex (overrides configured patterns)


@app.get("/parser/patterns")
def get_parser_patterns():
    """Return the parser patterns currently in use. Read-only — configured via PARSER_PATTERNS env var."""
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")
    return [
        {"name": p.name, "description": p.description, "regex": p.regex.pattern}
        for p in indexer.parser.patterns
    ]


@app.post("/parser/test")
def test_parser(req: ParserTestRequest):
    """
    Try the given filename against every configured pattern (or an ad-hoc
    regex if `regex` is provided). Returns which pattern matched and what
    each pattern would extract.
    """
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")
    from .services.filename_parser import FilenameParser, NamedPattern
    import re as _re

    filename = (req.filename or '').strip()
    if not filename:
        raise HTTPException(status_code=400, detail="filename is required")

    if req.regex:
        try:
            ad_hoc = _re.compile(req.regex, _re.IGNORECASE)
        except _re.error as e:
            raise HTTPException(status_code=400, detail=f"Invalid regex: {e}")
        patterns = [NamedPattern(name="(test pattern)", description="", regex=ad_hoc)]
    else:
        patterns = list(indexer.parser.patterns)

    test_parser_inst = FilenameParser(patterns)
    parsed = test_parser_inst.parse(filename)

    # Per-pattern breakdown so the UI can show which ones did/did not match
    attempts = []
    for p in patterns:
        m = p.regex.match(filename)
        attempts.append({
            "name": p.name,
            "matched": m is not None,
            "groups": m.groupdict() if m else None,
        })

    return {
        "filename": filename,
        "matched": parsed is not None,
        "matched_pattern": parsed.pattern_name if parsed else None,
        "parsed": (
            {
                "accession": parsed.accession,
                "block_id": parsed.block_id,
                "slide_number": parsed.slide_number,
                "stain_type": parsed.stain_type,
                "random_id": parsed.random_id,
                "year": parsed.year,
                "full_stem": parsed.full_stem,
            } if parsed else None
        ),
        "attempts": attempts,
    }


@app.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    """Get index statistics."""
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")
    return indexer.get_stats(db)


@app.get("/thumbnails/stats")
def get_thumbnail_stats():
    """Get thumbnail cache statistics."""
    cache_dir = settings.thumbnail_cache_path

    if not cache_dir.exists():
        return {
            "cached_count": 0,
            "total_slides": len(indexer.slide_hash_to_path),
            "cache_size_mb": 0
        }

    # Count cached thumbnails
    cached_files = list(cache_dir.glob("*.jpg"))
    # Exclude label files
    thumbnail_files = [f for f in cached_files if "_label" not in f.name]
    total_size = sum(f.stat().st_size for f in cached_files)

    return {
        "cached_count": len(thumbnail_files),
        "total_slides": len(indexer.slide_hash_to_path),
        "cache_size_mb": round(total_size / (1024 * 1024), 2)
    }


# ============================================================
# Indexing Endpoints
# ============================================================

def reconcile_auto_cohorts(db: Session, cohort_id: Optional[int] = None) -> int:
    """Ensure every "case-following" cohort contains all slides of the cases it
    tracks (the cases that already have at least one slide in the cohort).

    Called after indexing so newly-onboarded slides auto-join, and when the
    auto_add_cases flag is turned on so existing siblings back-fill immediately.
    Returns the number of slides added. If cohort_id is given, only that cohort
    is processed.
    """
    q = db.query(Cohort).filter(Cohort.auto_add_cases == True)  # noqa: E712
    if cohort_id is not None:
        q = q.filter(Cohort.id == cohort_id)
    cohorts = q.all()

    total_added = 0
    for cohort in cohorts:
        member_slide_ids = {s.id for s in cohort.slides}
        case_ids = {s.case_id for s in cohort.slides if s.case_id is not None}
        if not case_ids:
            continue
        siblings = db.query(Slide).filter(Slide.case_id.in_(case_ids)).all()
        to_add = [s for s in siblings if s.id not in member_slide_ids]
        if to_add:
            cohort.slides.extend(to_add)
            total_added += len(to_add)
            print(f"[AutoCohort] '{cohort.name}': added {len(to_add)} slide(s) for tracked cases")
    if total_added:
        db.commit()
    return total_added


@app.post("/index/full")
def run_full_index(db: Session = Depends(get_db)):
    """
    Run a full index of all slides.

    This scans the entire network drive and updates the database.
    May take several minutes for large collections.
    """
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    print("Starting full index...")
    stats = indexer.build_full_index(db)

    # Rebuild path cache
    cache_count = indexer.build_path_cache()
    stats['cache_rebuilt'] = cache_count

    stats['auto_cohort_slides_added'] = reconcile_auto_cohorts(db)

    print(f"Index complete: {stats}")
    return stats


@app.post("/index/incremental")
def run_incremental_index(db: Session = Depends(get_db)):
    """
    Index only new slides that aren't already in the database.

    Much faster than full index - use this to catch newly added files.
    """
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    print("Starting incremental index...")
    stats = indexer.build_incremental_index(db)
    stats['auto_cohort_slides_added'] = reconcile_auto_cohorts(db)
    print(f"Incremental index complete: {stats['new_slides_indexed']} new slides")
    return stats


@app.post("/index/refresh-cache")
def refresh_cache():
    """Refresh the in-memory path cache without re-indexing database."""
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    count = indexer.build_path_cache()
    return {"cached_slides": count}


@app.get("/index/ghost-slides")
def list_ghost_slides(db: Session = Depends(get_db)):
    """
    List slide records in the DB whose files no longer exist on disk.

    These are orphaned records — typically caused by filename changes,
    parser updates, or deleted files. Shows what would be removed by
    POST /index/cleanup.
    """
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    all_slides = db.query(Slide).options(
        joinedload(Slide.case),
        joinedload(Slide.tags),
        joinedload(Slide.job_slides),
    ).all()

    ghosts = []
    for slide in all_slides:
        if slide.slide_hash not in indexer.slide_hash_to_path:
            has_jobs = any(js.status in ("running", "transferring", "pending") for js in slide.job_slides)
            ghosts.append({
                "slidecap_id": slide.slidecap_id,
                "slide_hash": slide.slide_hash[:16] + "...",
                "block_id": slide.block_id,
                "stain_type": slide.stain_type,
                "case_slidecap_id": slide.case.slidecap_id if slide.case else None,
                "case_year": slide.case.year if slide.case else None,
                "tag_count": len(slide.tags),
                "tags": [t.name for t in slide.tags],
                "job_count": len(slide.job_slides),
                "has_active_jobs": has_jobs,
                "completed_analyses": [
                    js.job.model_name for js in slide.job_slides
                    if js.status == "completed" and js.job
                ],
            })

    # Group by case for readability
    by_case: dict[str, list] = {}
    for g in ghosts:
        key = g["case_slidecap_id"] or "unknown"
        by_case.setdefault(key, []).append(g)

    return {
        "ghost_count": len(ghosts),
        "total_slides_in_db": len(all_slides),
        "cached_slides": len(indexer.slide_hash_to_path),
        "by_case": {
            case_id: {
                "year": slides[0]["case_year"],
                "ghost_slides": slides,
            }
            for case_id, slides in sorted(by_case.items())
        },
    }


@app.post("/index/cleanup")
def cleanup_ghost_slides(
    dry_run: bool = Query(True, description="If true, only report what would be deleted. Set to false to actually delete."),
    db: Session = Depends(get_db),
):
    """
    Remove slide records from the DB whose files no longer exist on disk.

    Skips slides that have active (running/pending/transferring) analysis jobs.
    Also removes empty cases (cases with no remaining slides after cleanup).

    Use dry_run=true (default) to preview, dry_run=false to execute.
    """
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    all_slides = db.query(Slide).options(
        joinedload(Slide.case),
        joinedload(Slide.job_slides),
    ).all()

    to_delete = []
    skipped = []
    for slide in all_slides:
        if slide.slide_hash in indexer.slide_hash_to_path:
            continue  # File exists, keep it

        has_active = any(js.status in ("running", "transferring", "pending") for js in slide.job_slides)
        if has_active:
            skipped.append({
                "slidecap_id": slide.slidecap_id,
                "reason": "has active analysis jobs",
            })
            continue

        to_delete.append(slide)

    # Find cases that would become empty
    case_slide_counts: dict[int, int] = {}
    for slide in all_slides:
        case_slide_counts[slide.case_id] = case_slide_counts.get(slide.case_id, 0) + 1
    for slide in to_delete:
        case_slide_counts[slide.case_id] -= 1

    empty_case_ids = [cid for cid, count in case_slide_counts.items() if count <= 0]
    empty_cases = db.query(Case).filter(Case.id.in_(empty_case_ids)).all() if empty_case_ids else []

    result = {
        "dry_run": dry_run,
        "slides_to_remove": len(to_delete),
        "slides_skipped": skipped,
        "cases_to_remove": len(empty_cases),
        "removed_slides": [
            {"slidecap_id": s.slidecap_id, "block_id": s.block_id, "stain_type": s.stain_type}
            for s in to_delete
        ],
        "removed_cases": [
            {"slidecap_id": c.slidecap_id, "year": c.year}
            for c in empty_cases
        ],
    }

    if not dry_run:
        for slide in to_delete:
            db.delete(slide)
        for case in empty_cases:
            db.delete(case)
        db.flush()
        result["status"] = "deleted"
    else:
        result["status"] = "dry_run — POST with dry_run=false to execute"

    return result


# ============================================================
# Search Endpoints
# ============================================================

@app.get("/search")
def search_slides(
    db: Session = Depends(get_db),
    q: Optional[str] = Query(None, description="Search query (matches against filename). If empty, returns all slides matching filters."),
    year: Optional[int] = Query(None, description="Filter by year"),
    stain: Optional[str] = Query(None, description="Filter by stain type: HE (exact), IHC (prefix match), Special (not HE or IHC)"),
    tag: Optional[str] = Query(None, description="Filter by tag name"),
    limit: int = Query(500, le=500, description="Maximum results")
):
    """
    Search for slides by accession number or other filename components.

    Supports partial matching (e.g., "S24-123" will match "S24-12345").
    If no query is provided, returns all slides matching the filters (up to limit).
    """
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    # Use empty string if no query provided - indexer will return all slides
    search_query = q or ""

    results = indexer.search(
        db=db,
        query=search_query,
        year=year,
        stain_type=stain,
        limit=limit
    )

    # Filter by tag if specified
    if tag:
        tag_obj = db.query(Tag).filter_by(name=tag).first()
        if tag_obj:
            # Get slide hashes that have this tag
            tagged_hashes = set()
            for slide in tag_obj.slides:
                tagged_hashes.add(slide.slide_hash)
            # Filter results to only include slides with this tag
            results = [r for r in results if r.get('slide_hash') in tagged_hashes]
        else:
            # Tag doesn't exist, return empty results
            results = []

    # Enrich results with request sheet info
    if results:
        accessions = {_normalize_accession(r['accession_number']) for r in results if r.get('accession_number')}
        if accessions:
            req_rows = db.query(RequestRow).options(joinedload(RequestRow.sheet)).filter(
                RequestRow.accession_number.in_(accessions)
            ).all()
            # Also check normalized forms
            acc_to_sheets: dict[str, list] = {}
            for rr in req_rows:
                norm = _normalize_accession(rr.accession_number)
                if norm not in acc_to_sheets:
                    acc_to_sheets[norm] = []
                acc_to_sheets[norm].append({
                    'sheet_id': rr.sheet_id,
                    'sheet_name': rr.sheet.name if rr.sheet else None,
                    'case_status': rr.case_status,
                })
            for r in results:
                norm = _normalize_accession(r.get('accession_number', ''))
                r['request_sheets'] = acc_to_sheets.get(norm, [])

    return {
        "query": q,
        "filters": {"year": year, "stain": stain, "tag": tag},
        "count": len(results),
        "truncated": len(results) == limit,
        "results": results
    }


# ============================================================
# Bulk Tag Endpoints (must be before /slides/{slide_hash} routes)
# ============================================================

@app.post("/slides/bulk/tags/add")
def bulk_add_tags(request: BulkTagRequest, db: Session = Depends(get_db)):
    """Add tags to multiple slides at once."""
    updated = []
    not_found = []

    with get_lock().write_lock():
        for slide_hash in request.slide_hashes:
            slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
            if not slide:
                not_found.append(slide_hash)
                continue

            for tag_name in request.tags:
                tag = db.query(Tag).filter_by(name=tag_name).first()
                if not tag:
                    tag = Tag(name=tag_name, color=request.color)
                    db.add(tag)
                elif request.color and not tag.color:
                    # Backfill color if existing tag has none
                    tag.color = request.color

                if tag not in slide.tags:
                    slide.tags.append(tag)

            updated.append(slide_hash)

        db.commit()  # Commit while holding lock to ensure data is written

    return {
        "status": "ok",
        "updated": len(updated),
        "not_found": not_found
    }


@app.post("/slides/bulk/tags/remove")
def bulk_remove_tags(request: BulkTagRequest, db: Session = Depends(get_db)):
    """Remove tags from multiple slides at once."""
    updated = []
    not_found = []

    with get_lock().write_lock():
        for slide_hash in request.slide_hashes:
            slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
            if not slide:
                not_found.append(slide_hash)
                continue

            for tag_name in request.tags:
                tag = db.query(Tag).filter_by(name=tag_name).first()
                if tag and tag in slide.tags:
                    slide.tags.remove(tag)

            updated.append(slide_hash)

        db.commit()  # Commit while holding lock

    return {
        "status": "ok",
        "updated": len(updated),
        "not_found": not_found
    }


@app.put("/slides/bulk/tags")
def bulk_set_tags(request: BulkTagRequest, db: Session = Depends(get_db)):
    """Replace all tags on multiple slides (removes existing tags first)."""
    updated = []
    not_found = []

    with get_lock().write_lock():
        tags_to_set = []
        for tag_name in request.tags:
            tag = db.query(Tag).filter_by(name=tag_name).first()
            if not tag:
                tag = Tag(name=tag_name)
                db.add(tag)
            tags_to_set.append(tag)

        for slide_hash in request.slide_hashes:
            slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
            if not slide:
                not_found.append(slide_hash)
                continue

            slide.tags = tags_to_set.copy()
            updated.append(slide_hash)

        db.commit()  # Commit while holding lock

    return {
        "status": "ok",
        "updated": len(updated),
        "not_found": not_found
    }


@app.get("/slides/{slide_hash}")
def get_slide(slide_hash: str, db: Session = Depends(get_db)):
    """
    Get full details for a specific slide by its hash.

    Returns the same shape as `/search` rows so the frontend can drive the
    global SlideDetailsDialog by hash alone — adds request_sheets, file_path,
    file_size, and a small `analysis_jobs` list keyed off the JobSlide rows.
    """
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    slide = (
        db.query(Slide)
        .options(joinedload(Slide.case), joinedload(Slide.tags), joinedload(Slide.job_slides).joinedload(JobSlide.job))
        .filter_by(slide_hash=slide_hash)
        .first()
    )
    if not slide:
        raise HTTPException(status_code=404, detail="Slide not found")

    # Reuse the same dict-builder /search uses — keeps the frontend's `Slide`
    # type compatible whether you came from a search row or opened by hash.
    result = indexer._slide_to_result(slide)

    filepath = indexer.get_filepath(slide_hash)
    result["file_path"] = str(filepath) if filepath else None
    result["filepath_available"] = filepath is not None
    result["file_exists"] = bool(slide.file_exists)

    # Mirror /search's request-sheet enrichment so the dialog can show which
    # request sheets reference this slide's accession.
    accession = result.get("accession_number")
    if accession:
        norm = _normalize_accession(accession)
        req_rows = db.query(RequestRow).options(joinedload(RequestRow.sheet)).filter(
            RequestRow.accession_number == norm
        ).all()
        result["request_sheets"] = [
            {
                "sheet_id": rr.sheet_id,
                "sheet_name": rr.sheet.name if rr.sheet else None,
                "case_status": rr.case_status,
            }
            for rr in req_rows
        ]
    else:
        result["request_sheets"] = []

    # Per-job analysis status (used by the dialog's "completed analyses"
    # section to link out to specific job runs, not just model names).
    result["analysis_jobs"] = [
        {
            "id": js.job.id,
            "model": js.job.model_name,
            "status": js.status,
            "submitted_at": js.job.submitted_at.isoformat() if js.job.submitted_at else None,
        }
        for js in slide.job_slides if js.job
    ]

    return result


# ============================================================
# Slide Viewer Endpoints (OpenSlide)
# ============================================================

from openslide import open_slide
from openslide.deepzoom import DeepZoomGenerator

# ── Tile source via `large_image` ────────────────────────────────────────
#
# Replaced the openslide.DeepZoomGenerator + PIL re-encode pipeline (pre-
# large_image snapshot saved at backend/app/main.py.pre-large-image).
# Why: SVS files store JPEG-compressed tiles in their native pyramid. The
# old path decoded those JPEGs to RGBA, converted to RGB, re-encoded to
# JPEG (~20-30ms per tile), and shipped them. large_image's tiff source
# reads the raw JPEG bytes from the SVS and sends them unchanged — same
# bytes the scanner produced, no re-encode tax. Empirical (warm openslide
# vs warm tiff source on a 116k×38k SVS, mid-zoom): ~3 ms/tile → ~0 ms/tile.
import large_image

# Cache for large_image TileSource handles (keeps the slide open, IFD
# parsed, page cache warm — same idea as DeepZoomGenerator cache).
_ts_cache: dict[str, "large_image.tilesource.TileSource"] = {}

# Native SVS tile size. Most scanners write 256×256 tiles; using the same
# size in the DZI manifest means OSD asks for tiles at exactly the
# positions where they're already encoded — so tiff-source passthrough hits
# the fast path on every request.
TILE_SIZE = 256
# Overlap=0 because native SVS tiles have no overlap. OSD handles
# unoverlapped tiles fine; the only cost is no edge-blending at extreme
# zoom-in, which is invisible on H&E.
TILE_OVERLAP = 0
TILE_FORMAT = "jpeg"
TILE_JPEG_QUALITY = 80  # only used by the openslide-fallback path
TILE_CACHE_VERSION = "v3"  # bump to invalidate the on-disk tile cache
# v3: switched tile pipeline to large_image native passthrough; tile size
#     changed from 510 (with overlap) to 256 (without). v2 cache is
#     incompatible — left in place but ignored by the v3 path.

def _get_ts(slide_hash: str):
    """
    Get or create a large_image TileSource for a slide.

    Prefers the ``tiff`` source (true native-JPEG passthrough — bytes go
    SVS → HTTP unchanged, no decode/encode). Falls back to ``openslide``
    if tifffile can't open the file (rare formats, non-standard SVS, etc.).
    Either way it's substantially faster than the openslide.DeepZoomGenerator
    + PIL re-encode pipeline we used to run.
    """
    cached = _ts_cache.get(slide_hash)
    if cached is not None:
        return cached

    filepath = indexer.get_filepath(slide_hash)
    if not filepath or not filepath.exists():
        raise HTTPException(status_code=404, detail="Slide file not found")

    # Crucial: pass `encoding='JPEG'` BUT NOT `jpegQuality`. Setting a
    # quality forces re-encoding (~30 ms/tile); omitting it lets large_image
    # ship the SVS's native JPEG bytes unchanged (~0 ms/tile). The
    # openslide fallback does want jpegQuality since it has to re-encode
    # from RGBA anyway; we set it post-open so it only applies there.
    last_err: Exception | None = None
    for source in ("tiff", "openslide"):
        try:
            ts = large_image.open(str(filepath), sourceName=source, encoding="JPEG")
            if source == "openslide":
                # Re-encoding fallback path benefits from quality control.
                ts._jpegQuality = TILE_JPEG_QUALITY
            _ts_cache[slide_hash] = ts
            return ts
        except Exception as e:
            last_err = e
            continue
    raise HTTPException(status_code=500, detail=f"Failed to open slide via large_image: {last_err}")


def _dzi_max_level(width: int, height: int) -> int:
    """DZI max level for an image of WxH px (the highest level = full res)."""
    import math
    return max(0, math.ceil(math.log2(max(width, height))))


# Per-slide metadata cache. getMetadata() on large_image is cheap (~1 ms)
# but called once per tile request it adds up. This caches the only fields
# we need: width, height, level count, native tile size.
_md_cache: dict[str, tuple[int, int, int, int, int]] = {}


def _md_for(slide_hash: str) -> tuple[int, int, int, int, int]:
    """Returns (width, height, levels, tile_size, max_dzi_level). Cached."""
    cached = _md_cache.get(slide_hash)
    if cached is not None:
        return cached
    ts = _get_ts(slide_hash)
    md = ts.getMetadata()
    w, h, levels = md["sizeX"], md["sizeY"], md["levels"]
    ts_sz = md.get("tileWidth") or TILE_SIZE
    out = (w, h, levels, ts_sz, _dzi_max_level(w, h))
    _md_cache[slide_hash] = out
    return out


def _dzi_to_zxy(slide_hash: str, level: int, col: int, row: int):
    """
    Convert a DZI (level, col, row) request into large_image (x, y, z).

    DZI levels run 0..max_level where max_level = ceil(log2(max(W, H))).
    large_image's z runs 0..levels-1 where 0 is the smallest pyramid level
    available. The alignment between the two depends on how many levels
    large_image found in the SVS.

    Returns (z, x, y) suitable for ``ts.getTile``. If the DZI level is
    below large_image's smallest level (extreme zoom-out), clamps z to 0
    and returns col=0,row=0 — OSD's home view rarely lands there.
    """
    _, _, levels, _, max_dzi = _md_for(slide_hash)
    z = level - (max_dzi - (levels - 1))
    if z < 0:
        return 0, 0, 0
    if z >= levels:
        raise HTTPException(status_code=404, detail=f"Invalid DZI level {level} (max for this slide is {max_dzi})")
    return z, col, row


def _get_and_cache_embedded_thumbnail(slide_hash: str, filepath: Path) -> bytes:
    """
    Extract and cache the embedded thumbnail from the SVS file.
    Only stores one small file (~20-30KB) per slide.
    Returns JPEG bytes.
    """
    cache_dir = settings.thumbnail_cache_path
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"{slide_hash}.jpg"

    try:
        slide = open_slide(str(filepath))

        # Get the embedded thumbnail (these are pre-stored in SVS files)
        thumb = None
        if 'thumbnail' in slide.associated_images:
            thumb = slide.associated_images['thumbnail']
        elif 'macro' in slide.associated_images:
            thumb = slide.associated_images['macro']

        if thumb is None:
            # Fallback: get from lowest pyramid level (rare)
            level = slide.level_count - 1
            dims = slide.level_dimensions[level]
            thumb = slide.read_region((0, 0), level, dims)

        slide.close()

        # Convert and cache
        thumb = thumb.convert('RGB')
        buffer = BytesIO()
        thumb.save(buffer, format='JPEG', quality=85)
        image_bytes = buffer.getvalue()

        cache_file.write_bytes(image_bytes)
        return image_bytes

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to extract thumbnail: {e}")


@app.get("/slides/{slide_hash}/thumbnail.jpeg")
def get_slide_thumbnail(slide_hash: str, max_size: int = Query(1024, le=2048)):
    """
    Get the embedded thumbnail from the slide.
    Caches only the original embedded thumbnail (~20-30KB per slide).
    Resizes on-the-fly if a smaller size is requested.
    """
    from PIL import Image

    cache_dir = settings.thumbnail_cache_path
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"{slide_hash}.jpg"

    # Check cache first
    if cache_file.exists():
        image_bytes = cache_file.read_bytes()
    else:
        # Not cached - extract from SVS file
        filepath = indexer.get_filepath(slide_hash)
        if not filepath or not filepath.exists():
            raise HTTPException(status_code=404, detail="Slide file not found")
        image_bytes = _get_and_cache_embedded_thumbnail(slide_hash, filepath)

    # Resize on-the-fly if requested size is smaller than cached
    thumb = Image.open(BytesIO(image_bytes))
    if max(thumb.size) > max_size:
        ratio = max_size / max(thumb.size)
        new_size = (int(thumb.size[0] * ratio), int(thumb.size[1] * ratio))
        thumb = thumb.resize(new_size)
        buffer = BytesIO()
        thumb.save(buffer, format='JPEG', quality=85)
        image_bytes = buffer.getvalue()

    return Response(
        content=image_bytes,
        media_type="image/jpeg",
        headers={"Cache-Control": "public, max-age=86400"}
    )


@app.get("/slides/{slide_hash}/label.jpeg")
def get_slide_label(slide_hash: str, max_size: int = Query(256, le=512)):
    """
    Get the slide label image (the paper label on the physical slide).
    Returns 404 if no label is available.
    """
    # Check cache first
    cache_dir = settings.thumbnail_cache_path
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"{slide_hash}_label.jpg"

    if cache_file.exists():
        return Response(
            content=cache_file.read_bytes(),
            media_type="image/jpeg",
            headers={"Cache-Control": "public, max-age=86400"}
        )

    filepath = indexer.get_filepath(slide_hash)
    if not filepath or not filepath.exists():
        raise HTTPException(status_code=404, detail="Slide file not found")

    try:
        slide = open_slide(str(filepath))

        # Try to get label image
        label = None
        if 'label' in slide.associated_images:
            label = slide.associated_images['label']

        slide.close()

        if label is None:
            raise HTTPException(status_code=404, detail="No label image available")

        # Resize if needed
        if max(label.size) > max_size:
            ratio = max_size / max(label.size)
            new_size = (int(label.size[0] * ratio), int(label.size[1] * ratio))
            label = label.resize(new_size)

        # Convert and cache
        label = label.convert('RGB')
        buffer = BytesIO()
        label.save(buffer, format='JPEG', quality=90)
        image_bytes = buffer.getvalue()

        cache_file.write_bytes(image_bytes)

        return Response(
            content=image_bytes,
            media_type="image/jpeg",
            headers={"Cache-Control": "public, max-age=86400"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get label: {e}")


# Slides we've kicked off a background warm for in this process. Stops us
# from spawning N warms when OSD races multiple requests on first open.
_prewarmed_in_session: set[str] = set()
_prewarm_lock = threading.Lock()


def _schedule_warm_on_first_open(slide_hash: str) -> None:
    """
    On the FIRST DZI request per slide per process, kick off a background
    thread that pre-generates the bottom ~300 tiles for this slide. Idea:
    by the time the user has finished staring at the initial 20–30 visible
    tiles, the rest of the low-res mosaic is already cached on disk, and
    every pan/zoom they make is served from the cache (sub-ms).

    Idempotent within a process — once warmed, no-op. Tiles persist on disk
    across restarts, so future processes also see the cache.
    """
    with _prewarm_lock:
        if slide_hash in _prewarmed_in_session:
            return
        _prewarmed_in_session.add(slide_hash)

    def _run():
        try:
            n = prewarm_slide_tiles(slide_hash, max_tiles=300)
            if n:
                print(f"[tile-warm-on-open] {slide_hash[:12]}: generated {n} tiles in background")
        except Exception as e:
            print(f"[tile-warm-on-open] {slide_hash[:12]}: {e}")
    threading.Thread(target=_run, name=f"warm-{slide_hash[:8]}", daemon=True).start()


@app.get("/slides/{slide_hash}/dzi.json")
def get_slide_dzi(slide_hash: str):
    """
    Get Deep Zoom Image metadata for a slide.
    Returns tile size, format, and dimensions for OpenSeadragon.

    Also fires off a background tile pre-warm on first open of each slide
    in this process — see `_schedule_warm_on_first_open`. This is the
    "second open is instant" trick: after first open kicks off the warm,
    every future pan/zoom + future reopen comes from disk cache.
    """
    width, height, _, tile_size, _ = _md_for(slide_hash)

    _schedule_warm_on_first_open(slide_hash)

    return {
        "Image": {
            "xmlns": "http://schemas.microsoft.com/deepzoom/2008",
            "Format": TILE_FORMAT,
            "Overlap": str(TILE_OVERLAP),
            "TileSize": str(tile_size),
            "Size": {
                "Width": str(width),
                "Height": str(height)
            }
        }
    }


def _tile_cache_path(slide_hash: str, level: int, col: int, row: int) -> Path:
    """
    On-disk path for a cached tile. Namespaced by tile size + format so a
    config bump (TILE_CACHE_VERSION, TILE_SIZE) doesn't collide with old
    files — old cache becomes orphaned, not corrupt.
    """
    bucket = f"{TILE_CACHE_VERSION}_{TILE_SIZE}_{TILE_FORMAT}"
    return settings.local_data_path / "tile-cache" / slide_hash / bucket / str(level) / f"{col}_{row}.{TILE_FORMAT}"


def _generate_tile_bytes(slide_hash: str, level: int, col: int, row: int) -> bytes:
    """
    Fetch a single tile as JPEG bytes via large_image (no cache I/O).

    On the tiff source this is a native passthrough — the JPEG bytes stored
    in the SVS are returned unchanged. On the openslide fallback it's still
    way cheaper than the old PIL re-encode because large_image avoids the
    RGBA→RGB→JPEG round-trip.
    """
    ts = _get_ts(slide_hash)
    z, x, y = _dzi_to_zxy(slide_hash, level, col, row)
    return ts.getTile(x, y, z)


@app.get("/slides/{slide_hash}/tiles/{level}/{col}_{row}.jpeg")
def get_slide_tile(slide_hash: str, level: int, col: int, row: int):
    """
    Serve a single DZI tile.

    Two fast paths:
      1. Disk cache hit — static file serve, ~1 ms.
      2. large_image native passthrough — SVS stores JPEG tiles in its
         pyramid; we read the bytes and ship them unchanged, ~3-15 ms vs
         the 30-80 ms of the old openslide+PIL re-encode pipeline.

    Cold cache miss writes the bytes to disk for the next request.
    """
    cache_path = _tile_cache_path(slide_hash, level, col, row)
    if cache_path.exists():
        return FileResponse(
            str(cache_path),
            media_type="image/jpeg",
            headers={"Cache-Control": "public, max-age=86400"},
        )

    try:
        image_bytes = _generate_tile_bytes(slide_hash, level, col, row)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read tile: {e}")

    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_bytes(image_bytes)
    except Exception as cache_err:
        print(f"[tile-cache] write failed for {slide_hash[:12]} L{level} {col},{row}: {cache_err}")

    return Response(
        content=image_bytes,
        media_type="image/jpeg",
        headers={"Cache-Control": "public, max-age=86400"},
    )


def prewarm_slide_tiles(slide_hash: str, max_tiles: int = 300) -> int:
    """
    Pre-generate + disk-cache the lowest-resolution tiles so the initial
    "fit to screen" view in OSD pops up instantly. Walks DZI levels
    bottom-up (smallest image first) until we hit `max_tiles`, then stops —
    this caps disk usage and prewarm time regardless of slide size.

    Returns the number of tiles generated. Safe to no-op if the slide file
    is missing.
    """
    try:
        _get_ts(slide_hash)
    except HTTPException:
        print(f"[tile-prewarm] slide {slide_hash[:12]}: file missing — skipping")
        return 0

    width, height, _, tile_size, max_dzi = _md_for(slide_hash)

    generated = 0
    # Iterate low-res → high-res by DZI level. Same level math the tile
    # endpoint uses, so cached entries are the ones OSD will actually fetch.
    import math
    for level in range(max_dzi + 1):
        # Image dims at this DZI level
        scale = 2 ** (max_dzi - level)
        w_l = max(1, math.ceil(width / scale))
        h_l = max(1, math.ceil(height / scale))
        cols = max(1, math.ceil(w_l / tile_size))
        rows = max(1, math.ceil(h_l / tile_size))
        for row in range(rows):
            for col in range(cols):
                if generated >= max_tiles:
                    return generated
                cache_path = _tile_cache_path(slide_hash, level, col, row)
                if cache_path.exists():
                    continue
                try:
                    data = _generate_tile_bytes(slide_hash, level, col, row)
                    cache_path.parent.mkdir(parents=True, exist_ok=True)
                    cache_path.write_bytes(data)
                    generated += 1
                except Exception as e:
                    print(f"[tile-prewarm] {slide_hash[:12]} L{level} {col},{row} failed: {e}")
    return generated


# ============================================================
# Thumbnail Pre-generation
# ============================================================

# ============================================================
# Tag Endpoints
# ============================================================

class TagCreate(BaseModel):
    name: str
    color: Optional[str] = None  # Hex color like "#FF5733"
    category: Optional[str] = None


class TagUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None
    category: Optional[str] = None


@app.get("/tags")
def list_tags(db: Session = Depends(get_db)):
    """List all tags with usage counts."""
    tags = db.query(Tag).order_by(Tag.name).all()
    return [
        {
            "id": t.id,
            "name": t.name,
            "color": t.color,
            "category": t.category,
            "slide_count": len(t.slides),
            "case_count": len(t.cases)
        }
        for t in tags
    ]


@app.get("/tags/search")
def search_tags(db: Session = Depends(get_db), q: str = Query(..., min_length=1)):
    """Search/autocomplete tags by name prefix."""
    tags = db.query(Tag).filter(
        Tag.name.ilike(f"{q}%")
    ).order_by(Tag.name).limit(10).all()

    return [
        {
            "id": t.id,
            "name": t.name,
            "color": t.color,
            "category": t.category
        }
        for t in tags
    ]


@app.post("/tags")
def create_tag(tag_data: TagCreate, db: Session = Depends(get_db)):
    """Create a new tag."""
    existing = db.query(Tag).filter_by(name=tag_data.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Tag already exists")

    tag = Tag(name=tag_data.name, color=tag_data.color, category=tag_data.category)
    db.add(tag)
    db.commit()

    return {"id": tag.id, "name": tag.name, "color": tag.color, "category": tag.category}


@app.patch("/tags/{tag_id}")
def update_tag(tag_id: int, tag_data: TagUpdate, db: Session = Depends(get_db)):
    """Update a tag's name, color, or category."""
    tag = db.query(Tag).filter_by(id=tag_id).first()
    if not tag:
        raise HTTPException(status_code=404, detail="Tag not found")

    if tag_data.name is not None:
        # Check for name conflict
        existing = db.query(Tag).filter(Tag.name == tag_data.name, Tag.id != tag_id).first()
        if existing:
            raise HTTPException(status_code=400, detail="Tag name already exists")
        tag.name = tag_data.name

    if tag_data.color is not None:
        tag.color = tag_data.color

    if tag_data.category is not None:
        tag.category = tag_data.category

    db.commit()

    return {"id": tag.id, "name": tag.name, "color": tag.color, "category": tag.category}


@app.delete("/tags/{tag_id}")
def delete_tag(tag_id: int, db: Session = Depends(get_db)):
    """Delete a tag."""
    tag = db.query(Tag).filter_by(id=tag_id).first()
    if not tag:
        raise HTTPException(status_code=404, detail="Tag not found")

    db.delete(tag)
    db.commit()

    return {"status": "ok"}


@app.get("/tags/{tag_name}/slides")
def get_slides_by_tag(tag_name: str, db: Session = Depends(get_db), limit: int = Query(100, le=500)):
    """Get all slides with a given tag."""
    tag = db.query(Tag).filter_by(name=tag_name).first()
    if not tag:
        raise HTTPException(status_code=404, detail="Tag not found")

    slides = tag.slides[:limit]

    return {
        "tag": tag_name,
        "color": tag.color,
        "count": len(slides),
        "slides": [
            {
                "slide_hash": s.slide_hash,
                "case_hash": s.case.accession_hash,
                "year": s.case.year,
                "block_id": s.block_id,
                "stain_type": s.stain_type,
                "file_exists": bool(s.file_exists),
            }
            for s in slides
        ]
    }


@app.get("/slides/{slide_hash}/tags")
def get_slide_tags(slide_hash: str, db: Session = Depends(get_db)):
    """Get all tags for a slide."""
    slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
    if not slide:
        raise HTTPException(status_code=404, detail="Slide not found")

    return [
        {
            "id": t.id,
            "name": t.name,
            "color": t.color,
            "category": t.category
        }
        for t in slide.tags
    ]


class AddTagRequest(BaseModel):
    name: str
    color: Optional[str] = None  # Only used if creating new tag


@app.post("/slides/{slide_hash}/tags")
def add_tag_to_slide(slide_hash: str, tag_data: AddTagRequest, db: Session = Depends(get_db)):
    """Add a tag to a slide. Creates the tag if it doesn't exist."""
    slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
    if not slide:
        raise HTTPException(status_code=404, detail="Slide not found")

    with get_lock().write_lock():
        tag = db.query(Tag).filter_by(name=tag_data.name).first()
        if not tag:
            # Auto-create the tag with color
            tag = Tag(name=tag_data.name, color=tag_data.color)
            db.add(tag)
            db.flush()  # Get the ID
        elif tag_data.color and not tag.color:
            # Backfill color if existing tag has none
            tag.color = tag_data.color

        if tag not in slide.tags:
            slide.tags.append(tag)

        db.commit()

    return {
        "status": "ok",
        "tag": {
            "id": tag.id,
            "name": tag.name,
            "color": tag.color,
            "category": tag.category
        }
    }


@app.post("/slides/{slide_hash}/tags/{tag_name}")
def add_tag_to_slide_by_name(slide_hash: str, tag_name: str, db: Session = Depends(get_db)):
    """Add a tag to a slide by name (legacy endpoint)."""
    slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
    if not slide:
        raise HTTPException(status_code=404, detail="Slide not found")

    with get_lock().write_lock():
        tag = db.query(Tag).filter_by(name=tag_name).first()
        if not tag:
            # Auto-create the tag
            tag = Tag(name=tag_name)
            db.add(tag)

        if tag not in slide.tags:
            slide.tags.append(tag)
            db.commit()

    return {"status": "ok", "slide_hash": slide_hash, "tag": tag_name}


@app.delete("/slides/{slide_hash}/tags/{tag_name}")
def remove_tag_from_slide(slide_hash: str, tag_name: str, db: Session = Depends(get_db)):
    """Remove a tag from a slide."""
    slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
    if not slide:
        raise HTTPException(status_code=404, detail="Slide not found")

    with get_lock().write_lock():
        tag = db.query(Tag).filter_by(name=tag_name).first()
        if tag and tag in slide.tags:
            slide.tags.remove(tag)
            db.commit()

    return {"status": "ok"}


# ============================================================
# Project Endpoints
# ============================================================

@app.get("/projects")
def list_projects(db: Session = Depends(get_db)):
    """List all projects."""
    projects = db.query(Project).order_by(Project.created_at.desc()).all()
    return [
        {
            "id": p.id,
            "name": p.name,
            "description": p.description,
            "case_count": p.case_count,
            "slide_count": p.slide_count,
            "created_at": p.created_at.isoformat()
        }
        for p in projects
    ]


@app.post("/projects")
def create_project(db: Session = Depends(get_db), name: str = Query(...), description: Optional[str] = None, created_by: Optional[str] = None):
    """Create a new project."""
    with get_lock().write_lock():
        project = Project(name=name, description=description, created_by=created_by)
        db.add(project)
        db.commit()

    return {"id": project.id, "name": project.name}


@app.get("/projects/{project_id}")
def get_project(project_id: int, db: Session = Depends(get_db)):
    """Get project details including all cases."""
    project = db.query(Project).filter_by(id=project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    return {
        "id": project.id,
        "name": project.name,
        "description": project.description,
        "created_by": project.created_by,
        "created_at": project.created_at.isoformat(),
        "case_count": project.case_count,
        "slide_count": project.slide_count,
        "cases": [
            {
                "case_hash": c.accession_hash,
                "year": c.year,
                "slide_count": len(c.slides),
                "tags": [t.name for t in c.tags]
            }
            for c in project.cases
        ]
    }


@app.post("/projects/{project_id}/cases/{case_hash}")
def add_case_to_project(project_id: int, case_hash: str, db: Session = Depends(get_db)):
    """Add a case to a project."""
    project = db.query(Project).filter_by(id=project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    case = db.query(Case).filter_by(accession_hash=case_hash).first()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    with get_lock().write_lock():
        if case not in project.cases:
            project.cases.append(case)
            db.commit()

    return {"status": "ok", "project_id": project_id, "case_hash": case_hash}


# ============================================================
# Annotation Endpoints (filesystem-based, no DB table)
# ============================================================

def _annotations_dir(slide_hash: str) -> Path:
    """Return the annotations directory for a slide, creating it if needed."""
    d = settings.annotations_path / slide_hash
    d.mkdir(parents=True, exist_ok=True)
    return d


@app.get("/slides/{slide_hash}/annotations")
def list_annotations(slide_hash: str):
    """List annotation files for a slide."""
    ann_dir = settings.annotations_path / slide_hash
    if not ann_dir.is_dir():
        return []

    return [
        {
            "name": f.name,
            "size": f.stat().st_size,
        }
        for f in sorted(ann_dir.iterdir())
        if f.is_file()
    ]


@app.post("/slides/{slide_hash}/annotations")
async def upload_annotation(slide_hash: str, file: UploadFile = File(...)):
    """Upload an annotation file for a slide."""
    # Validate the slide exists
    db = get_session()
    try:
        slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
        if not slide:
            raise HTTPException(status_code=404, detail="Slide not found")
    finally:
        db.close()

    # Validate filename
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    safe_name = Path(file.filename).name
    if '..' in safe_name or '/' in safe_name:
        raise HTTPException(status_code=400, detail="Invalid filename")

    ann_dir = _annotations_dir(slide_hash)
    dest = ann_dir / safe_name

    content = await file.read()
    dest.write_bytes(content)

    return {"status": "ok", "filename": safe_name, "size": len(content)}


@app.delete("/slides/{slide_hash}/annotations/{filename}")
def delete_annotation(slide_hash: str, filename: str):
    """Delete a single annotation file."""
    safe_name = Path(filename).name
    file_path = settings.annotations_path / slide_hash / safe_name
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="Annotation file not found")

    file_path.unlink()
    return {"status": "ok"}


@app.get("/slides/{slide_hash}/annotations/{filename}")
def get_annotation_file(slide_hash: str, filename: str):
    """Download a single annotation file."""
    safe_name = Path(filename).name
    file_path = settings.annotations_path / slide_hash / safe_name
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="Annotation file not found")

    return FileResponse(str(file_path), filename=safe_name)


# ============================================================
# Cohort Endpoints
# ============================================================

class CohortCreate(BaseModel):
    name: str
    description: Optional[str] = None
    created_by: Optional[str] = None
    auto_add_cases: bool = False


class CohortUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    auto_add_cases: Optional[bool] = None


class CohortFromTag(BaseModel):
    name: str
    description: Optional[str] = None
    tag_name: str
    created_by: Optional[str] = None


class CohortAddSlides(BaseModel):
    slide_hashes: List[str]


class CohortFlagCreate(BaseModel):
    name: str
    case_hashes: List[str] = []


class CohortFlagPatch(BaseModel):
    add_case_hashes: List[str] = []
    remove_case_hashes: List[str] = []


@app.get("/cohorts")
def list_cohorts(db: Session = Depends(get_db)):
    """List all cohorts."""
    cohorts = db.query(Cohort).order_by(Cohort.created_at.desc()).all()
    return [
        {
            "id": c.id,
            "name": c.name,
            "description": c.description,
            "source_type": c.source_type,
            "slide_count": c.slide_count,
            "case_count": c.case_count,
            "auto_add_cases": c.auto_add_cases,
            "created_by": c.created_by,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None
        }
        for c in cohorts
    ]


@app.post("/cohorts")
def create_cohort(cohort_data: CohortCreate, db: Session = Depends(get_db)):
    """Create an empty cohort."""
    with get_lock().write_lock():
        cohort = Cohort(
            name=cohort_data.name,
            description=cohort_data.description,
            source_type='manual',
            created_by=cohort_data.created_by,
            auto_add_cases=cohort_data.auto_add_cases,
        )
        db.add(cohort)
        db.commit()

    return {
        "id": cohort.id,
        "name": cohort.name,
        "slide_count": 0,
        "auto_add_cases": cohort.auto_add_cases,
    }


@app.patch("/cohorts/{cohort_id}")
def update_cohort(cohort_id: int, data: CohortUpdate, db: Session = Depends(get_db)):
    """Update a cohort's name/description or its auto-add-cases setting.

    Turning auto_add_cases on immediately back-fills any existing sibling slides
    of the cohort's cases so the user doesn't have to add them by hand.
    """
    cohort = db.query(Cohort).filter_by(id=cohort_id).first()
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")
    turned_on = False
    with get_lock().write_lock():
        if data.name is not None:
            cohort.name = data.name
        if data.description is not None:
            cohort.description = data.description
        if data.auto_add_cases is not None:
            turned_on = data.auto_add_cases and not cohort.auto_add_cases
            cohort.auto_add_cases = data.auto_add_cases
        db.commit()

    added = reconcile_auto_cohorts(db, cohort_id=cohort_id) if turned_on else 0
    return {
        "id": cohort.id,
        "name": cohort.name,
        "description": cohort.description,
        "auto_add_cases": cohort.auto_add_cases,
        "slides_added": added,
    }


@app.get("/cohorts/{cohort_id}")
def get_cohort(cohort_id: int, db: Session = Depends(get_db)):
    """Get cohort details including all slides with enriched data."""
    cohort = db.query(Cohort).options(
        joinedload(Cohort.slides).joinedload(Slide.case),
        joinedload(Cohort.slides).joinedload(Slide.tags),
    ).filter_by(id=cohort_id).first()
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")

    slides_data = []
    for s in cohort.slides:
        accession_number = None
        slide_number = None
        filepath = indexer.get_filepath(s.slide_hash) if indexer else None
        if filepath:
            parsed = indexer.parser.parse(filepath.name)
            if parsed:
                accession_number = parsed.accession
                slide_number = parsed.slide_number

        slides_data.append({
            "slide_hash": s.slide_hash,
            "accession_number": accession_number,
            "block_id": s.block_id,
            "slide_number": slide_number,
            "stain_type": s.stain_type,
            "random_id": s.random_id,
            "year": s.case.year if s.case else None,
            "case_hash": s.case.accession_hash if s.case else None,
            "tags": [t.name for t in s.tags],
            "file_size_bytes": s.file_size_bytes,
        })

    return {
        "id": cohort.id,
        "name": cohort.name,
        "description": cohort.description,
        "source_type": cohort.source_type,
        "source_details": cohort.source_details,
        "created_by": cohort.created_by,
        "created_at": cohort.created_at.isoformat() if cohort.created_at else None,
        "updated_at": cohort.updated_at.isoformat() if cohort.updated_at else None,
        "slide_count": cohort.slide_count,
        "case_count": cohort.case_count,
        "auto_add_cases": cohort.auto_add_cases,
        "slides": slides_data
    }


@app.delete("/cohorts/{cohort_id}")
def delete_cohort(cohort_id: int, db: Session = Depends(get_db)):
    """Delete a cohort."""
    cohort = db.query(Cohort).filter_by(id=cohort_id).first()
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")

    with get_lock().write_lock():
        db.delete(cohort)
        db.commit()

    return {"status": "ok"}


@app.post("/cohorts/{cohort_id}/slides")
def add_slides_to_cohort(cohort_id: int, data: CohortAddSlides, db: Session = Depends(get_db)):
    """Add slides to a cohort."""
    cohort = db.query(Cohort).filter_by(id=cohort_id).first()
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")

    # ── READ phase (outside lock — NFS queries can be slow) ──────────────
    requested_hashes = list(dict.fromkeys(data.slide_hashes))  # dedupe, preserve order
    slides_by_hash = {
        s.slide_hash: s
        for s in db.query(Slide).filter(Slide.slide_hash.in_(requested_hashes)).all()
    }
    # Existing membership — one query instead of lazy-loading entire collection
    existing_ids: set[int] = {
        row[0]
        for row in db.execute(
            sa_text("SELECT slide_id FROM cohort_slides WHERE cohort_id = :cid"),
            {"cid": cohort_id},
        ).fetchall()
    }

    to_add = []
    not_found = []
    for h in requested_hashes:
        slide = slides_by_hash.get(h)
        if not slide:
            not_found.append(h)
        elif slide.id not in existing_ids:
            to_add.append(slide)

    # ── WRITE phase (inside lock — fast, no NFS reads) ───────────────────
    added = []
    with get_lock().write_lock():
        for slide in to_add:
            db.execute(
                sa_text(
                    "INSERT OR IGNORE INTO cohort_slides (cohort_id, slide_id) VALUES (:cid, :sid)"
                ),
                {"cid": cohort_id, "sid": slide.id},
            )
            added.append(slide.slide_hash)
        db.commit()

    db.refresh(cohort)
    return {
        "status": "ok",
        "added": len(added),
        "added_hashes": added,
        "not_found": not_found,
        "total_slides": cohort.slide_count,
        "total_cases": cohort.case_count
    }


@app.delete("/cohorts/{cohort_id}/slides")
def remove_slides_from_cohort(cohort_id: int, data: CohortAddSlides, db: Session = Depends(get_db)):
    """Remove slides from a cohort."""
    cohort = db.query(Cohort).filter_by(id=cohort_id).first()
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")

    # ── READ phase (outside lock) ─────────────────────────────────────────
    requested_hashes = list(data.slide_hashes)
    slides_by_hash = {
        s.slide_hash: s
        for s in db.query(Slide).filter(Slide.slide_hash.in_(requested_hashes)).all()
    }

    # ── WRITE phase (inside lock) ─────────────────────────────────────────
    removed = []
    with get_lock().write_lock():
        for slide_hash in requested_hashes:
            slide = slides_by_hash.get(slide_hash)
            if slide:
                result = db.execute(
                    sa_text("DELETE FROM cohort_slides WHERE cohort_id = :cid AND slide_id = :sid"),
                    {"cid": cohort_id, "sid": slide.id},
                )
                if result.rowcount > 0:
                    removed.append(slide_hash)
        db.commit()

    db.refresh(cohort)
    return {
        "status": "ok",
        "removed": len(removed),
        "removed_hashes": removed,
        "total_slides": cohort.slide_count,
        "total_cases": cohort.case_count
    }


class _ZipStreamWriter:
    """File-like object that feeds zip data to a queue in ~1 MB chunks."""

    def __init__(self, q: queue.Queue, chunk_size: int = 1024 * 1024):
        self._q = q
        self._chunk_size = chunk_size
        self._buf = bytearray()
        self._pos = 0

    def write(self, data: bytes) -> int:
        self._buf.extend(data)
        self._pos += len(data)
        while len(self._buf) >= self._chunk_size:
            self._q.put(bytes(self._buf[:self._chunk_size]))
            self._buf = self._buf[self._chunk_size:]
        return len(data)

    def tell(self) -> int:
        return self._pos

    def flush(self):
        if self._buf:
            self._q.put(bytes(self._buf))
            self._buf.clear()

    def close(self):
        self.flush()


# ── Patient tracking ──────────────────────────────────────────────────────

class PatientCreate(BaseModel):
    label: str
    note: Optional[str] = None

class PatientUpdate(BaseModel):
    label: Optional[str] = None
    note: Optional[str] = None

class PatientCaseAssign(BaseModel):
    case_hash: str
    surgery_label: str
    note: Optional[str] = None

class PatientCaseUpdate(BaseModel):
    surgery_label: Optional[str] = None
    note: Optional[str] = None

class PatientReorder(BaseModel):
    # Patient ids in the desired display order (first = top).
    patient_ids: list[int]


def _enrich_surgery(
    surgery: "CohortPatientCase",
    cohort_slide_ids: Optional[set] = None,
) -> dict:
    """Build the serialised surgery dict, resolving accession number from indexer cache.

    If cohort_slide_ids is provided, slide_count is restricted to slides that
    are members of that cohort (rather than every slide on the case).
    """
    case = surgery.case
    accession_number = None
    if indexer:
        for slide in case.slides:
            fp = indexer.get_filepath(slide.slide_hash)
            if fp:
                parsed = indexer.parser.parse(fp.name)
                if parsed:
                    accession_number = parsed.accession
                    break
    if cohort_slide_ids is not None:
        slide_count = sum(1 for s in case.slides if s.id in cohort_slide_ids)
    else:
        slide_count = len(case.slides)
    return {
        "id": surgery.id,
        "surgery_label": surgery.surgery_label,
        "case_hash": case.accession_hash,
        "case_id": case.slidecap_id,
        "accession_number": accession_number,
        "year": case.year,
        "slide_count": slide_count,
        "note": surgery.note,
    }


@app.get("/cohorts/{cohort_id}/patients")
def list_cohort_patients(cohort_id: int, db: Session = Depends(get_db)):
    """List all patients (and their surgery assignments) for a cohort."""
    cohort = (
        db.query(Cohort)
        .options(joinedload(Cohort.slides))
        .filter_by(id=cohort_id)
        .first()
    )
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")

    cohort_slide_ids = {s.id for s in cohort.slides}

    patients = (
        db.query(CohortPatient)
        .options(joinedload(CohortPatient.surgeries).joinedload(CohortPatientCase.case).joinedload(Case.slides))
        .filter_by(cohort_id=cohort_id)
        .order_by(CohortPatient.display_order, CohortPatient.label)
        .all()
    )
    return [
        {
            "id": p.id,
            "label": p.label,
            "note": p.note,
            "display_order": p.display_order,
            "surgeries": [_enrich_surgery(s, cohort_slide_ids) for s in p.surgeries],
        }
        for p in patients
    ]


@app.post("/cohorts/{cohort_id}/patients")
def create_cohort_patient(cohort_id: int, data: PatientCreate, db: Session = Depends(get_db)):
    """Create a new patient in a cohort."""
    if not db.query(Cohort).filter_by(id=cohort_id).first():
        raise HTTPException(status_code=404, detail="Cohort not found")
    with get_lock().write_lock():
        # Append new patients at the end of the manual order.
        max_order = (
            db.query(func.max(CohortPatient.display_order))
            .filter_by(cohort_id=cohort_id)
            .scalar()
        )
        next_order = (max_order if max_order is not None else -1) + 1
        patient = CohortPatient(cohort_id=cohort_id, label=data.label, note=data.note, display_order=next_order)
        db.add(patient)
        db.commit()
        db.refresh(patient)
    return {"id": patient.id, "label": patient.label, "note": patient.note,
            "display_order": patient.display_order, "surgeries": []}


@app.patch("/cohorts/{cohort_id}/patients/reorder")
def reorder_cohort_patients(cohort_id: int, data: PatientReorder, db: Session = Depends(get_db)):
    """Set the manual display order of patients in a cohort.

    Body lists patient ids in the desired order; display_order is assigned by
    position. Registered before /patients/{patient_id} so "reorder" isn't parsed
    as a patient id.
    """
    if not db.query(Cohort).filter_by(id=cohort_id).first():
        raise HTTPException(status_code=404, detail="Cohort not found")
    with get_lock().write_lock():
        patients = db.query(CohortPatient).filter_by(cohort_id=cohort_id).all()
        by_id = {p.id: p for p in patients}
        order = 0
        for pid in data.patient_ids:
            p = by_id.get(pid)
            if p is not None:
                p.display_order = order
                order += 1
        # Any patient not named in the list keeps a stable spot at the end.
        for p in patients:
            if p.id not in set(data.patient_ids):
                p.display_order = order
                order += 1
        db.commit()
    return {"status": "ok"}


@app.patch("/cohorts/{cohort_id}/patients/{patient_id}")
def update_cohort_patient(cohort_id: int, patient_id: int, data: PatientUpdate, db: Session = Depends(get_db)):
    """Rename a patient or update their note."""
    patient = db.query(CohortPatient).filter_by(id=patient_id, cohort_id=cohort_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    with get_lock().write_lock():
        if data.label is not None:
            patient.label = data.label
        if data.note is not None:
            patient.note = data.note
        db.commit()
    return {"id": patient.id, "label": patient.label, "note": patient.note}


@app.delete("/cohorts/{cohort_id}/patients/{patient_id}")
def delete_cohort_patient(cohort_id: int, patient_id: int, db: Session = Depends(get_db)):
    """Delete a patient and all their surgery assignments."""
    patient = db.query(CohortPatient).filter_by(id=patient_id, cohort_id=cohort_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    with get_lock().write_lock():
        db.delete(patient)
        db.commit()
    return {"status": "ok"}


@app.post("/cohorts/{cohort_id}/patients/{patient_id}/cases")
def assign_patient_case(cohort_id: int, patient_id: int, data: PatientCaseAssign, db: Session = Depends(get_db)):
    """Assign a case (surgery) to a patient. Moves it if already assigned to another patient."""
    patient = db.query(CohortPatient).filter_by(id=patient_id, cohort_id=cohort_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    case = db.query(Case).filter_by(accession_hash=data.case_hash).first()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    with get_lock().write_lock():
        # Remove any existing assignment for this case within the same cohort
        existing = (
            db.query(CohortPatientCase)
            .join(CohortPatient)
            .filter(CohortPatient.cohort_id == cohort_id, CohortPatientCase.case_id == case.id)
            .first()
        )
        if existing:
            if existing.patient_id == patient_id:
                existing.surgery_label = data.surgery_label
                if data.note is not None:
                    existing.note = data.note
                db.commit()
                return {"status": "ok", "surgery_label": data.surgery_label}
            db.delete(existing)

        assignment = CohortPatientCase(
            patient_id=patient_id,
            case_id=case.id,
            surgery_label=data.surgery_label,
            note=data.note,
        )
        db.add(assignment)
        db.commit()
    return {"status": "ok", "surgery_label": data.surgery_label}


@app.patch("/cohorts/{cohort_id}/patients/{patient_id}/cases/{case_hash}")
def update_patient_case(cohort_id: int, patient_id: int, case_hash: str, data: PatientCaseUpdate, db: Session = Depends(get_db)):
    """Update surgery label or note for a case assignment."""
    patient = db.query(CohortPatient).filter_by(id=patient_id, cohort_id=cohort_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    case = db.query(Case).filter_by(accession_hash=case_hash).first()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    assignment = db.query(CohortPatientCase).filter_by(patient_id=patient_id, case_id=case.id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Case not assigned to this patient")
    with get_lock().write_lock():
        if data.surgery_label is not None:
            assignment.surgery_label = data.surgery_label
        if data.note is not None:
            assignment.note = data.note
        db.commit()
    return {"status": "ok"}


@app.delete("/cohorts/{cohort_id}/patients/{patient_id}/cases/{case_hash}")
def remove_patient_case(cohort_id: int, patient_id: int, case_hash: str, db: Session = Depends(get_db)):
    """Remove a case from a patient's surgery list."""
    patient = db.query(CohortPatient).filter_by(id=patient_id, cohort_id=cohort_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    case = db.query(Case).filter_by(accession_hash=case_hash).first()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    assignment = db.query(CohortPatientCase).filter_by(patient_id=patient_id, case_id=case.id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    with get_lock().write_lock():
        db.delete(assignment)
        db.commit()
    return {"status": "ok"}


@app.get("/cohorts/{cohort_id}/export.csv")
def export_cohort_csv(cohort_id: int, db: Session = Depends(get_db)):
    """Export cohort slides as CSV, including patient_label and surgery_label columns."""
    cohort = db.query(Cohort).options(
        joinedload(Cohort.slides).joinedload(Slide.case).joinedload(Case.tags),
        joinedload(Cohort.slides).joinedload(Slide.tags),
        joinedload(Cohort.patients).joinedload(CohortPatient.surgeries).joinedload(CohortPatientCase.case),
    ).filter_by(id=cohort_id).first()
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")

    # Build case_hash → (patient_label, surgery_label)
    patient_map: dict[str, tuple[str, str]] = {}
    for patient in cohort.patients:
        for surgery in patient.surgeries:
            patient_map[surgery.case.accession_hash] = (patient.label, surgery.surgery_label)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "slide_hash", "filename", "accession", "year", "block_id",
        "slide_number", "stain_type", "random_id", "file_size_bytes",
        "slide_tags", "case_tags", "patient_label", "surgery_label",
    ])
    for slide in cohort.slides:
        filepath = indexer.get_filepath(slide.slide_hash) if indexer else None
        filename = filepath.name if filepath else ""
        if filepath and indexer:
            parsed = indexer.parser.parse(filepath.name)
            accession = parsed.accession if parsed else ""
            slide_number = parsed.slide_number if parsed else ""
        else:
            accession = ""
            slide_number = ""
        case_hash = slide.case.accession_hash if slide.case else ""
        patient_label, surgery_label = patient_map.get(case_hash, ("", ""))
        writer.writerow([
            slide.slide_hash,
            filename,
            accession,
            slide.case.year if slide.case else "",
            slide.block_id or "",
            slide_number,
            slide.stain_type or "",
            slide.random_id or "",
            slide.file_size_bytes or "",
            ";".join(t.name for t in slide.tags),
            ";".join(t.name for t in slide.case.tags) if slide.case else "",
            patient_label,
            surgery_label,
        ])
    safe_name = re.sub(r'[^\w\s-]', '', cohort.name).strip().replace(' ', '_')
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_cohort.csv"'},
    )


@app.get("/cohorts/{cohort_id}/export")
def export_cohort(cohort_id: int, db: Session = Depends(get_db)):
    """Stream cohort slides as a ZIP file, organized by accession number."""
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    cohort = db.query(Cohort).options(
        joinedload(Cohort.slides).joinedload(Slide.case),
    ).filter_by(id=cohort_id).first()
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")

    # Gather (filesystem_path, archive_name) pairs, skip missing files
    slides_info: list[tuple[str, str]] = []
    for s in cohort.slides:
        filepath = indexer.get_filepath(s.slide_hash)
        if not filepath or not filepath.exists():
            continue
        parsed = indexer.parser.parse(filepath.name)
        folder = parsed.accession if parsed else s.slide_hash[:12]
        arcname = f"{folder}/{filepath.name}"
        slides_info.append((str(filepath), arcname))

    if not slides_info:
        raise HTTPException(status_code=404, detail="No slide files available for export")

    def generate():
        q: queue.Queue = queue.Queue(maxsize=32)

        def writer():
            try:
                stream = _ZipStreamWriter(q)
                with zipfile.ZipFile(stream, 'w', zipfile.ZIP_STORED, allowZip64=True) as zf:
                    for filepath, arcname in slides_info:
                        zf.write(filepath, arcname)
                stream.flush()
            except Exception as e:
                print(f"ZIP export error: {e}")
            finally:
                q.put(None)  # sentinel

        t = threading.Thread(target=writer, daemon=True)
        t.start()

        while True:
            chunk = q.get()
            if chunk is None:
                break
            yield chunk

        t.join(timeout=5)

    # Sanitize cohort name for filename
    safe_name = re.sub(r'[^\w\s-]', '', cohort.name).strip().replace(' ', '_') or 'cohort'
    return StreamingResponse(
        generate(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}.zip"',
        },
    )


class SlidePullRequest(BaseModel):
    slide_hashes: List[str]


class SlidePullExportRequest(BaseModel):
    slide_hashes: List[str]
    output_dir: str
    bin_size: int = 100
    bin_prefix: str = "pull"
    # 'hardlink' (default, looks like a real file, fastest, same-volume only)
    # 'copy' (slowest but always works)
    # 'symlink' (instant but downstream tools like Halo may not follow)
    # 'auto' (legacy: symlink → hardlink → copy probe)
    method: str = "hardlink"
    # When True (default), slides already materialized in their target bin
    # (same filename, matching size) are left in place and reported as
    # "skipped" rather than re-linked/re-copied. This makes a re-run of the
    # same list resumable: an interrupted pull continues instead of redoing
    # all the work. Set False to force every slide to be re-materialized.
    skip_existing: bool = True


def _resolve_link_method(output_dir: Path, sample_src: Path, requested: str) -> str:
    """Pick the link method to use as the *preferred* method for this export.

    - 'hardlink' / 'symlink' / 'copy' — honor user request directly (the per-file
      _link_or_copy fallback chain still degrades gracefully if it fails partway).
    - 'auto' — legacy probe: try symlink, then hardlink, then copy.
    """
    if requested in ("hardlink", "symlink", "copy"):
        return requested

    probe_dir = output_dir / ".slidecap-probe"
    try:
        probe_dir.mkdir(exist_ok=True)
    except OSError:
        return "copy"
    try:
        sym_probe = probe_dir / f"sym-{os.getpid()}"
        try:
            if sym_probe.is_symlink() or sym_probe.exists():
                sym_probe.unlink()
            os.symlink(str(sample_src), str(sym_probe))
            try:
                sym_probe.unlink()
            except OSError:
                pass
            return "symlink"
        except (OSError, NotImplementedError):
            pass

        link_probe = probe_dir / f"hl-{os.getpid()}"
        try:
            if link_probe.exists():
                link_probe.unlink()
            os.link(str(sample_src), str(link_probe))
            try:
                link_probe.unlink()
            except OSError:
                pass
            return "hardlink"
        except (OSError, NotImplementedError):
            pass

        return "copy"
    finally:
        try:
            shutil.rmtree(probe_dir, ignore_errors=True)
        except Exception:
            pass


def _link_or_copy(src: Path, dst: Path, preferred: str, skip_existing: bool = False) -> str:
    """Materialize src at dst using preferred method, falling back to weaker methods. Returns method actually used.

    If skip_existing is True and dst is already a regular file whose size matches src,
    it is left untouched and "skipped" is returned (resumable re-runs). Symlinks and
    size-mismatched leftovers (e.g. a partial copy from an interrupted run) are always
    re-materialized so the target ends up correct.
    """
    if skip_existing and dst.exists() and not dst.is_symlink():
        try:
            if dst.stat().st_size == src.stat().st_size:
                return "skipped"
        except OSError:
            pass  # fall through and re-materialize

    if dst.is_symlink() or dst.exists():
        dst.unlink()

    if preferred == "symlink":
        chain = ["symlink", "hardlink", "copy"]
    elif preferred == "hardlink":
        chain = ["hardlink", "copy"]
    else:
        chain = ["copy"]

    last_err: Optional[Exception] = None
    for method in chain:
        try:
            if method == "symlink":
                os.symlink(str(src), str(dst))
            elif method == "hardlink":
                os.link(str(src), str(dst))
            else:
                shutil.copy2(str(src), str(dst))
            return method
        except (OSError, NotImplementedError) as e:
            last_err = e
            continue
    raise RuntimeError(f"All link methods failed for {src.name}: {last_err}")


_pull_history_lock = threading.Lock()


def _pull_history_path() -> Path:
    """JSON-lines log of every Export-to-Directory pull, one entry per line."""
    return settings.local_data_path / "pull-history.jsonl"


def _record_pull_history(entry: dict) -> None:
    """Append a pull record. Best-effort: never let history logging break an export."""
    try:
        with _pull_history_lock:
            with open(_pull_history_path(), "a", encoding="utf-8") as f:
                f.write(json.dumps(entry) + "\n")
    except Exception as e:
        print(f"[pull-history] failed to record entry: {e}")


@app.get("/slides/pull-history")
def pull_history(limit: int = 50):
    """Return past Export-to-Directory pulls, newest first."""
    path = _pull_history_path()
    if not path.exists():
        return {"entries": []}
    entries: list[dict] = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    entries.append(json.loads(line))
                except json.JSONDecodeError:
                    continue  # skip a corrupt line rather than failing the whole read
    except OSError as e:
        raise HTTPException(status_code=500, detail=f"Cannot read pull history: {e}")
    entries.reverse()  # file is append-order (oldest first) → newest first for the UI
    return {"entries": entries[: max(1, limit)]}


@app.post("/slides/pull-export")
def pull_export(data: SlidePullExportRequest):
    """Materialize selected slides under output_dir as pull-NNN/ subdirs, capped at bin_size slides each.
    Uses symlinks where supported, falls back to hardlinks, then copies. Writes a manifest per bin
    and a top-level pull-summary.csv recording method, missing slides, and any failures."""
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")
    if not data.slide_hashes:
        raise HTTPException(status_code=400, detail="No slides requested")

    bin_size = max(1, data.bin_size)
    bin_prefix = re.sub(r"[^\w\-]", "", data.bin_prefix or "pull") or "pull"

    try:
        output_root = Path(data.output_dir).expanduser().resolve()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid output_dir: {e}")

    try:
        output_root.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        raise HTTPException(status_code=400, detail=f"Cannot create output_dir: {e}")
    if not output_root.is_dir():
        raise HTTPException(status_code=400, detail=f"output_dir is not a directory: {output_root}")

    # Resolve slide hashes to filepaths
    resolved: list[dict] = []
    missing_hashes: list[str] = []
    for slide_hash in data.slide_hashes:
        filepath = indexer.get_filepath(slide_hash)
        if not filepath or not filepath.exists():
            missing_hashes.append(slide_hash)
            continue
        parsed = indexer.parser.parse(filepath.name)
        accession = parsed.accession if parsed else slide_hash[:12]
        resolved.append({
            "slide_hash": slide_hash,
            "src_path": filepath,
            "accession": accession,
            "filename": filepath.name,
        })

    if not resolved:
        raise HTTPException(status_code=404, detail="No slide files available on disk for the given hashes")

    method_request = (data.method or "hardlink").lower()
    if method_request not in ("auto", "hardlink", "symlink", "copy"):
        raise HTTPException(status_code=400, detail=f"Invalid method: {data.method}")
    preferred_method = _resolve_link_method(output_root, resolved[0]["src_path"], method_request)

    failures: list[dict] = []
    bin_reports: list[dict] = []

    for bin_idx in range(0, len(resolved), bin_size):
        chunk = resolved[bin_idx:bin_idx + bin_size]
        bin_num = (bin_idx // bin_size) + 1
        bin_dir = output_root / f"{bin_prefix}-{bin_num:03d}"
        bin_dir.mkdir(parents=True, exist_ok=True)

        manifest_rows: list[dict] = []
        bin_failures = 0
        bin_skipped = 0
        for s in chunk:
            dst = bin_dir / s["filename"]
            try:
                used = _link_or_copy(s["src_path"], dst, preferred_method, skip_existing=data.skip_existing)
                if used == "skipped":
                    bin_skipped += 1
                manifest_rows.append({
                    "slide_hash": s["slide_hash"],
                    "accession": s["accession"],
                    "filename": s["filename"],
                    "src_path": str(s["src_path"]),
                    "method": used,
                })
            except Exception as e:
                bin_failures += 1
                failures.append({
                    "slide_hash": s["slide_hash"],
                    "filename": s["filename"],
                    "bin": bin_dir.name,
                    "error": str(e),
                })

        with open(bin_dir / "manifest.csv", "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["slide_hash", "accession", "filename", "src_path", "method"])
            writer.writeheader()
            writer.writerows(manifest_rows)

        bin_reports.append({
            "bin": bin_dir.name,
            "path": str(bin_dir),
            "slides": len(manifest_rows),
            "skipped": bin_skipped,
            "failures": bin_failures,
        })

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_exported = sum(b["slides"] for b in bin_reports)
    total_skipped = sum(b["skipped"] for b in bin_reports)
    with open(output_root / "pull-summary.csv", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["exported_at", timestamp])
        w.writerow(["preferred_method", preferred_method])
        w.writerow(["skip_existing", data.skip_existing])
        w.writerow(["bin_size", bin_size])
        w.writerow(["total_requested", len(data.slide_hashes)])
        w.writerow(["total_exported", total_exported])
        w.writerow(["total_skipped", total_skipped])
        w.writerow(["missing_count", len(missing_hashes)])
        w.writerow(["failure_count", len(failures)])
        w.writerow([])
        w.writerow(["bin", "path", "slides", "skipped", "failures"])
        for b in bin_reports:
            w.writerow([b["bin"], b["path"], b["slides"], b["skipped"], b["failures"]])
        if missing_hashes:
            w.writerow([])
            w.writerow(["missing_slide_hashes"])
            for h in missing_hashes:
                w.writerow([h])
        if failures:
            w.writerow([])
            w.writerow(["failure_slide_hash", "filename", "bin", "error"])
            for fl in failures:
                w.writerow([fl["slide_hash"], fl["filename"], fl["bin"], fl["error"]])

    # Distinct accessions touched by this pull (for the "associated cases" history view).
    accessions = sorted({s["accession"] for s in resolved})

    result = {
        "output_dir": str(output_root),
        "preferred_method": preferred_method,
        "skip_existing": data.skip_existing,
        "bin_size": bin_size,
        "bin_count": len(bin_reports),
        "total_requested": len(data.slide_hashes),
        "total_exported": total_exported,
        "total_skipped": total_skipped,
        "missing_count": len(missing_hashes),
        "failure_count": len(failures),
        "bins": bin_reports,
        "failures": failures[:25],
        "missing_hashes": missing_hashes[:25],
        "summary_path": str(output_root / "pull-summary.csv"),
    }

    _record_pull_history({
        "id": datetime.now().strftime("%Y%m%d-%H%M%S-") + f"{os.getpid():d}",
        "exported_at": timestamp,
        "output_dir": str(output_root),
        "preferred_method": preferred_method,
        "method_requested": method_request,
        "skip_existing": data.skip_existing,
        "bin_size": bin_size,
        "bin_count": len(bin_reports),
        "total_requested": len(data.slide_hashes),
        "total_exported": total_exported,
        "total_skipped": total_skipped,
        "missing_count": len(missing_hashes),
        "failure_count": len(failures),
        "case_count": len(accessions),
        "accessions": accessions,
        "bins": [{"bin": b["bin"], "slides": b["slides"], "skipped": b["skipped"], "failures": b["failures"]} for b in bin_reports],
    })

    return result


@app.post("/slides/pull-download")
def pull_download(data: SlidePullRequest):
    """Stream a ZIP of selected slides, organized by accession number."""
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    if not data.slide_hashes:
        raise HTTPException(status_code=400, detail="No slides requested")

    # Gather (filesystem_path, archive_name) pairs, skip missing files
    slides_info: list[tuple[str, str]] = []
    for slide_hash in data.slide_hashes:
        filepath = indexer.get_filepath(slide_hash)
        if not filepath or not filepath.exists():
            continue
        parsed = indexer.parser.parse(filepath.name)
        folder = parsed.accession if parsed else slide_hash[:12]
        arcname = f"{folder}/{filepath.name}"
        slides_info.append((str(filepath), arcname))

    if not slides_info:
        raise HTTPException(status_code=404, detail="No slide files available for download")

    def generate():
        q: queue.Queue = queue.Queue(maxsize=32)

        def writer():
            try:
                stream = _ZipStreamWriter(q)
                with zipfile.ZipFile(stream, 'w', zipfile.ZIP_STORED, allowZip64=True) as zf:
                    for filepath, arcname in slides_info:
                        zf.write(filepath, arcname)
                stream.flush()
            except Exception as e:
                print(f"WSI Pull ZIP error: {e}")
            finally:
                q.put(None)

        t = threading.Thread(target=writer, daemon=True)
        t.start()

        while True:
            chunk = q.get()
            if chunk is None:
                break
            yield chunk

        t.join(timeout=5)

    timestamp = datetime.now().strftime("%Y-%m-%d")
    return StreamingResponse(
        generate(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="slide-pull-{timestamp}.zip"',
        },
    )


@app.post("/cohorts/{cohort_id}/export-analyses")
def export_cohort_analyses(
    cohort_id: int,
    analysis_name: Optional[str] = Query(None, description="Filter by analysis name"),
    db: Session = Depends(get_db),
):
    """
    Export post-processed analysis results for a cohort as a ZIP.
    Structure: {accession_number}/{analysis_name}/{files}
    """
    import subprocess
    import tempfile

    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    cohort = db.query(Cohort).options(
        joinedload(Cohort.slides).joinedload(Slide.case),
        joinedload(Cohort.slides).joinedload(Slide.job_slides).joinedload(JobSlide.job).joinedload(AnalysisJob.analysis),
    ).filter_by(id=cohort_id).first()
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")

    # Collect (output_dir, archive_prefix, postprocess_template) for each completed slide analysis
    export_items: list[tuple[Path, str, str | None]] = []

    for slide in cohort.slides:
        # Resolve accession number for folder name
        filepath = indexer.get_filepath(slide.slide_hash)
        if filepath:
            parsed = indexer.parser.parse(filepath.name)
            folder = parsed.accession if parsed else slide.slide_hash[:12]
        else:
            folder = slide.slide_hash[:12]

        for js in slide.job_slides:
            if js.status != "completed":
                continue
            if not js.local_output_path:
                continue
            local_out = Path(js.local_output_path)
            if not local_out.is_dir():
                continue

            job = js.job
            if not job:
                continue

            a_name = job.model_name
            if analysis_name and a_name != analysis_name:
                continue

            pp_template = job.analysis.postprocess_template if job.analysis else None
            arc_prefix = f"{folder}/{a_name}"
            export_items.append((local_out, arc_prefix, pp_template))

    if not export_items:
        raise HTTPException(status_code=404, detail="No completed analysis results available for export")

    def generate():
        q: queue.Queue = queue.Queue(maxsize=32)

        def writer():
            try:
                stream = _ZipStreamWriter(q)
                with zipfile.ZipFile(stream, 'w', zipfile.ZIP_DEFLATED, allowZip64=True) as zf:
                    for output_dir, arc_prefix, pp_template in export_items:
                        if pp_template:
                            # Run postprocess into a temp dir
                            with tempfile.TemporaryDirectory() as tmpdir:
                                cmd = pp_template.format(
                                    input_dir=str(output_dir),
                                    output_dir=tmpdir,
                                    filename_stem=output_dir.name,
                                )
                                try:
                                    subprocess.run(cmd, shell=True, check=True, timeout=300)
                                    src_dir = Path(tmpdir)
                                except Exception as e:
                                    print(f"Postprocess failed for {arc_prefix}: {e}, using raw files")
                                    src_dir = output_dir

                                for f in sorted(src_dir.iterdir()):
                                    if f.is_file():
                                        zf.write(str(f), f"{arc_prefix}/{f.name}")
                        else:
                            # No postprocessing — copy raw files
                            for f in sorted(output_dir.iterdir()):
                                if f.is_file():
                                    zf.write(str(f), f"{arc_prefix}/{f.name}")
                stream.flush()
            except Exception as e:
                print(f"Analysis export error: {e}")
            finally:
                q.put(None)

        t = threading.Thread(target=writer, daemon=True)
        t.start()

        while True:
            chunk = q.get()
            if chunk is None:
                break
            yield chunk

        t.join(timeout=5)

    safe_name = re.sub(r'[^\w\s-]', '', cohort.name).strip().replace(' ', '_') or 'cohort'
    return StreamingResponse(
        generate(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}_analyses.zip"',
        },
    )


@app.post("/cohorts/from-tag")
def create_cohort_from_tag(data: CohortFromTag, db: Session = Depends(get_db)):
    """Create a cohort from all slides with a specific tag."""
    tag = db.query(Tag).filter_by(name=data.tag_name).first()
    if not tag:
        raise HTTPException(status_code=404, detail="Tag not found")

    with get_lock().write_lock():
        cohort = Cohort(
            name=data.name,
            description=data.description,
            source_type='tag',
            source_details=f'{{"tag": "{data.tag_name}"}}',
            created_by=data.created_by
        )
        cohort.slides = list(tag.slides)
        db.add(cohort)
        db.commit()

    return {
        "id": cohort.id,
        "name": cohort.name,
        "slide_count": cohort.slide_count,
        "case_count": cohort.case_count
    }


@app.post("/cohorts/from-file")
async def create_cohort_from_file(
    name: str,
    file: UploadFile = File(...),
    description: Optional[str] = None,
    created_by: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Create a cohort by uploading a file with accession numbers or a manifest.
    Supports .txt (one per line), .csv, .xlsx.

    Single column: matches all slides for each accession number.
    Three columns (accession, block, stain): matches specific slides by all three fields.
    """
    # Read file content (strip BOM if present)
    content = await file.read()
    if content.startswith(b'\xef\xbb\xbf'):
        content = content[3:]
    filename = file.filename or ""

    # Parse rows — each row is a list of column values
    raw_rows = []

    if filename.endswith('.xlsx'):
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(cells):
                    raw_rows.append(cells)
        except ImportError:
            raise HTTPException(status_code=400, detail="openpyxl not installed for Excel support")
    elif filename.endswith('.csv'):
        import csv
        from io import StringIO
        text = content.decode('utf-8')
        reader = csv.reader(StringIO(text))
        for row in reader:
            cells = [c.strip() for c in row]
            if any(cells):
                raw_rows.append(cells)
    else:
        # Text file — split each line by tab, comma, or whitespace
        import re as _re
        text = content.decode('utf-8')
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            if '\t' in line:
                cells = [c.strip() for c in line.split('\t')]
            elif ',' in line:
                cells = [c.strip() for c in line.split(',')]
            else:
                # Split by whitespace (handles multiple spaces)
                cells = _re.split(r'\s+', line)
            raw_rows.append(cells)

    if not raw_rows:
        raise HTTPException(status_code=400, detail="No data found in file")

    # Detect mode: manifest (3 cols) vs accession list (1 col)
    is_manifest = len(raw_rows[0]) >= 3 and all(len(r) >= 3 for r in raw_rows)

    def normalize_accession(acc: str) -> str:
        """Normalize BS-?YY- prefix so BS18- and BS-18- both match."""
        return _normalize_accession(acc)

    matching_slides = []
    matching_slide_ids = set()  # track by id to avoid duplicates

    if is_manifest:
        # Manifest mode: each row is (accession, block, stain)
        manifest_rows = [(r[0], r[1], r[2]) for r in raw_rows]
        rows_matched = []
        rows_not_matched = []

        for acc, block, stain in manifest_rows:
            acc_norm = normalize_accession(acc)
            block_upper = block.upper()
            stain_lower = stain.lower()
            found = False
            near_misses = []

            for slide_hash, filepath in indexer.slide_hash_to_path.items():
                parsed = indexer.parser.parse(filepath.name)
                if not parsed:
                    continue
                if normalize_accession(parsed.accession) == acc_norm:
                    # Accession matched — check block + stain
                    if (parsed.block_id.upper() == block_upper
                            and parsed.stain_type.lower() == stain_lower):
                        slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
                        if slide and slide.id not in matching_slide_ids:
                            matching_slides.append(slide)
                            matching_slide_ids.add(slide.id)
                        found = True
                    else:
                        near_misses.append(f"block={parsed.block_id},stain={parsed.stain_type}")

            if found:
                rows_matched.append(f"{acc},{block},{stain}")
            else:
                detail = f"{acc},{block},{stain}"
                if near_misses:
                    detail += f" (accession found but slides have: {'; '.join(near_misses[:5])})"
                rows_not_matched.append(detail)

        # Create cohort
        with get_lock().write_lock():
            cohort = Cohort(
                name=name,
                description=description,
                source_type='upload',
                source_details=json.dumps({
                    "filename": filename,
                    "mode": "manifest",
                    "rows_requested": len(manifest_rows),
                    "rows_matched": len(rows_matched),
                    "rows_not_matched": rows_not_matched[:50]
                }),
                created_by=created_by
            )
            cohort.slides = matching_slides
            db.add(cohort)
            db.commit()

        return {
            "id": cohort.id,
            "name": cohort.name,
            "slide_count": cohort.slide_count,
            "case_count": cohort.case_count,
            "rows_requested": len(manifest_rows),
            "rows_matched": len(rows_matched),
            "rows_not_matched": rows_not_matched
        }
    else:
        # Accession list mode (original behavior)
        accessions = [r[0] for r in raw_rows if r[0]]

        if not accessions:
            raise HTTPException(status_code=400, detail="No accession numbers found in file")

        found_accessions = set()
        not_found_accessions = []

        for accession in accessions:
            acc_norm = normalize_accession(accession)
            found = False

            for slide_hash, filepath in indexer.slide_hash_to_path.items():
                parsed = indexer.parser.parse(filepath.name)
                if parsed and normalize_accession(parsed.accession) == acc_norm:
                    slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
                    if slide and slide.id not in matching_slide_ids:
                        matching_slides.append(slide)
                        matching_slide_ids.add(slide.id)
                        found_accessions.add(acc_norm)
                        found = True

            if not found and acc_norm not in found_accessions:
                not_found_accessions.append(accession)

        # Create cohort
        with get_lock().write_lock():
            cohort = Cohort(
                name=name,
                description=description,
                source_type='upload',
                source_details=json.dumps({
                    "filename": filename,
                    "mode": "accession_list",
                    "accessions_requested": len(accessions),
                    "accessions_found": len(found_accessions),
                    "accessions_not_found": not_found_accessions[:50]
                }),
                created_by=created_by
            )
            cohort.slides = matching_slides
            db.add(cohort)
            db.commit()

        return {
            "id": cohort.id,
            "name": cohort.name,
            "slide_count": cohort.slide_count,
            "case_count": cohort.case_count,
            "accessions_requested": len(accessions),
            "accessions_found": len(found_accessions),
            "accessions_not_found": not_found_accessions
        }


@app.post("/cohorts/from-filter")
def create_cohort_from_filter(
    db: Session = Depends(get_db),
    name: str = Query(...),
    description: Optional[str] = None,
    created_by: Optional[str] = None,
    query: Optional[str] = None,
    year: Optional[int] = None,
    stain: Optional[str] = None,
    tag: Optional[str] = None
):
    """Create a cohort from slides matching filter criteria."""
    # Use the search function to find matching slides
    results = indexer.search(
        db=db,
        query=query or "",
        year=year,
        stain_type=stain,
        tags=[tag] if tag else None,
        limit=10000  # Higher limit for cohort building
    )

    if not results:
        raise HTTPException(status_code=400, detail="No slides match the filter criteria")

    # Get slide objects
    slide_hashes = [r['slide_hash'] for r in results]
    slides = db.query(Slide).filter(Slide.slide_hash.in_(slide_hashes)).all()

    # Create cohort
    with get_lock().write_lock():
        cohort = Cohort(
            name=name,
            description=description,
            source_type='filter',
            source_details=json.dumps({
                "query": query,
                "year": year,
                "stain": stain,
                "tag": tag
            }),
            created_by=created_by
        )
        cohort.slides = slides
        db.add(cohort)
        db.commit()

    return {
        "id": cohort.id,
        "name": cohort.name,
        "slide_count": cohort.slide_count,
        "case_count": cohort.case_count
    }


# ============================================================
# Analysis Registry Endpoints
# ============================================================

class AnalysisCreate(BaseModel):
    name: str
    version: str = '1.0'
    description: Optional[str] = None
    kind: str = 'cellvit'  # plugin id from backend/app/analyses/
    script_path: Optional[str] = None
    working_directory: Optional[str] = None
    env_setup: Optional[str] = None
    command_template: Optional[str] = None
    postprocess_template: Optional[str] = None
    parameters_schema: Optional[str] = None
    default_parameters: Optional[str] = None
    gpu_required: bool = True
    estimated_runtime_minutes: int = 60
    transforms: Optional[str] = None  # JSON-encoded list of {match, ops}; if unset, kind's default_rules apply


class AnalysisUpdate(BaseModel):
    name: Optional[str] = None
    version: Optional[str] = None
    description: Optional[str] = None
    kind: Optional[str] = None
    script_path: Optional[str] = None
    working_directory: Optional[str] = None
    env_setup: Optional[str] = None
    command_template: Optional[str] = None
    postprocess_template: Optional[str] = None
    parameters_schema: Optional[str] = None
    default_parameters: Optional[str] = None
    gpu_required: Optional[bool] = None
    estimated_runtime_minutes: Optional[int] = None
    active: Optional[bool] = None
    transforms: Optional[str] = None


@app.get("/analyses/kinds")
def list_analysis_kinds():
    """List registered analysis-kind plugins. Used by the UI's Kind dropdown.

    Registered above /analyses/{analysis_id} so FastAPI matches `kinds` here
    instead of trying to coerce it to an int and 422-ing.
    """
    from . import analyses as _ax
    return _ax.list_kinds()


@app.get("/analyses")
def list_analyses(
    db: Session = Depends(get_db),
    active_only: bool = Query(False, description="Only return active analyses"),
):
    """List all registered analyses."""
    query = db.query(Analysis)
    if active_only:
        query = query.filter(Analysis.active == True)
    analyses = query.order_by(Analysis.name).all()
    return [
        {
            "id": a.id,
            "name": a.name,
            "version": a.version,
            "description": a.description,
            "kind": a.kind,
            "script_path": a.script_path,
            "working_directory": a.working_directory,
            "env_setup": a.env_setup,
            "command_template": a.command_template,
            "postprocess_template": a.postprocess_template,
            "parameters_schema": a.parameters_schema,
            "default_parameters": a.default_parameters,
            "gpu_required": a.gpu_required,
            "estimated_runtime_minutes": a.estimated_runtime_minutes,
            "active": a.active,
            "created_at": a.created_at.isoformat() if a.created_at else None,
            "job_count": len(a.jobs),
            "transforms": a.transforms,
        }
        for a in analyses
    ]


@app.post("/analyses")
def create_analysis(data: AnalysisCreate, db: Session = Depends(get_db)):
    """Register a new analysis pipeline."""
    from . import analyses as _ax
    existing = db.query(Analysis).filter_by(name=data.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Analysis with this name already exists")
    if _ax.get_kind(data.kind) is None:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown analysis kind {data.kind!r}. Use GET /analyses/kinds for valid ids.",
        )

    analysis = Analysis(
        name=data.name,
        version=data.version,
        description=data.description,
        kind=data.kind,
        script_path=data.script_path,
        working_directory=data.working_directory,
        env_setup=data.env_setup,
        command_template=data.command_template,
        postprocess_template=data.postprocess_template,
        parameters_schema=data.parameters_schema,
        default_parameters=data.default_parameters,
        gpu_required=data.gpu_required,
        estimated_runtime_minutes=data.estimated_runtime_minutes,
        transforms=data.transforms,
    )
    db.add(analysis)
    db.commit()

    return {
        "id": analysis.id,
        "name": analysis.name,
        "version": analysis.version,
    }


@app.get("/analyses/{analysis_id}")
def get_analysis(analysis_id: int, db: Session = Depends(get_db)):
    """Get a single analysis by ID."""
    analysis = db.query(Analysis).filter_by(id=analysis_id).first()
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")

    return {
        "id": analysis.id,
        "name": analysis.name,
        "version": analysis.version,
        "description": analysis.description,
        "kind": analysis.kind,
        "script_path": analysis.script_path,
        "working_directory": analysis.working_directory,
        "env_setup": analysis.env_setup,
        "command_template": analysis.command_template,
        "postprocess_template": analysis.postprocess_template,
        "parameters_schema": analysis.parameters_schema,
        "default_parameters": analysis.default_parameters,
        "gpu_required": analysis.gpu_required,
        "estimated_runtime_minutes": analysis.estimated_runtime_minutes,
        "active": analysis.active,
        "created_at": analysis.created_at.isoformat() if analysis.created_at else None,
        "job_count": len(analysis.jobs),
        "transforms": analysis.transforms,
    }


@app.patch("/analyses/{analysis_id}")
def update_analysis(analysis_id: int, data: AnalysisUpdate, db: Session = Depends(get_db)):
    """Update an analysis pipeline."""
    from . import analyses as _ax
    analysis = db.query(Analysis).filter_by(id=analysis_id).first()
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")

    if data.name is not None:
        existing = db.query(Analysis).filter(Analysis.name == data.name, Analysis.id != analysis_id).first()
        if existing:
            raise HTTPException(status_code=400, detail="Analysis name already exists")
        analysis.name = data.name
    if data.version is not None:
        analysis.version = data.version
    if data.description is not None:
        analysis.description = data.description
    if data.kind is not None:
        if _ax.get_kind(data.kind) is None:
            raise HTTPException(
                status_code=400,
                detail=f"Unknown analysis kind {data.kind!r}. Use GET /analyses/kinds for valid ids.",
            )
        analysis.kind = data.kind
    if data.script_path is not None:
        analysis.script_path = data.script_path
    if data.working_directory is not None:
        analysis.working_directory = data.working_directory
    if data.env_setup is not None:
        analysis.env_setup = data.env_setup
    if data.command_template is not None:
        analysis.command_template = data.command_template
    if data.postprocess_template is not None:
        analysis.postprocess_template = data.postprocess_template
    if data.parameters_schema is not None:
        analysis.parameters_schema = data.parameters_schema
    if data.default_parameters is not None:
        analysis.default_parameters = data.default_parameters
    if data.gpu_required is not None:
        analysis.gpu_required = data.gpu_required
    if data.estimated_runtime_minutes is not None:
        analysis.estimated_runtime_minutes = data.estimated_runtime_minutes
    if data.active is not None:
        analysis.active = data.active
    if data.transforms is not None:
        # Validate it's JSON-decodable so bad payloads error early
        try:
            json.loads(data.transforms)
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"transforms must be valid JSON: {e}")
        analysis.transforms = data.transforms

    db.commit()

    return {
        "id": analysis.id,
        "name": analysis.name,
        "version": analysis.version,
        "active": analysis.active,
    }


@app.delete("/analyses/{analysis_id}")
def delete_analysis(analysis_id: int, db: Session = Depends(get_db)):
    """Soft-delete an analysis (sets active=False)."""
    analysis = db.query(Analysis).filter_by(id=analysis_id).first()
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")

    analysis.active = False
    db.commit()

    return {"status": "ok", "id": analysis_id, "active": False}


# ============================================================
# Job Submission & Management Endpoints
# ============================================================

class JobSubmitRequest(BaseModel):
    analysis_id: int
    slide_hashes: List[str]
    gpu_index: int = 0
    remote_wsi_dir: str = "/tmp/slidecap_wsi"
    remote_output_dir: str = "/tmp/slidecap_output"
    parameters: Optional[str] = None
    submitted_by: Optional[str] = None


class CohortJobSubmitRequest(BaseModel):
    analysis_id: int
    gpu_index: int = 0
    remote_wsi_dir: str = "/tmp/slidecap_wsi"
    remote_output_dir: str = "/tmp/slidecap_output"
    parameters: Optional[str] = None
    submitted_by: Optional[str] = None
    case_hashes: Optional[List[str]] = None  # If set, only submit slides from these cases


class JobCancelRequest(BaseModel):
    job_ids: List[int]


def _submit_jobs_demo(data: "JobSubmitRequest", db: Session) -> dict:
    """
    Demo-mode job submission. Creates real DB records but skips all
    cluster work; demo_mod.simulate_job_run walks the job through
    transferring -> running -> completed on a 5s + 5s timer.
    """
    if not _demo_cluster_connected:
        raise HTTPException(status_code=503, detail="Not connected to cluster. Connect first via /cluster/connect")

    analysis = db.query(Analysis).filter_by(id=data.analysis_id, active=True).first()
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found or inactive")

    job = AnalysisJob(
        slidecap_id=generate_slidecap_id(db, "JB"),
        analysis_id=analysis.id,
        model_name=analysis.name,
        model_version=analysis.version,
        parameters=data.parameters,
        gpu_index=data.gpu_index,
        remote_wsi_dir=data.remote_wsi_dir,
        remote_output_dir=data.remote_output_dir,
        output_path=str(settings.local_data_path / "demo-outputs"),
        status="pending",
        submitted_by=data.submitted_by,
    )
    db.add(job)
    db.flush()

    errors: list[str] = []
    # (job_slide_id, slide_label, slide_hash) — hash is passed so the demo
    # output generator can size the cells.geojson to this slide's actual
    # dimensions, putting the overlay on tissue instead of in dead space.
    slide_specs: list[tuple[int, str, str]] = []

    for slide_hash in data.slide_hashes:
        slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
        if not slide:
            errors.append(f"Slide not found: {slide_hash[:12]}")
            continue

        # Build a non-PHI label for the demo output dir name
        label = slide.slidecap_id or slide_hash[:10]
        # Try filename if we have one, but fall back to the slidecap id
        slide_path = indexer.get_filepath(slide_hash) if indexer else None
        filename = slide_path.name if slide_path else f"{label}.svs"

        job_slide = JobSlide(
            job_id=job.id,
            slide_id=slide.id,
            remote_wsi_path=f"{data.remote_wsi_dir}/{job.id}/{filename}",
            remote_output_path=f"{data.remote_output_dir}/{job.id}",
            filename=filename,
            status="pending",
        )
        db.add(job_slide)
        db.flush()
        slide_specs.append((job_slide.id, label, slide_hash))

    db.commit()

    demo_mod.simulate_job_run(job.id, slide_specs)

    return {
        "job_id": job.id,
        "slides_created": len(slide_specs),
        "old_records_cleaned": 0,
        "errors": errors,
        "cluster_connected": True,
    }


@app.post("/jobs/submit")
def submit_jobs(data: JobSubmitRequest, db: Session = Depends(get_db)):
    """
    Submit a multi-slide analysis job.

    Creates DB records immediately (pending), then spawns a background thread
    to do rsync + tmux start. This avoids holding the DB lock during long transfers.
    """
    if _is_demo():
        return _submit_jobs_demo(data, db)

    if not cluster_service or not cluster_service.is_connected:
        raise HTTPException(status_code=503, detail="Not connected to cluster. Connect first via /cluster/connect")

    analysis = db.query(Analysis).filter_by(id=data.analysis_id, active=True).first()
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found or inactive")

    params = None
    if data.parameters:
        try:
            params = json.loads(data.parameters)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid JSON in parameters")

    # Snapshot analysis config for the background thread (detached from session)
    analysis_snapshot = {
        "id": analysis.id,
        "name": analysis.name,
        "version": analysis.version,
        "script_path": analysis.script_path,
        "working_directory": analysis.working_directory,
        "env_setup": analysis.env_setup,
        "command_template": analysis.command_template,
        "gpu_required": analysis.gpu_required,
    }

    # --- Phase 1: Create all DB records and commit (fast, releases DB) ---
    job = AnalysisJob(
        slidecap_id=generate_slidecap_id(db, "JB"),
        analysis_id=analysis.id,
        model_name=analysis.name,
        model_version=analysis.version,
        parameters=data.parameters,
        gpu_index=data.gpu_index,
        remote_wsi_dir=data.remote_wsi_dir,
        remote_output_dir=data.remote_output_dir,
        output_path=str(settings.analyses_path),
        status="pending",
        submitted_by=data.submitted_by,
    )
    db.add(job)
    db.flush()

    errors = []
    slides_created = 0
    old_records_cleaned = 0

    # Check if cluster has the network drive mounted (skip transfers if so)
    cluster_mount = settings.CLUSTER_NETWORK_MOUNT  # e.g. "/ligonlab"
    direct_mount = False
    local_to_cluster_prefix = None
    if cluster_mount:
        # Build path mapping: local NETWORK_ROOT -> cluster mount path
        # e.g. "/Volumes/DFCI-LIGONLAB/Ligon Lab/test_directory_pt_slides" -> "/ligonlab"
        local_to_cluster_prefix = (str(settings.NETWORK_ROOT), cluster_mount.rstrip("/"))
        direct_mount = True
        print(f"[Job] Direct mount mode: {local_to_cluster_prefix[0]} -> {local_to_cluster_prefix[1]}")

    # List of (job_slide_id, slide_hash, local_path, remote_wsi_path) for bg thread
    slides_to_process: list[tuple[int, str, str, str]] = []

    for slide_hash in data.slide_hashes:
        slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
        if not slide:
            errors.append(f"Slide not found: {slide_hash[:12]}")
            continue

        slide_path = indexer.get_filepath(slide_hash) if indexer else None
        if not slide_path:
            errors.append(f"No filepath for {slide_hash[:12]}")
            continue

        # Clean up old completed/failed JobSlide records for same analysis
        old_job_slides = (
            db.query(JobSlide)
            .join(AnalysisJob)
            .filter(
                JobSlide.slide_id == slide.id,
                AnalysisJob.analysis_id == analysis.id,
                JobSlide.status.in_(["completed", "failed"]),
            )
            .all()
        )
        for old_js in old_job_slides:
            db.delete(old_js)
            old_records_cleaned += 1
            remaining = db.query(JobSlide).filter(
                JobSlide.job_id == old_js.job_id, JobSlide.id != old_js.id
            ).count()
            if remaining == 0:
                old_job = db.query(AnalysisJob).filter_by(id=old_js.job_id).first()
                if old_job:
                    db.delete(old_job)

        # Batch dirs: all slides share one WSI dir and one output dir per job
        remote_wsi_batch_dir = f"{data.remote_wsi_dir}/{job.id}"
        remote_output_batch_dir = f"{data.remote_output_dir}/{job.id}"

        remote_wsi_path = f"{remote_wsi_batch_dir}/{slide_path.name}"

        job_slide = JobSlide(
            job_id=job.id,
            slide_id=slide.id,
            remote_wsi_path=remote_wsi_path,
            remote_output_path=remote_output_batch_dir,
            filename=slide_path.name,
            status="pending",
        )
        db.add(job_slide)
        db.flush()

        slides_to_process.append((job_slide.id, slide_hash, str(slide_path), remote_wsi_path))
        slides_created += 1

    db.commit()  # Commit and release DB immediately

    # --- Phase 2: Background thread for transfer + job launch ---
    job_id = job.id
    gpu_index = data.gpu_index
    remote_wsi_batch_dir = f"{data.remote_wsi_dir}/{job_id}"
    remote_output_batch_dir = f"{data.remote_output_dir}/{job_id}"

    def _run_submissions():
        n = len(slides_to_process)
        transfer_ok: list[int] = []
        transfer_errors: dict[int, str] = {}

        if direct_mount:
            # ── Direct mount: create symlinks on cluster instead of rsyncing ──
            print(f"[Job {job_id}] Direct mount — creating symlinks for {n} slides (no transfer needed)")
            bg_db = get_session()
            try:
                # Create the batch WSI dir on the cluster
                cluster_service.run_command(f"mkdir -p {remote_wsi_batch_dir}")

                for js_id, _hash, local_path_str, _remote in slides_to_process:
                    # Translate local server path to cluster mount path
                    cluster_path = local_path_str.replace(
                        local_to_cluster_prefix[0], local_to_cluster_prefix[1]
                    )
                    symlink_dest = f"{remote_wsi_batch_dir}/{Path(local_path_str).name}"

                    # Create symlink on the cluster: batch_dir/slide.svs -> /ligonlab/.../slide.svs
                    _, stderr, rc = cluster_service.run_command(
                        f"ln -sf {cluster_path!r} {symlink_dest!r}"
                    )
                    if rc != 0:
                        transfer_errors[js_id] = f"Symlink failed: {stderr}"
                    else:
                        transfer_ok.append(js_id)

                # Update DB statuses
                for js_id in transfer_ok:
                    js = bg_db.query(JobSlide).filter_by(id=js_id).first()
                    if js:
                        js.status = "transferring"  # brief transitional status
                for js_id, msg in transfer_errors.items():
                    js = bg_db.query(JobSlide).filter_by(id=js_id).first()
                    if js:
                        js.status = "failed"
                        js.error_message = msg
                bg_db.commit()
            except Exception:
                bg_db.rollback()
            finally:
                bg_db.close()
        else:
            # ── Rsync mode: transfer slides over SSH ──
            # Set all to 'transferring' upfront in one commit so the UI reflects progress.
            bg_db = get_session()
            try:
                for (js_id, _hash, _path, _remote) in slides_to_process:
                    js = bg_db.query(JobSlide).filter_by(id=js_id).first()
                    if js:
                        js.status = "transferring"
                bg_db.commit()
            except Exception:
                bg_db.rollback()
            finally:
                bg_db.close()

            # Rsync slides in parallel — no DB session held during transfer
            def _transfer_one(args):
                idx, js_id, local_path_str = args
                local_path = Path(local_path_str)
                if not local_path.exists():
                    return js_id, f"Local file not found: {local_path}"
                try:
                    mb = local_path.stat().st_size / 1e6
                    print(f"[Job {job_id}/Transfer {idx+1}/{n}] Rsyncing {local_path.name} ({mb:.0f} MB)")
                    cluster_service.rsync_slide(local_path, remote_wsi_batch_dir)
                    print(f"[Job {job_id}/Transfer {idx+1}/{n}] Done: {local_path.name}")
                    return js_id, None
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    return js_id, f"Transfer failed: {e}"

            work = [(i, js_id, lp) for i, (js_id, _sh, lp, _rp) in enumerate(slides_to_process)]
            with ThreadPoolExecutor(max_workers=3) as pool:
                for js_id, error in pool.map(_transfer_one, work):
                    if error:
                        transfer_errors[js_id] = error
                    else:
                        transfer_ok.append(js_id)

            # Commit rsync failures immediately so they show up
            if transfer_errors:
                bg_db = get_session()
                try:
                    for js_id, msg in transfer_errors.items():
                        js = bg_db.query(JobSlide).filter_by(id=js_id).first()
                        if js:
                            js.status = "failed"
                            js.error_message = msg
                    bg_db.commit()
                except Exception:
                    bg_db.rollback()
                finally:
                    bg_db.close()

        if not transfer_ok:
            print(f"[Job {job_id}] All {'symlinks' if direct_mount else 'transfers'} failed.")
            _recompute_job_status_standalone(job_id)
            return

        print(f"[Job {job_id}] {'Symlinks' if direct_mount else 'Transfers'} complete ({len(transfer_ok)}/{n}). Starting batch analysis...")

        # --- Phase B: Start tmux session then mark all successful slides as running ---
        # Both the tmux launch and the DB update happen in one block so they stay in sync.
        bg_db = get_session()
        try:
            _analysis = bg_db.query(Analysis).filter_by(id=analysis_snapshot["id"]).first()
            if not _analysis:
                for js_id in transfer_ok:
                    js = bg_db.query(JobSlide).filter_by(id=js_id).first()
                    if js:
                        js.status = "failed"
                        js.error_message = "Analysis not found"
                bg_db.commit()
                _recompute_job_status_standalone(job_id)
                return

            session_name = cluster_service.start_job(
                analysis=_analysis,
                job_id=job_id,
                remote_wsi_dir=remote_wsi_batch_dir,
                remote_output_dir=remote_output_batch_dir,
                gpu_index=gpu_index,
                parameters=params,
            )

            started_at = datetime.utcnow()
            for js_id in transfer_ok:
                js = bg_db.query(JobSlide).filter_by(id=js_id).first()
                if js:
                    js.cluster_job_id = session_name
                    js.status = "running"
                    js.started_at = started_at
            bg_db.commit()
            print(f"[Job {job_id}] Started tmux session '{session_name}' for {len(transfer_ok)} slides")

        except Exception as e:
            import traceback
            traceback.print_exc()
            bg_db.rollback()
            # Try once more with a fresh session to record the failure
            err_db = get_session()
            try:
                for js_id in transfer_ok:
                    js = err_db.query(JobSlide).filter_by(id=js_id).first()
                    if js:
                        js.status = "failed"
                        js.error_message = f"Job start failed: {e}"
                err_db.commit()
            except Exception:
                err_db.rollback()
            finally:
                err_db.close()
        finally:
            bg_db.close()

        _recompute_job_status_standalone(job_id)

    t = threading.Thread(target=_run_submissions, daemon=True)
    t.start()

    return {
        "job_id": job_id,
        "slides_created": slides_created,
        "old_records_cleaned": old_records_cleaned,
        "errors": errors,
        "cluster_connected": cluster_service.is_connected,
    }


def _recompute_job_status_standalone(job_id: int):
    """Open a fresh DB session, recompute parent job status, commit, close."""
    db = get_session()
    try:
        job = db.query(AnalysisJob).options(joinedload(AnalysisJob.slides)).filter_by(id=job_id).first()
        if job:
            _recompute_job_status(job)
            db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


def _recompute_job_status(job: AnalysisJob):
    """Derive parent job status from its child JobSlides."""
    if not job.slides:
        return
    statuses = [js.status for js in job.slides]
    if any(s in ("running", "transferring") for s in statuses):
        job.status = "running"
        if not job.started_at:
            job.started_at = datetime.utcnow()
    elif all(s == "completed" for s in statuses):
        job.status = "completed"
        job.completed_at = datetime.utcnow()
    elif any(s == "failed" for s in statuses) and not any(s in ("running", "transferring", "pending") for s in statuses):
        job.status = "failed"
        job.completed_at = datetime.utcnow()
    elif all(s == "pending" for s in statuses):
        job.status = "pending"


@app.get("/cohorts/{cohort_id}/flags")
def get_cohort_flags(cohort_id: int, db: Session = Depends(get_db)):
    """List all flags for a cohort."""
    flags = db.query(CohortFlag).filter_by(cohort_id=cohort_id).order_by(CohortFlag.created_at).all()
    return [{"id": f.id, "name": f.name, "case_hashes": f.get_case_hashes()} for f in flags]


@app.post("/cohorts/{cohort_id}/flags")
def create_cohort_flag(cohort_id: int, data: CohortFlagCreate, db: Session = Depends(get_db)):
    """Create a new cohort flag, optionally with initial case_hashes."""
    cohort = db.query(Cohort).filter_by(id=cohort_id).first()
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")
    name = data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Flag name cannot be empty")
    flag = CohortFlag(cohort_id=cohort_id, name=name)
    flag.set_case_hashes(data.case_hashes)
    db.add(flag)
    db.commit()
    db.refresh(flag)
    return {"id": flag.id, "name": flag.name, "case_hashes": flag.get_case_hashes()}


@app.patch("/cohorts/{cohort_id}/flags/{flag_id}")
def update_cohort_flag(cohort_id: int, flag_id: int, data: CohortFlagPatch, db: Session = Depends(get_db)):
    """Add or remove case_hashes from a flag."""
    flag = db.query(CohortFlag).filter_by(id=flag_id, cohort_id=cohort_id).first()
    if not flag:
        raise HTTPException(status_code=404, detail="Flag not found")
    current = set(flag.get_case_hashes())
    current.update(data.add_case_hashes)
    current -= set(data.remove_case_hashes)
    flag.set_case_hashes(list(current))
    db.commit()
    return {"id": flag.id, "name": flag.name, "case_hashes": flag.get_case_hashes()}


@app.delete("/cohorts/{cohort_id}/flags/{flag_id}")
def delete_cohort_flag(cohort_id: int, flag_id: int, db: Session = Depends(get_db)):
    """Delete a cohort flag."""
    flag = db.query(CohortFlag).filter_by(id=flag_id, cohort_id=cohort_id).first()
    if not flag:
        raise HTTPException(status_code=404, detail="Flag not found")
    db.delete(flag)
    db.commit()
    return {"status": "ok"}


@app.get("/cohorts/{cohort_id}/analysis-status")
def get_cohort_analysis_status(cohort_id: int, db: Session = Depends(get_db)):
    """Get per-slide analysis completion status for all slides in a cohort."""
    cohort = db.query(Cohort).options(
        joinedload(Cohort.slides)
    ).filter_by(id=cohort_id).first()
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")

    if not cohort.slides:
        return {"slides": {}}

    slide_ids = [s.id for s in cohort.slides]
    slide_id_to_hash = {s.id: s.slide_hash for s in cohort.slides}

    job_slides = (
        db.query(JobSlide)
        .options(joinedload(JobSlide.job))
        .filter(JobSlide.slide_id.in_(slide_ids))
        .all()
    )

    STATUS_PRIORITY = {"completed": 4, "running": 3, "transferring": 3, "pending": 2, "failed": 1}

    result: dict = {}
    for js in job_slides:
        slide_hash = slide_id_to_hash.get(js.slide_id)
        if not slide_hash or not js.job:
            continue
        analysis_key = js.job.model_name
        if slide_hash not in result:
            result[slide_hash] = {}
        existing = result[slide_hash].get(analysis_key)
        new_priority = STATUS_PRIORITY.get(js.status, 0)
        if existing is None or new_priority > STATUS_PRIORITY.get(existing["status"], 0):
            result[slide_hash][analysis_key] = {
                "status": js.status,
                "job_id": js.job.id,
                "analysis_id": js.job.analysis_id,
                "analysis_name": js.job.model_name,
            }

    return {"slides": result}


@app.post("/jobs/submit-cohort/{cohort_id}")
def submit_cohort_jobs(cohort_id: int, data: CohortJobSubmitRequest, db: Session = Depends(get_db)):
    """Submit a multi-slide analysis job for all (or selected) slides in a cohort."""
    cohort = db.query(Cohort).options(
        joinedload(Cohort.slides).joinedload(Slide.case)
    ).filter_by(id=cohort_id).first()
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")

    if data.case_hashes:
        case_hash_set = set(data.case_hashes)
        slide_hashes = [
            s.slide_hash for s in cohort.slides
            if s.case and s.case.accession_hash in case_hash_set
        ]
    else:
        slide_hashes = [s.slide_hash for s in cohort.slides]

    submit_data = JobSubmitRequest(
        analysis_id=data.analysis_id,
        slide_hashes=slide_hashes,
        gpu_index=data.gpu_index,
        remote_wsi_dir=data.remote_wsi_dir,
        remote_output_dir=data.remote_output_dir,
        parameters=data.parameters,
        submitted_by=data.submitted_by,
    )
    return submit_jobs(submit_data, db)


@app.get("/jobs")
def list_jobs(
    db: Session = Depends(get_db),
    status: Optional[str] = Query(None),
    analysis_id: Optional[int] = Query(None),
    slide_hashes: Optional[str] = Query(None, description="Comma-separated slide hashes to filter jobs containing these slides"),
    limit: int = Query(200, le=1000),
):
    """List analysis jobs with slide progress counts."""
    from sqlalchemy import func
    from sqlalchemy import case as sa_case

    # Aggregate counts in SQL — avoids loading log_tail blobs for every slide
    counts_sq = (
        db.query(
            JobSlide.job_id,
            func.count(JobSlide.id).label("total"),
            func.sum(sa_case((JobSlide.status == "completed", 1), else_=0)).label("completed"),
            func.sum(sa_case((JobSlide.status == "failed", 1), else_=0)).label("failed"),
        )
        .group_by(JobSlide.job_id)
        .subquery()
    )

    query = (
        db.query(
            AnalysisJob,
            func.coalesce(counts_sq.c.total, 0),
            func.coalesce(counts_sq.c.completed, 0),
            func.coalesce(counts_sq.c.failed, 0),
        )
        .outerjoin(counts_sq, AnalysisJob.id == counts_sq.c.job_id)
        # Eager-load the parent Analysis so we can include its `kind` in the
        # row payload without N+1 lazy-loads. The frontend uses kind to decide
        # whether a job's slides offer renderer buttons (UMAP / PCA / …).
        .options(joinedload(AnalysisJob.analysis))
    )

    if status:
        query = query.filter(AnalysisJob.status == status)
    if analysis_id:
        query = query.filter(AnalysisJob.analysis_id == analysis_id)
    if slide_hashes:
        wanted = [h.strip() for h in slide_hashes.split(",") if h.strip()]
        matching_jobs_sq = (
            db.query(JobSlide.job_id)
            .join(Slide, JobSlide.slide_id == Slide.id)
            .filter(Slide.slide_hash.in_(wanted))
            .subquery()
        )
        query = query.filter(AnalysisJob.id.in_(matching_jobs_sq))

    rows = query.order_by(AnalysisJob.submitted_at.desc()).limit(limit).all()

    return [
        {
            "id": j.id,
            "analysis_id": j.analysis_id,
            "analysis_kind": j.analysis.kind if j.analysis else None,
            "model_name": j.model_name,
            "model_version": j.model_version,
            "parameters": j.parameters,
            "gpu_index": j.gpu_index,
            "status": j.status,
            "submitted_by": j.submitted_by,
            "submitted_at": j.submitted_at.isoformat() if j.submitted_at else None,
            "started_at": j.started_at.isoformat() if j.started_at else None,
            "completed_at": j.completed_at.isoformat() if j.completed_at else None,
            "error_message": j.error_message,
            "slide_count": total,
            "completed_count": completed,
            "failed_count": failed,
        }
        for j, total, completed, failed in rows
    ]


@app.get("/jobs/{job_id}")
def get_job(job_id: int, db: Session = Depends(get_db)):
    """Get a single job with nested slides detail."""
    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.analysis),
        joinedload(AnalysisJob.slides).joinedload(JobSlide.slide)
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    slide_count = len(job.slides)
    completed_count = sum(1 for js in job.slides if js.status == "completed")
    failed_count = sum(1 for js in job.slides if js.status == "failed")
    postprocess_available = bool(job.analysis and job.analysis.postprocess_template)

    return {
        "id": job.id,
        "analysis_id": job.analysis_id,
        "analysis_kind": job.analysis.kind if job.analysis else None,
        "model_name": job.model_name,
        "model_version": job.model_version,
        "parameters": job.parameters,
        "gpu_index": job.gpu_index,
        "status": job.status,
        "submitted_by": job.submitted_by,
        "submitted_at": job.submitted_at.isoformat() if job.submitted_at else None,
        "started_at": job.started_at.isoformat() if job.started_at else None,
        "completed_at": job.completed_at.isoformat() if job.completed_at else None,
        "error_message": job.error_message,
        "slide_count": slide_count,
        "completed_count": completed_count,
        "failed_count": failed_count,
        "postprocess_available": postprocess_available,
        "slides": [
            {
                "id": js.id,
                "slide_hash": js.slide.slide_hash if js.slide else None,
                "slide_id": js.slide.slidecap_id if js.slide else None,
                "case_id": js.slide.case.slidecap_id if js.slide and js.slide.case else None,
                "filename": (indexer.get_filepath(js.slide.slide_hash).name if indexer and js.slide and indexer.get_filepath(js.slide.slide_hash) else None),
                # parse() returns None when the filename doesn't match the
                # parser pattern (used to crash here when an anchor was
                # stored locally under a non-standard name). Belt-and-
                # suspenders: surface None as `accession_number` rather
                # than 500'ing the whole endpoint.
                "accession_number": (lambda fp: (
                    (indexer.parser.parse(fp.name).accession if indexer.parser.parse(fp.name) else None) if fp else None
                ))(indexer.get_filepath(js.slide.slide_hash) if indexer and js.slide else None),
                "block_id": js.slide.block_id if js.slide else None,
                "stain_type": js.slide.stain_type if js.slide else None,
                "cluster_job_id": js.cluster_job_id,
                "status": js.status,
                "started_at": js.started_at.isoformat() if js.started_at else None,
                "completed_at": js.completed_at.isoformat() if js.completed_at else None,
                "error_message": js.error_message,
                "log_tail": js.log_tail,
                "remote_output_path": js.remote_output_path,
                "local_output_path": js.local_output_path,
                "cell_stats": json.loads(js.cell_stats) if js.cell_stats else _parse_cell_stats(js.log_tail, js.local_output_path, js.filename),
            }
            for js in job.slides
        ],
    }


@app.get("/jobs/{job_id}/log")
def get_job_log(
    job_id: int,
    slide_id: Optional[int] = Query(None, description="Optional JobSlide ID for specific slide log"),
    db: Session = Depends(get_db),
):
    """Fetch log from the cluster for a job or specific slide within it."""
    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.slides)
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # If slide_id specified, get that specific slide's log
    if slide_id is not None:
        job_slide = next((js for js in job.slides if js.id == slide_id), None)
        if not job_slide:
            raise HTTPException(status_code=404, detail="JobSlide not found in this job")
        slides_to_check = [job_slide]
    else:
        slides_to_check = job.slides

    if not cluster_service or not cluster_service.is_connected:
        logs = []
        for js in slides_to_check:
            logs.append({"slide_id": js.id, "log": js.log_tail or "(not connected)", "source": "cached"})
        return {"job_id": job_id, "slides": logs}

    results = []
    for js in slides_to_check:
        remote_out = js.remote_output_path
        if not remote_out:
            results.append({"slide_id": js.id, "log": "(no remote output path)", "source": "none"})
            continue
        try:
            stdout, stderr, exit_code = cluster_service.run_command(f"cat {remote_out}/run.log 2>&1")
            _, _, tmux_code = cluster_service.run_command(
                f"tmux has-session -t {js.cluster_job_id} 2>/dev/null" if js.cluster_job_id else "false"
            )
            results.append({
                "slide_id": js.id,
                "log": stdout if exit_code == 0 else f"(no run.log found)\nstderr: {stderr}",
                "tmux_alive": tmux_code == 0,
                "cluster_job_id": js.cluster_job_id,
                "source": "live",
            })
        except Exception as e:
            results.append({"slide_id": js.id, "log": f"Error fetching log: {e}", "source": "error"})

    return {"job_id": job_id, "slides": results}


@app.post("/jobs/cancel")
def cancel_jobs(data: JobCancelRequest, db: Session = Depends(get_db)):
    """Cancel jobs by their IDs. Cancels all active child slides."""
    cancelled = 0
    errors = []

    for job_id in data.job_ids:
        job = db.query(AnalysisJob).options(
            joinedload(AnalysisJob.slides)
        ).filter_by(id=job_id).first()
        if not job:
            errors.append(f"Job {job_id} not found")
            continue

        if job.status not in ("queued", "running", "pending", "transferring"):
            errors.append(f"Job {job_id} is already {job.status}")
            continue

        # Cancel all active child slides
        for js in job.slides:
            if js.status in ("pending", "transferring", "running", "queued"):
                if cluster_service and cluster_service.is_connected and js.cluster_job_id:
                    try:
                        cluster_service.cancel_job(js.cluster_job_id)
                    except Exception:
                        pass
                js.status = "failed"
                js.error_message = "Cancelled by user"
                js.completed_at = datetime.utcnow()

        job.status = "failed"
        job.error_message = "Cancelled by user"
        job.completed_at = datetime.utcnow()
        cancelled += 1

    db.commit()

    return {"cancelled": cancelled, "errors": errors}


@app.post("/jobs/{job_id}/retry")
def retry_job(job_id: int, db: Session = Depends(get_db)):
    """
    Retry failed/cancelled slides in an existing job without re-uploading slides
    that already made it to the cluster.

    Reuses the same remote directories (keyed by job_id), so rsync will:
      - Skip files that are already fully uploaded (size+mtime match)
      - Resume files that were only partially uploaded (--partial flag)
    """
    if not cluster_service or not cluster_service.is_connected:
        raise HTTPException(status_code=503, detail="Not connected to cluster. Connect first.")

    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.slides).joinedload(JobSlide.slide)
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    if job.status in ("running", "transferring", "pending"):
        raise HTTPException(status_code=400, detail=f"Job is currently {job.status} — wait for it to finish or cancel it first")

    analysis = db.query(Analysis).filter_by(id=job.analysis_id, active=True).first()
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis pipeline not found or inactive")

    # Collect slides that need to be retried
    slides_to_retry = [js for js in job.slides if js.status in ("failed", "cancelled")]
    if not slides_to_retry:
        raise HTTPException(status_code=400, detail="No failed slides to retry")

    # Resolve local paths now (indexer lookup must happen in the request thread)
    slides_to_process: list[tuple[int, str, str, str]] = []
    skipped = []
    for js in slides_to_retry:
        if not js.slide:
            continue
        slide_hash = js.slide.slide_hash
        local_path = indexer.get_filepath(slide_hash) if indexer else None
        if not local_path:
            skipped.append(slide_hash[:12])
            continue
        slides_to_process.append((js.id, slide_hash, str(local_path), js.remote_wsi_path or ""))

    if not slides_to_process:
        raise HTTPException(status_code=400, detail="No local files found for failed slides")

    # Reset slides to pending
    for js in slides_to_retry:
        js.status = "pending"
        js.error_message = None
        js.cluster_job_id = None
        js.started_at = None
        js.completed_at = None
        js.log_tail = None
    job.status = "pending"
    job.error_message = None
    job.completed_at = None
    db.commit()

    # Snapshot what the background thread needs (detached from DB session)
    analysis_snapshot = {
        "id": analysis.id,
        "name": analysis.name,
        "version": analysis.version,
        "script_path": analysis.script_path,
        "working_directory": analysis.working_directory,
        "env_setup": analysis.env_setup,
        "command_template": analysis.command_template,
        "gpu_required": analysis.gpu_required,
    }
    gpu_index = job.gpu_index or 0
    params = json.loads(job.parameters) if job.parameters else None
    remote_wsi_batch_dir = f"{job.remote_wsi_dir}/{job_id}"
    remote_output_batch_dir = f"{job.remote_output_dir}/{job_id}"

    def _run_retry():
        n = len(slides_to_process)

        # Mark all as transferring upfront
        bg_db = get_session()
        try:
            for (js_id, *_) in slides_to_process:
                js = bg_db.query(JobSlide).filter_by(id=js_id).first()
                if js:
                    js.status = "transferring"
            bg_db.commit()
        except Exception:
            bg_db.rollback()
        finally:
            bg_db.close()

        transfer_ok: list[int] = []
        transfer_errors: dict[int, str] = {}

        def _retry_one(args):
            idx, js_id, local_path_str = args
            local_path = Path(local_path_str)
            if not local_path.exists():
                return js_id, f"Local file not found: {local_path}"
            try:
                local_size = local_path.stat().st_size
                remote_path = f"{remote_wsi_batch_dir}/{local_path.name}"
                stdout, _, rc = cluster_service.run_command(
                    f"stat -c%s '{remote_path}' 2>/dev/null"
                )
                if rc == 0 and stdout.strip().isdigit() and int(stdout.strip()) == local_size:
                    print(f"[Job {job_id}/Retry {idx+1}/{n}] Already on cluster: {local_path.name} — skipping")
                    return js_id, None
                mb = local_size / 1e6
                print(f"[Job {job_id}/Retry {idx+1}/{n}] Rsyncing {local_path.name} ({mb:.0f} MB)")
                cluster_service.rsync_slide(local_path, remote_wsi_batch_dir)
                print(f"[Job {job_id}/Retry {idx+1}/{n}] Done: {local_path.name}")
                return js_id, None
            except Exception as e:
                import traceback; traceback.print_exc()
                return js_id, f"Transfer failed: {e}"

        retry_work = [(i, js_id, lp) for i, (js_id, _sh, lp, _rp) in enumerate(slides_to_process)]
        with ThreadPoolExecutor(max_workers=3) as pool:
            for js_id, error in pool.map(_retry_one, retry_work):
                if error:
                    transfer_errors[js_id] = error
                else:
                    transfer_ok.append(js_id)

        if transfer_errors:
            bg_db = get_session()
            try:
                for js_id, msg in transfer_errors.items():
                    js = bg_db.query(JobSlide).filter_by(id=js_id).first()
                    if js:
                        js.status = "failed"
                        js.error_message = msg
                bg_db.commit()
            except Exception:
                bg_db.rollback()
            finally:
                bg_db.close()

        if not transfer_ok:
            print(f"[Job {job_id}/Retry] All transfers failed.")
            _recompute_job_status_standalone(job_id)
            return

        print(f"[Job {job_id}/Retry] Transfers complete ({len(transfer_ok)}/{n}). Starting job...")

        bg_db = get_session()
        try:
            _analysis = bg_db.query(Analysis).filter_by(id=analysis_snapshot["id"]).first()
            if not _analysis:
                raise RuntimeError("Analysis not found")

            session_name = cluster_service.start_job(
                analysis=_analysis,
                job_id=job_id,
                remote_wsi_dir=remote_wsi_batch_dir,
                remote_output_dir=remote_output_batch_dir,
                gpu_index=gpu_index,
                parameters=params,
            )

            started_at = datetime.utcnow()
            for js_id in transfer_ok:
                js = bg_db.query(JobSlide).filter_by(id=js_id).first()
                if js:
                    js.cluster_job_id = session_name
                    js.status = "running"
                    js.started_at = started_at
            bg_db.commit()
            print(f"[Job {job_id}/Retry] Started tmux session '{session_name}'")

        except Exception as e:
            import traceback; traceback.print_exc()
            bg_db.rollback()
            err_db = get_session()
            try:
                for js_id in transfer_ok:
                    js = err_db.query(JobSlide).filter_by(id=js_id).first()
                    if js:
                        js.status = "failed"
                        js.error_message = f"Job start failed: {e}"
                err_db.commit()
            except Exception:
                err_db.rollback()
            finally:
                err_db.close()
        finally:
            bg_db.close()

        _recompute_job_status_standalone(job_id)

    t = threading.Thread(target=_run_retry, daemon=True)
    t.start()

    return {
        "job_id": job_id,
        "retrying": len(slides_to_process),
        "skipped_no_path": skipped,
    }


@app.delete("/jobs/{job_id}")
def delete_job(job_id: int, delete_files: bool = False, db: Session = Depends(get_db)):
    """Hard-delete a job and all its JobSlide records from the database.
    If delete_files=true, also removes output directories from disk."""
    import shutil

    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.slides)
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    if job.status in ("running", "transferring"):
        raise HTTPException(status_code=400, detail="Cannot delete a running job. Cancel it first.")

    files_deleted = 0
    if delete_files:
        for js in job.slides:
            if js.local_output_path:
                p = Path(js.local_output_path)
                if p.is_dir():
                    shutil.rmtree(p, ignore_errors=True)
                    files_deleted += 1

    with get_lock().write_lock():
        for js in job.slides:
            db.delete(js)
        db.delete(job)
        db.commit()

    return {"status": "ok", "deleted_job_id": job_id, "directories_removed": files_deleted}


class JobDeleteBulkRequest(BaseModel):
    job_ids: List[int]


@app.post("/jobs/delete-bulk")
def delete_jobs_bulk(data: JobDeleteBulkRequest, db: Session = Depends(get_db)):
    """Hard-delete multiple jobs and their JobSlide records."""
    deleted = 0
    errors = []

    with get_lock().write_lock():
        for job_id in data.job_ids:
            job = db.query(AnalysisJob).options(
                joinedload(AnalysisJob.slides)
            ).filter_by(id=job_id).first()
            if not job:
                errors.append(f"Job {job_id} not found")
                continue
            if job.status in ("running", "transferring"):
                errors.append(f"Job {job_id} is still {job.status}")
                continue
            for js in job.slides:
                db.delete(js)
            db.delete(job)
            deleted += 1

        db.commit()

    return {"deleted": deleted, "errors": errors}


@app.post("/jobs/refresh")
def refresh_job_statuses():
    """Trigger an immediate job status refresh from the cluster."""
    if not job_poller:
        raise HTTPException(status_code=503, detail="Job poller not running")
    if not cluster_service or not cluster_service.is_connected:
        raise HTTPException(status_code=503, detail="Not connected to cluster")

    job_poller.poll_now()
    return {"status": "ok", "message": "Status refresh triggered"}


@app.post("/jobs/{job_id}/transfer-results")
def transfer_job_results(job_id: int, db: Session = Depends(get_db)):
    """Manually trigger result transfer from cluster for completed slides in a job."""
    if not job_poller:
        raise HTTPException(status_code=503, detail="Job poller not running")
    if not cluster_service or not cluster_service.is_connected:
        raise HTTPException(status_code=503, detail="Not connected to cluster")

    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.slides).joinedload(JobSlide.slide)
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # Collect slides that need transfer (no local output yet)
    slides_needing_transfer = [
        js for js in job.slides
        if not (js.local_output_path and Path(js.local_output_path).exists()
                and any(Path(js.local_output_path).iterdir()))
    ]

    if not slides_needing_transfer:
        return {"transferred": 0, "errors": [], "message": "All slides already transferred"}

    try:
        job_poller._transfer_and_distribute(slides_needing_transfer, db)
        db.commit()
        transferred = sum(1 for js in slides_needing_transfer if js.status == "completed")
        errors = [f"Slide {js.id} ({js.filename}): {js.error_message}"
                  for js in slides_needing_transfer if js.status == "failed" and js.error_message]
    except Exception as e:
        errors = [str(e)]
        transferred = 0

    return {"transferred": transferred, "errors": errors}


# ============================================================
# Cluster Connection Endpoints
# ============================================================

class ClusterConnectRequest(BaseModel):
    host: str
    port: int = 22
    username: str
    password: str


@app.post("/cluster/connect")
def cluster_connect(data: ClusterConnectRequest):
    """Connect to the GPU cluster via SSH."""
    if _is_demo():
        global _demo_cluster_connected
        # Brief artificial delay so the spinner is visible during the demo
        import time as _t; _t.sleep(0.6)
        _demo_cluster_connected = True
        return demo_mod.mock_cluster_connect(data.host, data.username)

    if not cluster_service:
        raise HTTPException(status_code=503, detail="Cluster service not initialized")

    try:
        result = cluster_service.connect(data.host, data.port, data.username, data.password)
        # Immediately query GPU status
        try:
            gpus = cluster_service.get_gpu_status()
            result["gpus"] = gpus
        except Exception as e:
            result["gpus"] = []
            result["gpu_error"] = str(e)
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/cluster/disconnect")
def cluster_disconnect():
    """Disconnect from the GPU cluster."""
    if _is_demo():
        global _demo_cluster_connected
        _demo_cluster_connected = False
        return {"status": "ok", "connected": False}
    if cluster_service:
        cluster_service.disconnect()
    return {"status": "ok", "connected": False}


@app.get("/cluster/status")
def cluster_status():
    """Get cluster connection status. Actively probes the SSH connection."""
    if _is_demo():
        return demo_mod.mock_cluster_status(_demo_cluster_connected)

    if not cluster_service:
        return {"connected": False}

    info = cluster_service.connection_info
    if info["connected"]:
        if cluster_service.ping():
            try:
                info["gpus"] = cluster_service.get_gpu_status()
            except Exception:
                info["gpus"] = []
        else:
            cluster_service.disconnect()
            info["connected"] = False
            info["gpus"] = []
    return info


@app.get("/cluster/gpus")
def cluster_gpus():
    """Get current GPU status from the cluster."""
    if _is_demo():
        if not _demo_cluster_connected:
            raise HTTPException(status_code=503, detail="Not connected to cluster")
        return demo_mod.fake_gpus()

    if not cluster_service or not cluster_service.is_connected:
        raise HTTPException(status_code=503, detail="Not connected to cluster")

    try:
        return cluster_service.get_gpu_status()
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to query GPUs: {e}")


# ============================================================
# Seed / Pre-configuration Endpoints
# ============================================================

@app.post("/analyses/seed-cellvit")
def seed_cellvit(db: Session = Depends(get_db)):
    """Pre-populate registry with the CellViT analysis pipeline."""
    existing = db.query(Analysis).filter_by(name="CellViT").first()
    if existing:
        return {"status": "already_exists", "id": existing.id}

    # Resolve postprocess script path relative to this file
    scripts_dir = Path(__file__).resolve().parent.parent / "scripts"
    postprocess_cmd = f"python {scripts_dir / 'postprocess_cellvit.py'} --input-dir {{input_dir}} --output-dir {{output_dir}}"

    analysis = Analysis(
        name="CellViT",
        version="SAM-H-x40",
        description="Cell detection and segmentation using CellViT with SAM-H backbone at 40x magnification",
        kind="cellvit",  # transforms come from the kind's default_rules — see analyses/cellvit.py
        # (UNI seeder lives below this function; kinds are independent.)
        script_path="/ligonlab/Prem/CellViT_CCNU/CellViT-plus-plus/run_cellvit_resume.sh",
        working_directory="/ligonlab/Prem/CellViT_CCNU/CellViT-plus-plus",
        env_setup="source cellvit_env/bin/activate && export TMPDIR=/ligonlab/Prem/CellViT_CCNU/tmp && export RAY_EXPERIMENTAL_NOSET_CUDA_VISIBLE_DEVICES=1",
        command_template="./run_cellvit_resume.sh {wsi_dir} {outdir} ./checkpoints/CellViT-SAM-H-x40-AMP.pth {gpu} {batch_size}",
        postprocess_template=postprocess_cmd,
        default_parameters='{"batch_size": 4}',
        gpu_required=True,
        estimated_runtime_minutes=120,
    )
    db.add(analysis)
    db.commit()

    return {"status": "created", "id": analysis.id, "name": analysis.name}


@app.post("/analyses/seed-uni")
def seed_uni(db: Session = Depends(get_db)):
    """
    Pre-populate the registry with a UNI analysis pipeline.

    Like seed-cellvit, this is idempotent — re-running returns the existing row.
    Mainline fields (script_path / command_template) are placeholders; the
    operator edits them in Analysis Registry to point at their actual cluster
    paths. The important bit for the new architecture is `kind="uni"` so the
    UMAP/PCA renderers in analyses/uni.py become available.
    """
    existing = db.query(Analysis).filter_by(name="UNI").first()
    if existing:
        return {"status": "already_exists", "id": existing.id}

    analysis = Analysis(
        name="UNI",
        version="v2-20x-256px",
        description="Foundation-model embeddings (UNI v2). Per-patch 1024-d CLS tokens, viewable as UMAP/PCA scatter linked to slide coords.",
        kind="uni",
        # No transforms override — the kind has no default rules either; UNI
        # outputs are uncompressed. All the value is in the per-slide renderers.
        script_path="<set me in Analysis Registry>",
        working_directory="<set me in Analysis Registry>",
        env_setup="<set me in Analysis Registry>",
        command_template="<set me in Analysis Registry>",
        default_parameters='{"patch_size": 256, "magnification": "20x"}',
        gpu_required=True,
        estimated_runtime_minutes=45,
    )
    db.add(analysis)
    db.commit()

    return {"status": "created", "id": analysis.id, "name": analysis.name}


# ============================================================
# Results Endpoints
# ------------------------------------------------------------------


def _parse_cell_stats(log_tail: Optional[str],
                      local_output_path: Optional[str] = None,
                      filename: Optional[str] = None) -> Optional[dict]:
    """Extract per-slide cell count stats from a CellViT log.

    Preferentially reads the full run.log from local_output_path and extracts
    the section for the specific slide (identified by filename stem), so that
    stats are correct even in multi-slide batch jobs where log_tail only
    contains the last slide's output.

    Falls back to scanning log_tail directly.
    """
    import ast
    import re as _re

    def _extract_stats_from_text(text: str) -> Optional[dict]:
        """Find last stats dict in the given text block."""
        for line in reversed(text.splitlines()):
            line = line.strip()
            if " - " in line:
                line = line.split(" - ", 2)[-1].strip()
            if line.startswith("{") and line.endswith("}"):
                try:
                    parsed = ast.literal_eval(line)
                    if isinstance(parsed, dict) and all(isinstance(v, (int, float)) for v in parsed.values()):
                        return parsed
                except Exception:
                    continue
        return None

    # Try reading per-slide section from the local run.log first
    if local_output_path and filename:
        run_log = Path(local_output_path) / "run.log"
        if run_log.exists():
            try:
                full_log = run_log.read_text(errors="replace")
                stem = Path(filename).stem
                # Find the section of the log that belongs to this slide
                wsi_pattern = _re.compile(r"Processing WSI:\s+(\S+)")
                sections: list[tuple[str, int]] = []
                for m in wsi_pattern.finditer(full_log):
                    sections.append((Path(m.group(1)).stem, m.start()))
                for i, (s, start) in enumerate(sections):
                    if s == stem:
                        end = sections[i + 1][1] if i + 1 < len(sections) else len(full_log)
                        result = _extract_stats_from_text(full_log[start:end])
                        if result is not None:
                            return result
                        break
                # Fallback: no section match — scan whole log
                result = _extract_stats_from_text(full_log)
                if result is not None:
                    return result
            except Exception:
                pass

    # Final fallback: use log_tail stored in DB
    return _extract_stats_from_text(log_tail or "")
# ============================================================

@app.get("/results/search")
def search_results(
    q: str = Query(..., min_length=1, description="Accession number search query"),
    db: Session = Depends(get_db),
):
    """
    Search for analysis results by accession number.
    Returns slides matching the query with their completed analyses.
    """
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    query_upper = q.strip().upper()

    # Find matching slides via the path cache (same pattern as /search)
    matching: list[dict] = []
    seen_hashes: set[str] = set()

    for slide_hash, filepath in indexer.slide_hash_to_path.items():
        parsed = indexer.parser.parse(filepath.name)
        if not parsed:
            continue
        if query_upper not in parsed.accession.upper():
            continue
        if slide_hash in seen_hashes:
            continue
        seen_hashes.add(slide_hash)

        slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
        if not slide:
            continue

        # Get completed analyses for this slide
        completed = (
            db.query(JobSlide)
            .join(AnalysisJob)
            .filter(
                JobSlide.slide_id == slide.id,
                JobSlide.status == "completed",
            )
            .order_by(JobSlide.completed_at.desc())
            .all()
        )

        if not completed:
            continue

        matching.append({
            "slide_hash": slide.slide_hash,
            "slide_id": slide.slidecap_id,
            "case_id": slide.case.slidecap_id if slide.case else None,
            "accession_number": parsed.accession,
            "block_id": slide.block_id,
            "stain_type": slide.stain_type,
            "year": slide.case.year if slide.case else None,
            "results": [
                {
                    "job_id": js.job_id,
                    "job_slide_id": js.id,
                    "analysis_name": js.job.model_name,
                    "version": js.job.model_version or "",
                    "status": js.status,
                    "completed_at": js.completed_at.isoformat() if js.completed_at else None,
                    "output_path": js.local_output_path or js.remote_output_path,
                    "cell_stats": json.loads(js.cell_stats) if js.cell_stats else _parse_cell_stats(js.log_tail, js.local_output_path, js.filename),
                }
                for js in completed
            ],
        })

    return {"query": q, "count": len(matching), "results": matching}


@app.get("/slides/{slide_hash}/results")
def get_slide_results(slide_hash: str, db: Session = Depends(get_db)):
    """List completed analysis results for a slide."""
    slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()
    if not slide:
        raise HTTPException(status_code=404, detail="Slide not found")

    completed_job_slides = (
        db.query(JobSlide)
        .join(AnalysisJob)
        .options(joinedload(JobSlide.job).joinedload(AnalysisJob.analysis))
        .filter(
            JobSlide.slide_id == slide.id,
            JobSlide.status == "completed",
        )
        .order_by(JobSlide.completed_at.desc())
        .all()
    )

    return [
        {
            "job_id": js.job_id,
            "job_slide_id": js.id,
            "analysis_name": js.job.model_name,
            "version": js.job.model_version or "",
            # analysis_kind lets the frontend decide which renderer buttons
            # to show for this row (UMAP / PCA / …). Pulled off the Analysis
            # row joined above.
            "analysis_kind": js.job.analysis.kind if js.job.analysis else None,
            "status": js.status,
            "completed_at": js.completed_at.isoformat() if js.completed_at else None,
            "output_path": js.local_output_path or js.remote_output_path,
            "cell_stats": _parse_cell_stats(js.log_tail, js.local_output_path, js.filename),
        }
        for js in completed_job_slides
    ]


@app.get("/results/{job_id}/files")
def list_result_files(
    job_id: int,
    slide_hash: Optional[str] = Query(None, description="Slide hash to get per-slide output"),
    db: Session = Depends(get_db),
):
    """List output files for a completed job. Use slide_hash for per-slide results."""
    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.slides).joinedload(JobSlide.slide)
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # Determine output directory (prefer local path if available)
    output_dir = None
    if slide_hash:
        js = next((js for js in job.slides if js.slide and js.slide.slide_hash == slide_hash), None)
        if js:
            if js.local_output_path:
                output_dir = Path(js.local_output_path)
            elif js.remote_output_path:
                output_dir = Path(js.remote_output_path)
    elif job.output_path:
        output_dir = Path(job.output_path)

    if not output_dir or not output_dir.exists():
        return []

    return _build_file_tree(output_dir, output_dir)




@app.get("/results/{job_id}/file/{filename:path}")
def get_result_file(
    job_id: int,
    filename: str,
    slide_hash: Optional[str] = Query(None),
    apply_transforms: bool = Query(False, description="Apply this analysis's transform rules to the file (e.g. snappy decompress, geometry fix)"),
    decompress: bool = Query(False, description="Legacy alias for apply_transforms (kept for older clients)"),
    db: Session = Depends(get_db),
):
    """
    Serve a single result file for download or preview.

    When `apply_transforms=true` (or the legacy `decompress=true`), the bytes
    are piped through the Analysis's configured transform rules before being
    sent. See backend/app/analyses/ for the per-kind op registry.
    """
    from . import analyses as _ax

    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.slides).joinedload(JobSlide.slide),
        joinedload(AnalysisJob.analysis),
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # Determine output directory (prefer local path if available)
    output_dir = None
    if slide_hash:
        js = next((js for js in job.slides if js.slide and js.slide.slide_hash == slide_hash), None)
        if js:
            if js.local_output_path:
                output_dir = Path(js.local_output_path)
            elif js.remote_output_path:
                output_dir = Path(js.remote_output_path)
    elif job.output_path:
        output_dir = Path(job.output_path)

    if not output_dir:
        raise HTTPException(status_code=404, detail="No output path for this job")

    # Resolve path safely (prevent traversal)
    try:
        file_path = _safe_resolve(output_dir, filename)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid file path")

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="File not found")

    if apply_transforms or decompress:
        # Resolve the kind plugin from the Analysis row, falling back to the
        # default kind for legacy rows. The kind owns its op registry, and
        # `resolve_rules` falls back to its baked-in default_rules when the
        # Analysis row has no `transforms` override.
        kind_id = job.analysis.kind if job.analysis else None
        kind = _ax.get_kind(kind_id)
        if kind is None:
            raise HTTPException(
                status_code=500,
                detail=f"Unknown analysis kind {kind_id!r} — no plugin registered. "
                       f"Check backend/app/analyses/.",
            )
        rules = _ax.resolve_rules(kind, job.analysis.transforms if job.analysis else None)
        if rules:
            try:
                transformed, applied = _ax.apply_rules(kind, file_path.name, file_path.read_bytes(), rules)
            except _ax.TransformError as e:
                # Surface the specific op that failed — never silently serve raw bytes
                # when transforms were requested, that just hides the bug.
                raise HTTPException(status_code=500, detail=str(e))
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Transform pipeline error: {e}")
            if applied:
                # Strip the matched extensions to produce a sensible output filename
                out_name = file_path.name
                # Convention: snappy/gzip ops imply we'd remove that suffix from the served name
                for suffix in ('.snappy', '.gz'):
                    if out_name.endswith(suffix):
                        out_name = out_name.removesuffix(suffix)
                        break
                media = 'application/geo+json' if out_name.endswith('.geojson') else 'application/octet-stream'
                return Response(
                    content=transformed,
                    media_type=media,
                    headers={
                        "Content-Disposition": f'inline; filename="{out_name}"',
                        "X-Transforms-Applied": ",".join(applied),
                        # Transformed bytes are deterministic per (file mtime,
                        # rule set). Let the browser cache for an hour so
                        # reopening an overlay is instant. We don't bust on
                        # rule edits — the user re-fetches via shift-reload
                        # in that rare case.
                        "Cache-Control": "private, max-age=3600",
                    },
                )

    return FileResponse(str(file_path), filename=file_path.name)


@app.get("/results/{job_id}/render/{renderer_id}")
def render_result(
    job_id: int,
    renderer_id: str,
    slide_hash: str = Query(..., description="Which slide in the job to render"),
    recompute: bool = Query(False, description="Bust the on-disk projection cache"),
    db: Session = Depends(get_db),
):
    """
    Invoke a per-slide renderer declared by the Analysis's kind plugin.

    Unlike `/results/{job_id}/file/...` which serves a single file through
    optional bytes→bytes ops, this endpoint hands the kind's renderer the
    slide's whole output directory + stem so it can read multiple sibling
    files (e.g. UNI's UMAP joins features.h5 + patches.h5). The renderer
    returns arbitrary JSON the frontend's plot component knows how to draw.
    """
    from . import analyses as _ax

    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.slides).joinedload(JobSlide.slide),
        joinedload(AnalysisJob.analysis),
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if not job.analysis:
        raise HTTPException(status_code=400, detail="Job has no associated analysis")

    kind = _ax.get_kind(job.analysis.kind)
    if kind is None:
        raise HTTPException(status_code=500, detail=f"Unknown analysis kind {job.analysis.kind!r}")
    renderer = kind.renderers.get(renderer_id)
    if renderer is None:
        available = [r["id"] for r in kind.list_renderers()]
        raise HTTPException(
            status_code=404,
            detail=f"Renderer {renderer_id!r} not declared by kind {kind.id!r}. Available: {available}",
        )

    # Locate the slide's output dir + source filename (same logic as get_result_file).
    js = next((js for js in job.slides if js.slide and js.slide.slide_hash == slide_hash), None)
    if not js:
        raise HTTPException(status_code=404, detail=f"Slide {slide_hash} not in job {job_id}")
    output_dir = Path(js.local_output_path or js.remote_output_path or "")
    if not output_dir or not output_dir.exists():
        raise HTTPException(
            status_code=404,
            detail=f"No local output dir for slide {slide_hash} (path: {output_dir or '<unset>'})",
        )
    # `filename` is the WSI name as recorded by the cluster runner; strip its
    # extension to get the stem that UNI uses for all its sibling files.
    if not js.filename:
        raise HTTPException(status_code=400, detail="JobSlide has no filename — cannot derive stem")
    stem = Path(js.filename).stem

    try:
        result = renderer.fn(output_dir, stem, {"recompute": recompute})
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except RuntimeError as e:
        # Renderer raised something diagnostic (missing dep, bad h5 schema, …).
        # 500 with the original message so the UI can surface it.
        raise HTTPException(status_code=500, detail=str(e))

    # Browser-cache renderer JSON for an hour — same rationale as the
    # transformed file endpoint. Server-side, the UNI renderer also caches
    # to disk next to the source .h5, so even on a cold cache miss the
    # second-and-later requests skip the UMAP fit.
    return JSONResponse(
        content={"renderer": renderer_id, "kind": kind.id, "stem": stem, "data": result},
        headers={"Cache-Control": "private, max-age=3600"},
    )


@app.get("/transforms/ops")
def list_transform_ops(kind: Optional[str] = Query(None, description="Scope ops to a specific analysis kind id (e.g. 'cellvit'). Required.")):
    """
    List the transform ops available for a given analysis kind. Used by the
    Analysis Registry UI to populate its per-rule op picker.

    `kind` is required — the global, kind-less listing was removed alongside
    the per-kind plugin refactor, because mixing UNI ops into a CellViT
    pipeline (or vice-versa) silently produces wrong output.
    """
    from . import analyses as _ax
    if not kind:
        raise HTTPException(
            status_code=400,
            detail="kind query parameter is required (e.g. ?kind=cellvit). "
                   "Use GET /analyses/kinds to list available kinds.",
        )
    k = _ax.get_kind(kind)
    if k is None:
        raise HTTPException(status_code=404, detail=f"Unknown analysis kind {kind!r}")
    return k.list_ops()


@app.delete("/results/{job_id}/file/{filename:path}")
def delete_result_file(
    job_id: int,
    filename: str,
    slide_hash: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Delete a specific output file from disk for a job slide."""
    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.slides).joinedload(JobSlide.slide)
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    output_dir = None
    if slide_hash:
        js = next((js for js in job.slides if js.slide and js.slide.slide_hash == slide_hash), None)
        if js:
            output_dir = _resolve_job_slide_output(js)
    elif job.output_path:
        output_dir = Path(job.output_path)

    if not output_dir:
        raise HTTPException(status_code=404, detail="No output path for this job")

    try:
        file_path = _safe_resolve(output_dir, filename)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid file path")

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="File not found")

    file_path.unlink()
    return {"status": "ok", "deleted": filename}


@app.get("/results/{job_id}/download-folder")
def download_result_folder_zip(
    job_id: int,
    slide_hash: Optional[str] = Query(None),
    folder_path: str = Query(""),
    db: Session = Depends(get_db),
):
    """Download a subfolder of a slide's output as a ZIP file."""
    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.slides).joinedload(JobSlide.slide)
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    output_dir = None
    if slide_hash:
        js = next((js for js in job.slides if js.slide and js.slide.slide_hash == slide_hash), None)
        if js:
            output_dir = _resolve_job_slide_output(js)
    elif job.output_path:
        output_dir = Path(job.output_path)

    if not output_dir:
        raise HTTPException(status_code=404, detail="No output path for this job")

    try:
        target = _safe_resolve(output_dir, folder_path) if folder_path else output_dir.resolve()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid folder path")

    if not target.is_dir():
        raise HTTPException(status_code=404, detail="Folder not found")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        _add_files_to_zip(zf, target, target.name)
    buf.seek(0)

    zip_name = f"{target.name}.zip"
    return Response(
        content=buf.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{zip_name}"'},
    )


def _resolve_job_slide_output(js) -> Optional[Path]:
    """Get the output directory for a JobSlide, if it exists locally."""
    for attr in ("local_output_path", "remote_output_path"):
        val = getattr(js, attr, None)
        if val:
            p = Path(val)
            if p.is_dir():
                return p
    return None


_RESULT_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".gif", ".svg"}


def _build_file_tree(base_dir: Path, current_dir: Path) -> list:
    """Recursively build a file/folder tree for a result output directory."""
    result = []
    try:
        entries = sorted(current_dir.iterdir(), key=lambda e: (e.is_file(), e.name.lower()))
    except PermissionError:
        return result
    for entry in entries:
        rel = entry.relative_to(base_dir).as_posix()
        if entry.is_dir():
            result.append({
                "name": entry.name,
                "type": "dir",
                "path": rel,
                "children": _build_file_tree(base_dir, entry),
            })
        else:
            result.append({
                "name": entry.name,
                "type": "file",
                "path": rel,
                "size": entry.stat().st_size,
                "is_image": entry.suffix.lower() in _RESULT_IMAGE_EXTS,
            })
    return result


def _safe_resolve(output_dir: Path, rel_path: str) -> Path:
    """
    Resolve a relative path within output_dir, raising ValueError on
    traversal attempts (``..`` segments that escape the chroot).

    Does NOT follow symlinks — we deliberately keep symlink targets
    out of the traversal check because the demo's redacted-anchor dir
    is a symlink farm pointing back at NETWORK_ROOT. Resolving symlinks
    would land the path outside ``output_dir`` and falsely trip the
    traversal guard. We trust symlink targets because we built them; we
    don't trust the user-supplied ``rel_path``, which is what this guards.
    """
    import os
    # normpath collapses `..` and `.` against the logical path without
    # following symlinks. relative_to then verifies the result is still
    # inside output_dir as a string-level check.
    logical = Path(os.path.normpath(str(output_dir / rel_path)))
    logical.relative_to(output_dir)  # raises ValueError if outside
    return logical


def _add_files_to_zip(zf: zipfile.ZipFile, output_dir: Path, arc_prefix: str):
    """Recursively add all files from output_dir to the ZIP as-is."""
    for f in sorted(output_dir.rglob("*")):
        if not f.is_file():
            continue
        rel = f.relative_to(output_dir).as_posix()
        zf.write(str(f), f"{arc_prefix}/{rel}")


@app.get("/jobs/{job_id}/download-zip")
def download_job_zip(job_id: int, db: Session = Depends(get_db)):
    """Stream a ZIP of all completed slides' output files for a job."""
    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.slides).joinedload(JobSlide.slide)
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # Collect slides with local output
    slides_with_output: list[tuple[str, Path]] = []
    for js in job.slides:
        output_dir = _resolve_job_slide_output(js)
        if output_dir and js.slide:
            slides_with_output.append((js.slide.slide_hash, output_dir))

    if not slides_with_output:
        raise HTTPException(status_code=404, detail="No local output files available")

    def generate():
        q: queue.Queue = queue.Queue(maxsize=32)

        def writer():
            try:
                stream = _ZipStreamWriter(q)
                with zipfile.ZipFile(stream, 'w', zipfile.ZIP_DEFLATED, allowZip64=True) as zf:
                    for slide_hash, output_dir in slides_with_output:
                        _add_files_to_zip(zf, output_dir, slide_hash[:12])
                stream.flush()
            except Exception as e:
                print(f"Job ZIP export error: {e}")
            finally:
                q.put(None)

        t = threading.Thread(target=writer, daemon=True)
        t.start()
        while True:
            chunk = q.get()
            if chunk is None:
                break
            yield chunk
        t.join(timeout=5)

    return StreamingResponse(
        generate(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="job_{job_id}_results.zip"',
        },
    )


class CartItem(BaseModel):
    job_id: int
    slide_hash: str
    filename: str


class CartDownloadRequest(BaseModel):
    items: List[CartItem]


@app.post("/results/download-cart")
def download_cart(data: CartDownloadRequest, db: Session = Depends(get_db)):
    """Stream a ZIP of selected files from the download cart, decompressing .snappy."""
    if not data.items:
        raise HTTPException(status_code=400, detail="Cart is empty")

    # Resolve each item to a file path
    resolved: list[tuple[Path, str]] = []  # (file_path, arcname)
    for item in data.items:
        job = db.query(AnalysisJob).options(
            joinedload(AnalysisJob.slides).joinedload(JobSlide.slide)
        ).filter_by(id=item.job_id).first()
        if not job:
            continue

        js = next((js for js in job.slides if js.slide and js.slide.slide_hash == item.slide_hash), None)
        if not js:
            continue

        output_dir = _resolve_job_slide_output(js)
        if not output_dir:
            continue

        try:
            file_path = _safe_resolve(output_dir, item.filename)
        except ValueError:
            continue
        if not file_path.exists() or not file_path.is_file():
            continue

        arc_prefix = item.slide_hash[:12]
        rel = file_path.relative_to(output_dir).as_posix()
        resolved.append((file_path, f"{arc_prefix}/{rel}"))

    if not resolved:
        raise HTTPException(status_code=404, detail="No files found for the selected items")

    def generate():
        q: queue.Queue = queue.Queue(maxsize=32)

        def writer():
            try:
                stream = _ZipStreamWriter(q)
                with zipfile.ZipFile(stream, 'w', zipfile.ZIP_DEFLATED, allowZip64=True) as zf:
                    for file_path, arcname in resolved:
                        zf.write(str(file_path), arcname)
                stream.flush()
            except Exception as e:
                print(f"Cart ZIP export error: {e}")
            finally:
                q.put(None)

        t = threading.Thread(target=writer, daemon=True)
        t.start()
        while True:
            chunk = q.get()
            if chunk is None:
                break
            yield chunk
        t.join(timeout=5)

    return StreamingResponse(
        generate(),
        media_type="application/zip",
        headers={
            "Content-Disposition": 'attachment; filename="selected_results.zip"',
        },
    )


@app.get("/jobs/{job_id}/output-filenames")
def get_job_output_filenames(
    job_id: int,
    slide_hashes: Optional[str] = Query(None, description="Comma-separated slide hashes to restrict to"),
    db: Session = Depends(get_db),
):
    """Return output filenames grouped by slide, with annotation counts."""
    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.slides).joinedload(JobSlide.slide).joinedload(Slide.case)
    ).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    wanted = None
    if slide_hashes:
        wanted = set(h.strip() for h in slide_hashes.split(",") if h.strip())

    groups: list[dict] = []
    for js in job.slides:
        if not js.slide:
            continue
        sh = js.slide.slide_hash
        if wanted and sh not in wanted:
            continue

        # Resolve a human-readable label: accession_block_stain
        label = sh[:12] + "..."
        if indexer:
            fp = indexer.get_filepath(sh)
            if fp:
                parsed = indexer.parser.parse(fp.name)
                if parsed:
                    parts = [parsed.accession]
                    if parsed.block_id:
                        parts.append(parsed.block_id)
                    if parsed.stain_type:
                        parts.append(parsed.stain_type)
                    label = '_'.join(parts)

        # Output files — prefer local (network drive), fall back to remote (cluster).
        # Walk recursively so nested layouts (UNI's
        # `20x_256px_0px_overlap/features_uni_v2/<stem>.h5` etc.) are visible
        # in the file list, not just top-level entries. Filenames are stored
        # as POSIX-relative paths so the frontend's download/overlay clicks
        # can pass them straight through to the per-file endpoint.
        files: list[str] = []
        is_local = bool(js.local_output_path and Path(js.local_output_path).is_dir())
        output_dir = _resolve_job_slide_output(js)
        if output_dir:
            for f in sorted(output_dir.rglob("*")):
                if f.is_file():
                    files.append(f.relative_to(output_dir).as_posix())

        # Annotation count
        ann_dir = settings.annotations_path / sh
        annotation_count = 0
        if ann_dir.is_dir():
            annotation_count = sum(1 for f in ann_dir.iterdir() if f.is_file())

        groups.append({
            "slide_hash": sh,
            "label": label,
            "files": files,
            "annotation_count": annotation_count,
            "is_local": is_local,
        })

    return groups


class DownloadBundleRequest(BaseModel):
    slide_hashes: List[str]
    job_id: int
    include_filenames: List[str]
    # wsi_slide_hashes: explicit per-slide WSI inclusion (preferred).
    # Fallback: if empty but include_wsi=True, include WSI for all slide_hashes.
    wsi_slide_hashes: List[str] = []
    include_wsi: bool = False  # legacy fallback
    include_annotations: bool = False
    # If True and the analysis has a postprocess_template, run it on-the-fly
    # per slide via a temp dir before adding files to the ZIP.
    apply_postprocess: bool = False


@app.post("/download-bundle")
def download_bundle(data: DownloadBundleRequest, db: Session = Depends(get_db)):
    """
    Stream a ZIP with selected output files (as-is) and optionally the
    original WSI for each requested slide.

    ZIP structure: {accession_block_stain}/{filename}
    """
    job = db.query(AnalysisJob).options(
        joinedload(AnalysisJob.analysis),
        joinedload(AnalysisJob.slides).joinedload(JobSlide.slide).joinedload(Slide.case)
    ).filter_by(id=data.job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    pp_template = (
        job.analysis.postprocess_template
        if data.apply_postprocess and job.analysis
        else None
    )

    requested = set(data.slide_hashes)
    include_set = set(data.include_filenames)

    # Build WSI set: explicit per-slide list takes priority; fall back to
    # include_wsi=True meaning "all requested slides".
    if data.wsi_slide_hashes:
        wsi_set = set(data.wsi_slide_hashes)
    elif data.include_wsi:
        wsi_set = requested
    else:
        wsi_set = set()

    # If WSI is requested but the path cache might be stale, refresh it once
    if wsi_set and indexer:
        missing = all(indexer.get_filepath(h) is None for h in wsi_set)
        if missing:
            print("WSI requested but no paths in cache — refreshing path cache")
            indexer.build_path_cache()

    # Resolve per-slide: (folder_name, slide_hash, output_dir, wsi_path)
    items: list[tuple[str, str, Optional[Path], Optional[Path]]] = []
    for js in job.slides:
        if not js.slide or js.slide.slide_hash not in requested:
            continue

        slide_hash = js.slide.slide_hash
        output_dir = _resolve_job_slide_output(js)

        # Resolve folder name: accession_block_stain if possible, else hash prefix
        wsi_path = None
        folder = slide_hash[:12]
        if indexer:
            fp = indexer.get_filepath(slide_hash)
            if fp:
                parsed = indexer.parser.parse(fp.name)
                if parsed:
                    parts = [parsed.accession]
                    if parsed.block_id:
                        parts.append(parsed.block_id)
                    if parsed.stain_type:
                        parts.append(parsed.stain_type)
                    folder = '_'.join(parts)
                if slide_hash in wsi_set:
                    wsi_path = fp

        items.append((folder, slide_hash, output_dir, wsi_path))

    if not items:
        raise HTTPException(status_code=404, detail="No matching slides found")

    def generate():
        q: queue.Queue = queue.Queue(maxsize=32)

        def writer():
            import tempfile
            try:
                stream = _ZipStreamWriter(q)
                with zipfile.ZipFile(stream, 'w', zipfile.ZIP_STORED, allowZip64=True) as zf:
                    for folder, slide_hash, output_dir, wsi_path in items:
                        # Add selected output files (with optional on-the-fly postprocessing)
                        if output_dir:
                            if pp_template:
                                # Run postprocess into a temp dir, then add outputs
                                with tempfile.TemporaryDirectory() as tmpdir:
                                    cmd = pp_template.format(
                                        input_dir=str(output_dir),
                                        output_dir=tmpdir,
                                        filename_stem=output_dir.name,
                                    )
                                    try:
                                        subprocess.run(cmd, shell=True, check=True, timeout=300)
                                        src_dir = Path(tmpdir)
                                    except Exception as e:
                                        print(f"Postprocess failed for {folder}: {e}, using raw files")
                                        src_dir = output_dir
                                    for f in sorted(src_dir.iterdir()):
                                        if f.is_file() and (include_set and f.name in include_set or not include_set):
                                            zf.write(str(f), f"{folder}/{f.name}")
                            else:
                                for f in sorted(output_dir.iterdir()):
                                    if not f.is_file() or f.name not in include_set:
                                        continue
                                    zf.write(str(f), f"{folder}/{f.name}")
                        # Add annotations
                        if data.include_annotations:
                            ann_dir = settings.annotations_path / slide_hash
                            if ann_dir.is_dir():
                                for f in sorted(ann_dir.iterdir()):
                                    if f.is_file():
                                        zf.write(str(f), f"{folder}/annotations/{f.name}")
                        # Add WSI
                        if wsi_path:
                            try:
                                zf.write(str(wsi_path), f"{folder}/{wsi_path.name}")
                            except Exception as e:
                                print(f"Warning: could not add WSI {wsi_path}: {e}")
                stream.flush()
            except Exception as e:
                print(f"Bundle ZIP error: {e}")
            finally:
                q.put(None)

        t = threading.Thread(target=writer, daemon=True)
        t.start()
        while True:
            chunk = q.get()
            if chunk is None:
                break
            yield chunk
        t.join(timeout=5)

    return StreamingResponse(
        generate(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="job_{data.job_id}_bundle.zip"',
        },
    )


# ============================================================
# Dashboard & Staging Endpoints
# ============================================================


class SortRequest(BaseModel):
    filenames: List[str] = []
    tags: List[str] = []        # Tags to apply to sorted slides after indexing
    tag_color: Optional[str] = None  # Color for newly-created tags


# ── Sort background job state ────────────────────────────────────────────────
_sort_state: dict = {
    "running": False,
    "done": False,
    "total": 0,
    "current": 0,
    "current_file": "",
    "sorted": 0,
    "skipped": 0,
    "errors": [],
}


def _run_sort_background(filenames: list, tags: list = None, tag_color: str = None) -> None:
    """
    Two-phase background sort for maximum speed:
      Phase 1 – OS rename every file (atomic, instant on same filesystem)
      Phase 2 – Batch-index all moved files in a single DB transaction
    """
    global _sort_state
    _sort_state.update({
        "running": True,
        "done": False,
        "total": len(filenames),
        "current": 0,
        "current_file": "",
        "sorted": 0,
        "skipped": 0,
        "errors": [],
    })

    staging = settings.staging_path
    moved: list = []  # (dest_path, parsed)

    # Phase 1: rename/move (no data copy on same-FS network volume)
    for i, fn in enumerate(filenames):
        _sort_state["current"] = i + 1
        _sort_state["current_file"] = fn

        f = staging / fn
        if not f.is_file():
            _sort_state["skipped"] += 1
            _sort_state["errors"].append(f"{fn}: file not found")
            continue

        parsed = indexer.parser.parse(fn)
        if not parsed:
            _sort_state["skipped"] += 1
            _sort_state["errors"].append(f"{fn}: cannot parse filename")
            continue

        year_dir = settings.slides_path / str(parsed.year)
        canonical_name = parsed.full_stem + ".svs"
        dest = year_dir / canonical_name

        if dest.exists():
            _sort_state["skipped"] += 1
            _sort_state["errors"].append(f"{fn}: already exists in destination")
            continue

        try:
            year_dir.mkdir(parents=True, exist_ok=True)
            shutil.move(str(f), str(dest))
            moved.append((dest, parsed))
            _sort_state["sorted"] += 1
        except Exception as e:
            _sort_state["errors"].append(f"{fn}: {str(e)}")
            _sort_state["skipped"] += 1

    # Phase 2: single batch DB commit
    new_slides = []
    if moved:
        _sort_state["current_file"] = "Indexing…"
        db = get_session()
        try:
            for dest, _parsed in moved:
                result = indexer.index_file(db, dest)
                if result:
                    _, slide = result
                    indexer.slide_hash_to_path[slide.slide_hash] = dest
                    new_slides.append(slide)
            db.commit()
        except Exception as e:
            db.rollback()
            _sort_state["errors"].append(f"Indexing error: {str(e)}")
        finally:
            db.close()

    # Phase 3: apply tags to newly sorted slides
    if tags and new_slides:
        _sort_state["current_file"] = "Applying tags…"
        db = get_session()
        try:
            with get_lock().write_lock():
                tag_objects = []
                for tag_name in tags:
                    tag = db.query(Tag).filter_by(name=tag_name).first()
                    if not tag:
                        tag = Tag(name=tag_name, color=tag_color)
                        db.add(tag)
                        db.flush()
                    tag_objects.append(tag)

                for slide in new_slides:
                    # Re-attach slide to this session
                    slide = db.merge(slide)
                    for tag in tag_objects:
                        if tag not in slide.tags:
                            slide.tags.append(tag)

                db.commit()
        except Exception as e:
            db.rollback()
            _sort_state["errors"].append(f"Tagging error: {str(e)}")
        finally:
            db.close()

    _sort_state.update({"running": False, "done": True, "current_file": ""})


@app.get("/staging/sort/status")
def staging_sort_status():
    """Poll current sort job progress."""
    return _sort_state


@app.delete("/staging/file/{filename}")
def staging_delete_file(filename: str):
    """Delete a single file from the staging folder."""
    safe_name = Path(filename).name  # strip any path components
    f = settings.staging_path / safe_name
    print(f"[staging-delete] Attempting to delete: {f}")
    if not f.exists():
        print(f"[staging-delete] File not found: {f}")
        raise HTTPException(status_code=404, detail=f"File not found in staging: {safe_name}")
    try:
        f.unlink()
        print(f"[staging-delete] Deleted: {f}")
    except PermissionError as e:
        print(f"[staging-delete] Permission denied: {f} — {e}")
        raise HTTPException(status_code=500, detail=f"Permission denied: {safe_name}. File may be locked or read-only.")
    except OSError as e:
        print(f"[staging-delete] OS error: {f} — {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete {safe_name}: {e}")
    return {"deleted": safe_name}


@app.get("/staging/scan")
def staging_scan():
    """Scan staging folder for .svs files and parse their filenames."""
    if _is_demo():
        global _demo_staging_files
        _demo_staging_files = demo_mod.mock_staging_scan()
        return _demo_staging_files

    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    staging = settings.staging_path
    if not staging.exists():
        return []

    results = []
    # Track slide hashes seen within staging to detect duplicates
    seen_hashes: dict[str, str] = {}  # slide_hash -> first filename
    print(f"[staging-scan] Cache has {len(indexer.slide_hash_to_path)} indexed slides, slides_path={settings.slides_path}")

    # os.scandir caches DirEntry.stat() from the directory listing on SMB,
    # avoiding a separate stat() network call per file.
    with os.scandir(staging) as it:
        entries = sorted(it, key=lambda e: e.name)

    for entry in entries:
        if not entry.is_file(follow_symlinks=False) or not entry.name.lower().endswith('.svs'):
            continue

        try:
            size = entry.stat(follow_symlinks=False).st_size
        except OSError:
            size = 0

        result_entry = {
            "filename": entry.name,
            "size_bytes": size,
            "parsed": False,
            "accession": None,
            "block_id": None,
            "slide_number": None,
            "stain_type": None,
            "year": None,
            "destination": None,
            "conflict": False,
            "conflict_reason": None,
        }

        parsed = indexer.parser.parse(entry.name)
        if parsed:
            canonical_name = parsed.full_stem + ".svs"
            slide_hash = indexer.hasher.hash_slide_stem(parsed.full_stem)

            # Check for conflict: already indexed OR duplicate within staging
            is_indexed_conflict = slide_hash in indexer.slide_hash_to_path
            is_staging_duplicate = slide_hash in seen_hashes

            # Filesystem fallback: check if destination file already exists
            # (catches cases where hash cache is stale)
            year_dir = settings.slides_path / str(parsed.year)
            dest = year_dir / canonical_name
            is_dest_conflict = not is_indexed_conflict and dest.exists()

            conflict = is_indexed_conflict or is_staging_duplicate or is_dest_conflict
            conflict_reason = None
            if is_indexed_conflict:
                conflict_reason = "already indexed"
                print(f"[staging-scan] CONFLICT (indexed): {entry.name} -> hash={slide_hash[:12]}... stem={parsed.full_stem}")
            elif is_dest_conflict:
                conflict_reason = "already indexed"
                print(f"[staging-scan] CONFLICT (dest exists): {entry.name} -> {dest}")
            elif is_staging_duplicate:
                conflict_reason = f"duplicate of {seen_hashes[slide_hash]}"
                print(f"[staging-scan] CONFLICT (staging dup): {entry.name} -> duplicate of {seen_hashes[slide_hash]}")
            else:
                print(f"[staging-scan] OK: {entry.name} -> hash={slide_hash[:12]}... stem={parsed.full_stem}")

            if not is_staging_duplicate:
                seen_hashes[slide_hash] = entry.name

            result_entry.update({
                "parsed": True,
                "accession": parsed.accession,
                "block_id": parsed.block_id,
                "slide_number": parsed.slide_number,
                "stain_type": parsed.stain_type,
                "year": parsed.year,
                "destination": f"slides/{parsed.year}/{canonical_name}",
                "conflict": conflict,
                "conflict_reason": conflict_reason,
            })

        results.append(result_entry)

    conflicts = sum(1 for r in results if r["conflict"])
    print(f"[staging-scan] Done: {len(results)} files, {conflicts} conflicts")
    return results


@app.post("/staging/sort")
def staging_sort(req: SortRequest):
    """Start sort in background. Returns immediately — poll /staging/sort/status for progress."""
    if _is_demo():
        if _sort_state.get("running"):
            raise HTTPException(status_code=409, detail="Sort already in progress")
        # Use either the requested filenames or the cached scan
        filenames = req.filenames or [f["filename"] for f in _demo_staging_files]
        if not filenames:
            return {"started": False, "total": 0, "message": "No files to sort"}
        demo_mod.start_demo_sort(_sort_state, filenames)
        return {"started": True, "total": len(filenames)}

    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    if _sort_state.get("running"):
        raise HTTPException(status_code=409, detail="Sort already in progress")

    staging = settings.staging_path
    if not staging.exists():
        return {"started": False, "total": 0, "message": "Staging folder not found"}

    if req.filenames:
        filenames = [fn for fn in req.filenames if (staging / fn).is_file()]
    else:
        filenames = [
            f.name for f in staging.iterdir()
            if f.is_file() and f.name.lower().endswith('.svs')
        ]

    if not filenames:
        return {"started": False, "total": 0, "message": "No files to sort"}

    threading.Thread(target=_run_sort_background, args=(filenames, req.tags, req.tag_color), daemon=True).start()
    return {"started": True, "total": len(filenames)}


@app.get("/export/slides.csv")
def export_slides_csv(db: Session = Depends(get_db)):
    """Export all slides with metadata and tags as a CSV file."""
    slides = db.query(Slide).options(
        joinedload(Slide.tags),
        joinedload(Slide.case).joinedload(Case.tags),
    ).all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "slide_hash", "filename", "accession", "year",
        "block_id", "slide_number", "stain_type", "random_id",
        "file_size_bytes", "slide_tags", "case_tags",
    ])

    for slide in slides:
        filepath = indexer.get_filepath(slide.slide_hash) if indexer else None
        filename = filepath.name if filepath else ""
        if filepath and indexer:
            parsed = indexer.parser.parse(filepath.name)
            accession = parsed.accession if parsed else ""
            slide_number = parsed.slide_number if parsed else ""
        else:
            accession = ""
            slide_number = ""

        writer.writerow([
            slide.slide_hash,
            filename,
            accession,
            slide.case.year if slide.case else "",
            slide.block_id or "",
            slide_number,
            slide.stain_type or "",
            slide.random_id or "",
            slide.file_size_bytes or "",
            ";".join(t.name for t in slide.tags),
            ";".join(t.name for t in slide.case.tags) if slide.case else "",
        ])

    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="slides_export.csv"'},
    )


@app.post("/import/slides-csv")
async def import_slides_csv(
    file: UploadFile = File(...),
    mode: str = Query("merge", enum=["merge", "replace"]),
    db: Session = Depends(get_db),
):
    """
    Bulk-import slide tags from a CSV file.

    Required column: slide_hash
    Optional columns: slide_tags (semicolon-separated), case_tags (semicolon-separated)

    mode='merge'   — add CSV tags to existing tags (default)
    mode='replace' — replace all slide/case tags with CSV values
    """
    raw = await file.read()
    text = raw.decode("utf-8-sig")  # handle Excel BOM

    reader = csv.DictReader(io.StringIO(text))
    if "slide_hash" not in (reader.fieldnames or []):
        raise HTTPException(status_code=400, detail="CSV must contain a 'slide_hash' column")

    matched = 0
    unmatched = 0
    tags_added = 0

    for row in reader:
        slide_hash = row.get("slide_hash", "").strip()
        if not slide_hash:
            continue

        slide = db.query(Slide).options(
            joinedload(Slide.tags),
            joinedload(Slide.case).joinedload(Case.tags),
        ).filter_by(slide_hash=slide_hash).first()

        if not slide:
            unmatched += 1
            continue
        matched += 1

        def _apply_tags(obj, raw_str: str):
            nonlocal tags_added
            names = [n.strip() for n in raw_str.split(";") if n.strip()]
            if not names:
                return
            if mode == "replace":
                obj.tags.clear()
            existing = {t.name for t in obj.tags}
            for name in names:
                if name in existing:
                    continue
                tag = db.query(Tag).filter_by(name=name).first()
                if not tag:
                    tag = Tag(name=name)
                    db.add(tag)
                    db.flush()
                obj.tags.append(tag)
                existing.add(name)
                tags_added += 1

        _apply_tags(slide, row.get("slide_tags", ""))
        if slide.case:
            _apply_tags(slide.case, row.get("case_tags", ""))

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

    return {"matched": matched, "unmatched": unmatched, "tags_added": tags_added, "mode": mode}


@app.get("/dashboard/summary")
def dashboard_summary(db: Session = Depends(get_db)):
    """Aggregated dashboard data: library stats, staging, recent jobs, storage."""
    if _is_demo():
        # Return a fast canned payload + recent jobs from the demo DB
        payload = demo_mod.mock_dashboard_summary()
        recent_jobs = (
            db.query(AnalysisJob)
            .options(joinedload(AnalysisJob.slides))
            .order_by(AnalysisJob.submitted_at.desc())
            .limit(5)
            .all()
        )
        payload["recent_jobs"] = [
            {
                "id": j.id,
                "model_name": j.model_name,
                "status": j.status,
                "slide_count": len(j.slides),
                "submitted_at": j.submitted_at.isoformat() if j.submitted_at else None,
                "completed_at": j.completed_at.isoformat() if j.completed_at else None,
            }
            for j in recent_jobs
        ]
        return payload

    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    # Library stats
    stats = indexer.get_stats(db)

    # Year breakdown from DB
    year_counts = {}
    for case in db.query(Case).all():
        yr = case.year
        if yr:
            slide_count = len(case.slides)
            year_counts[yr] = year_counts.get(yr, 0) + slide_count

    # Staging info
    staging = settings.staging_path
    staging_count = 0
    staging_size = 0
    if staging.exists():
        for f in staging.iterdir():
            if f.is_file() and f.name.lower().endswith('.svs'):
                staging_count += 1
                staging_size += f.stat().st_size

    # Recent jobs
    recent_jobs = (
        db.query(AnalysisJob)
        .options(joinedload(AnalysisJob.slides))
        .order_by(AnalysisJob.submitted_at.desc())
        .limit(5)
        .all()
    )
    jobs_data = [
        {
            "id": j.id,
            "model_name": j.model_name,
            "status": j.status,
            "slide_count": len(j.slides),
            "submitted_at": j.submitted_at.isoformat() if j.submitted_at else None,
            "completed_at": j.completed_at.isoformat() if j.completed_at else None,
        }
        for j in recent_jobs
    ]

    # Storage sizes (lightweight: just sum file sizes)
    def dir_size_mb(path: Path) -> int:
        if not path.exists():
            return 0
        total = 0
        try:
            for dirpath, _, filenames in os.walk(path):
                for fn in filenames:
                    try:
                        total += os.path.getsize(os.path.join(dirpath, fn))
                    except OSError:
                        pass
        except OSError:
            pass
        return total // (1024 * 1024)

    return {
        "library": {
            "total_slides": stats['total_slides'],
            "total_cases": stats['total_cases'],
            "years": year_counts,
        },
        "staging": {
            "count": staging_count,
            "total_size_bytes": staging_size,
        },
        "recent_jobs": jobs_data,
        "storage": {
            "network_root": settings.NETWORK_ROOT,
            "slides_size_mb": dir_size_mb(settings.slides_path),
            "analyses_size_mb": dir_size_mb(settings.analyses_path),
            "staging_size_mb": dir_size_mb(settings.staging_path),
        },
    }


# ============================================================
# Request Tracker
# ============================================================

class RequestSheetCreate(BaseModel):
    name: str
    description: Optional[str] = None
    created_by: Optional[str] = None

class RequestSheetUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    # When provided, replace the sheet's auto-tags with exactly this list.
    # Pass [] to clear all. None means "no change".
    auto_tag_ids: Optional[List[int]] = None

class RequestRowCreate(BaseModel):
    accession_number: str
    case_status: Optional[str] = 'Not Started'
    all_blocks: Optional[str] = None
    blocks_available: Optional[str] = None
    order_id: Optional[str] = None
    is_consult: bool = False
    blocks_hes_requested: Optional[str] = None
    hes_requested: int = 0
    non_hes_requested: int = 0
    ihc_stains_requested: Optional[str] = None
    block_hes_received: Optional[str] = None
    hes_received: int = 0
    unaccounted_blocks: Optional[str] = None
    non_hes_received: int = 0
    fs_received: int = 0
    uss_received: int = 0
    ihc_received: int = 0
    ihc_stains_received: Optional[str] = None
    recut_blocks: Optional[str] = None
    recut_status: Optional[str] = None
    hes_scanned: Optional[str] = None
    he_scanning_status: Optional[str] = None
    non_hes_scanned: Optional[str] = None
    slide_location: Optional[str] = None
    notes: Optional[str] = None

class RequestRowUpdate(BaseModel):
    accession_number: Optional[str] = None
    case_status: Optional[str] = None
    all_blocks: Optional[str] = None
    blocks_available: Optional[str] = None
    order_id: Optional[str] = None
    is_consult: Optional[bool] = None
    blocks_hes_requested: Optional[str] = None
    hes_requested: Optional[int] = None
    non_hes_requested: Optional[int] = None
    ihc_stains_requested: Optional[str] = None
    block_hes_received: Optional[str] = None
    hes_received: Optional[int] = None
    unaccounted_blocks: Optional[str] = None
    non_hes_received: Optional[int] = None
    fs_received: Optional[int] = None
    uss_received: Optional[int] = None
    ihc_received: Optional[int] = None
    ihc_stains_received: Optional[str] = None
    recut_blocks: Optional[str] = None
    recut_status: Optional[str] = None
    hes_scanned: Optional[str] = None
    he_scanning_status: Optional[str] = None
    non_hes_scanned: Optional[str] = None
    slide_location: Optional[str] = None
    notes: Optional[str] = None

class ImportCohortRequest(BaseModel):
    cohort_id: int


def _normalize_accession(acc: str) -> str:
    """Normalize accession: BS-26-D12345 → BS26-D12345."""
    return re.sub(r'^([A-Z]{2})-(\d{2})-', r'\1\2-', acc.strip().upper())


def _serialize_request_row(r: RequestRow) -> dict:
    return {
        "id": r.id,
        "sheet_id": r.sheet_id,
        "accession_number": r.accession_number,
        "case_status": r.case_status,
        "all_blocks": r.all_blocks,
        "blocks_available": r.blocks_available,
        "order_id": r.order_id,
        "is_consult": r.is_consult,
        "blocks_hes_requested": r.blocks_hes_requested,
        "hes_requested": r.hes_requested,
        "non_hes_requested": r.non_hes_requested,
        "ihc_stains_requested": r.ihc_stains_requested,
        "block_hes_received": r.block_hes_received,
        "hes_received": r.hes_received,
        "unaccounted_blocks": r.unaccounted_blocks,
        "non_hes_received": r.non_hes_received,
        "fs_received": r.fs_received,
        "uss_received": r.uss_received,
        "ihc_received": r.ihc_received,
        "ihc_stains_received": r.ihc_stains_received,
        "recut_blocks": r.recut_blocks,
        "recut_status": r.recut_status,
        "hes_scanned": r.hes_scanned,
        "he_scanning_status": r.he_scanning_status,
        "non_hes_scanned": r.non_hes_scanned,
        "slide_location": r.slide_location,
        "notes": r.notes,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "updated_at": r.updated_at.isoformat() if r.updated_at else None,
    }


@app.get("/request-sheets")
def list_request_sheets(db: Session = Depends(get_db)):
    sheets = db.query(RequestSheet).order_by(RequestSheet.updated_at.desc()).all()
    return [{
        "id": s.id,
        "name": s.name,
        "description": s.description,
        "case_count": s.case_count,
        "created_by": s.created_by,
        "created_at": s.created_at.isoformat() if s.created_at else None,
        "updated_at": s.updated_at.isoformat() if s.updated_at else None,
    } for s in sheets]


@app.get("/request-sheets/case-warnings")
def case_warnings(
    accession: str = Query(...),
    sheet_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """
    Check if a case (by accession number) already has slides in the database
    and/or appears in other request sheets.
    """
    warnings: list[dict] = []

    # 1. Check if slides already exist in the database for this accession
    if indexer:
        results = indexer.search(db, query=accession, limit=50)
        exact = [r for r in results if _normalize_accession(r.get('accession_number', '')) == _normalize_accession(accession)]
        if exact:
            stains = {}
            for s in exact:
                st = s.get('stain_type', 'Unknown')
                stains[st] = stains.get(st, 0) + 1
            warnings.append({
                "type": "already_scanned",
                "message": f"{len(exact)} slide(s) already in database",
                "slide_count": len(exact),
                "stain_breakdown": stains,
            })

    # 2. Check if this accession appears in other request sheets
    normalized = _normalize_accession(accession)
    other_rows = (
        db.query(RequestRow)
        .join(RequestSheet)
        .filter(RequestRow.accession_number == normalized)
        .filter(RequestRow.sheet_id != sheet_id)
        .all()
    )
    if other_rows:
        sheets_info = []
        for r in other_rows:
            sheet = db.query(RequestSheet).filter_by(id=r.sheet_id).first()
            sheets_info.append({
                "sheet_id": r.sheet_id,
                "sheet_name": sheet.name if sheet else f"Sheet #{r.sheet_id}",
                "case_status": r.case_status,
            })
        warnings.append({
            "type": "duplicate_request",
            "message": f"Also in {len(sheets_info)} other sheet(s)",
            "sheets": sheets_info,
        })

    return {"warnings": warnings}


_BLOCK_SPLIT_RE = re.compile(r"[,;\s/]+")


def _parse_block_list(raw: Optional[str]) -> list[str]:
    """
    Split a free-text block field into a list of canonical block IDs.

    Handles comma, semicolon, slash, and whitespace separators and
    uppercases each token. Empty/None input returns [].
    """
    if not raw:
        return []
    seen: set[str] = set()
    out: list[str] = []
    for tok in _BLOCK_SPLIT_RE.split(raw):
        norm = tok.strip().upper()
        if norm and norm not in seen:
            seen.add(norm)
            out.append(norm)
    return out


def _apply_auto_tags_to_accession(db: Session, accession: str, tags: list) -> int:
    """Add each tag to the Case matching this accession. Returns count of tags newly added."""
    if not hasher or not tags:
        return 0
    acc_hash = hasher.hash_accession(_normalize_accession(accession))
    case = db.query(Case).filter_by(accession_hash=acc_hash).first()
    if not case:
        return 0
    added = 0
    existing = set(case.tags)
    for tag in tags:
        if tag not in existing:
            case.tags.append(tag)
            existing.add(tag)
            added += 1
    return added


def _apply_auto_tags_to_sheet(db: Session, sheet: "RequestSheet") -> int:
    """Apply every sheet.auto_tags entry to every row's accession.
    Returns the count of (case, tag) pairs newly created."""
    tags = list(sheet.auto_tags)
    if not tags:
        return 0
    added = 0
    for row in sheet.rows:
        added += _apply_auto_tags_to_accession(db, row.accession_number, tags)
    return added


@app.get("/request-sheets/{sheet_id}/coverage")
def request_sheet_coverage(sheet_id: int, db: Session = Depends(get_db)):
    """
    For every row in a request sheet, report which scanned slides exist on
    the row's accession, grouped by block. Read-only — does not write back
    to `blocks_available`. The frontend renders this as a side-by-side hint
    next to the user-entered fields.
    """
    sheet = db.query(RequestSheet).filter_by(id=sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    if not indexer:
        raise HTTPException(status_code=503, detail="Indexer not initialized")

    rows = db.query(RequestRow).filter_by(sheet_id=sheet_id).all()
    coverage: dict[int, dict] = {}

    for row in rows:
        acc_norm = _normalize_accession(row.accession_number)
        results = indexer.search(db, query=acc_norm, limit=500)
        # Restrict to exact accession matches (search is fuzzy)
        scanned = [r for r in results if _normalize_accession(r.get("accession_number", "")) == acc_norm]

        # Group scanned slides by block
        blocks_scanned_map: dict[str, list[dict]] = {}
        for s in scanned:
            b = (s.get("block_id") or "").upper() or "(no block)"
            blocks_scanned_map.setdefault(b, []).append({
                "stain_type": s.get("stain_type"),
                "slide_number": s.get("slide_number"),
            })

        requested = _parse_block_list(row.all_blocks)
        scanned_block_ids = set(blocks_scanned_map.keys())

        # Coverage: requested blocks that have at least one scanned slide
        requested_covered = [b for b in requested if b in scanned_block_ids]
        requested_missing = [b for b in requested if b not in scanned_block_ids]
        # Scanned blocks not asked for (informational)
        extra_blocks = sorted(scanned_block_ids - set(requested))

        coverage[row.id] = {
            "requested_blocks": requested,
            "scanned_blocks": sorted(scanned_block_ids),
            "requested_covered": requested_covered,
            "requested_missing": requested_missing,
            "extra_blocks": extra_blocks,
            "slides_total": len(scanned),
            "slides_by_block": blocks_scanned_map,
        }

    return {"sheet_id": sheet_id, "coverage": coverage}


@app.post("/request-sheets")
def create_request_sheet(req: RequestSheetCreate, db: Session = Depends(get_db)):
    sheet = RequestSheet(name=req.name, description=req.description, created_by=req.created_by)
    db.add(sheet)
    db.flush()
    return {
        "id": sheet.id,
        "name": sheet.name,
        "description": sheet.description,
        "case_count": 0,
        "created_by": sheet.created_by,
        "created_at": sheet.created_at.isoformat() if sheet.created_at else None,
        "updated_at": sheet.updated_at.isoformat() if sheet.updated_at else None,
    }


@app.get("/request-sheets/{sheet_id}")
def get_request_sheet(sheet_id: int, db: Session = Depends(get_db)):
    sheet = db.query(RequestSheet).options(
        joinedload(RequestSheet.rows),
        joinedload(RequestSheet.auto_tags),
    ).filter_by(id=sheet_id).first()
    if not sheet:
        raise HTTPException(404, "Sheet not found")
    return {
        "id": sheet.id,
        "name": sheet.name,
        "description": sheet.description,
        "case_count": sheet.case_count,
        "created_by": sheet.created_by,
        "created_at": sheet.created_at.isoformat() if sheet.created_at else None,
        "updated_at": sheet.updated_at.isoformat() if sheet.updated_at else None,
        "auto_tags": [
            {"id": t.id, "name": t.name, "color": t.color}
            for t in sheet.auto_tags
        ],
        "rows": [_serialize_request_row(r) for r in sheet.rows],
    }


@app.patch("/request-sheets/{sheet_id}")
def update_request_sheet(sheet_id: int, req: RequestSheetUpdate, db: Session = Depends(get_db)):
    sheet = db.query(RequestSheet).options(
        joinedload(RequestSheet.rows),
        joinedload(RequestSheet.auto_tags),
    ).filter_by(id=sheet_id).first()
    if not sheet:
        raise HTTPException(404, "Sheet not found")
    if req.name is not None:
        sheet.name = req.name
    if req.description is not None:
        sheet.description = req.description

    auto_tag_applied = 0
    if req.auto_tag_ids is not None:
        # Replace the whole set of auto-tags
        new_ids = list(set(req.auto_tag_ids))
        new_tags = db.query(Tag).filter(Tag.id.in_(new_ids)).all() if new_ids else []
        if len(new_tags) != len(new_ids):
            raise HTTPException(404, "One or more tags not found")
        sheet.auto_tags = new_tags
        db.flush()
        # Apply immediately to every current row's case
        auto_tag_applied = _apply_auto_tags_to_sheet(db, sheet)

    sheet.updated_at = datetime.utcnow()
    db.flush()
    return {"ok": True, "auto_tag_applied": auto_tag_applied}


@app.post("/request-sheets/{sheet_id}/apply-auto-tag")
def apply_auto_tag_now(sheet_id: int, db: Session = Depends(get_db)):
    """Re-apply every auto-tag to every row's case (idempotent)."""
    sheet = db.query(RequestSheet).options(
        joinedload(RequestSheet.rows),
        joinedload(RequestSheet.auto_tags),
    ).filter_by(id=sheet_id).first()
    if not sheet:
        raise HTTPException(404, "Sheet not found")
    if not sheet.auto_tags:
        return {"ok": True, "auto_tag_applied": 0, "message": "No auto-tags set"}
    applied = _apply_auto_tags_to_sheet(db, sheet)
    return {"ok": True, "auto_tag_applied": applied}


@app.delete("/request-sheets/{sheet_id}")
def delete_request_sheet(sheet_id: int, db: Session = Depends(get_db)):
    sheet = db.query(RequestSheet).filter_by(id=sheet_id).first()
    if not sheet:
        raise HTTPException(404, "Sheet not found")
    db.delete(sheet)
    return {"ok": True}


@app.post("/request-sheets/{sheet_id}/rows")
def create_request_row(sheet_id: int, req: RequestRowCreate, db: Session = Depends(get_db)):
    sheet = db.query(RequestSheet).options(joinedload(RequestSheet.auto_tags)).filter_by(id=sheet_id).first()
    if not sheet:
        raise HTTPException(404, "Sheet not found")
    normalized = _normalize_accession(req.accession_number)
    existing = db.query(RequestRow).filter_by(sheet_id=sheet_id, accession_number=normalized).first()
    if existing:
        raise HTTPException(409, f"Accession {normalized} already exists in this sheet")
    data = req.model_dump()
    data['accession_number'] = normalized
    row = RequestRow(sheet_id=sheet_id, **data)
    db.add(row)
    sheet.updated_at = datetime.utcnow()
    # If the sheet has auto-tags, apply them to this new row's case
    if sheet.auto_tags:
        _apply_auto_tags_to_accession(db, normalized, list(sheet.auto_tags))
    db.flush()
    return _serialize_request_row(row)


@app.patch("/request-sheets/{sheet_id}/rows/{row_id}")
def update_request_row(sheet_id: int, row_id: int, req: RequestRowUpdate, db: Session = Depends(get_db)):
    row = db.query(RequestRow).filter_by(id=row_id, sheet_id=sheet_id).first()
    if not row:
        raise HTTPException(404, "Row not found")
    updates = req.model_dump(exclude_unset=True)
    for field, value in updates.items():
        setattr(row, field, value)
    row.updated_at = datetime.utcnow()
    sheet = db.query(RequestSheet).filter_by(id=sheet_id).first()
    if sheet:
        sheet.updated_at = datetime.utcnow()
    db.flush()
    return _serialize_request_row(row)


@app.delete("/request-sheets/{sheet_id}/rows/{row_id}")
def delete_request_row(sheet_id: int, row_id: int, db: Session = Depends(get_db)):
    row = db.query(RequestRow).filter_by(id=row_id, sheet_id=sheet_id).first()
    if not row:
        raise HTTPException(404, "Row not found")
    db.delete(row)
    sheet = db.query(RequestSheet).filter_by(id=sheet_id).first()
    if sheet:
        sheet.updated_at = datetime.utcnow()
    return {"ok": True}


@app.post("/request-sheets/{sheet_id}/import-cohort")
def import_cohort_to_sheet(sheet_id: int, req: ImportCohortRequest, db: Session = Depends(get_db)):
    """Import cases from a cohort into a request sheet."""
    sheet = db.query(RequestSheet).filter_by(id=sheet_id).first()
    if not sheet:
        raise HTTPException(404, "Sheet not found")
    cohort = db.query(Cohort).options(joinedload(Cohort.slides).joinedload(Slide.case)).filter_by(id=req.cohort_id).first()
    if not cohort:
        raise HTTPException(404, "Cohort not found")

    # Group slides by case
    case_slides: dict = {}
    for slide in cohort.slides:
        case = slide.case
        if not case:
            continue
        if case.accession_hash not in case_slides:
            case_slides[case.accession_hash] = {"case": case, "slides": []}
        case_slides[case.accession_hash]["slides"].append(slide)

    added = 0
    skipped = 0
    for acc_hash, data in case_slides.items():
        accession_number = None
        block_ids = set()
        for slide in data["slides"]:
            filepath = indexer.get_filepath(slide.slide_hash) if indexer else None
            if filepath:
                parsed = indexer.parser.parse(Path(filepath).name)
                if parsed and parsed.accession:
                    accession_number = parsed.accession
                if parsed and parsed.block_id:
                    block_ids.add(parsed.block_id)
            elif slide.block_id:
                block_ids.add(slide.block_id)

        if not accession_number:
            accession_number = f"HASH:{acc_hash[:12]}"
        else:
            accession_number = _normalize_accession(accession_number)

        existing = db.query(RequestRow).filter_by(sheet_id=sheet_id, accession_number=accession_number).first()
        if existing:
            skipped += 1
            continue

        blocks_str = ";".join(sorted(block_ids)) if block_ids else None
        row = RequestRow(
            sheet_id=sheet_id,
            accession_number=accession_number,
            case_status="Not Started",
            all_blocks=blocks_str,
            blocks_available=blocks_str,
            hes_requested=len(block_ids),
        )
        db.add(row)
        added += 1

    sheet.updated_at = datetime.utcnow()
    db.flush()
    return {"added": added, "skipped": skipped}


@app.get("/request-sheets/{sheet_id}/export.csv")
def export_request_sheet_csv(sheet_id: int, db: Session = Depends(get_db)):
    sheet = db.query(RequestSheet).options(joinedload(RequestSheet.rows)).filter_by(id=sheet_id).first()
    if not sheet:
        raise HTTPException(404, "Sheet not found")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Accession Number", "Case Status",
        "All Blocks", "Blocks Available", "Order ID", "Consult?",
        "Blocks H&Es Requested", "H&Es Requested", "Non H&Es Requested", "IHC Stains Requested",
        "Block H&Es Received", "H&Es Received", "Unaccounted Blocks",
        "Non H&Es Received", "FS Received", "USS Received", "IHC Received", "IHC Stains Received",
        "Recut Blocks", "Recut Status",
        "H&Es Scanned?", "H&E Scanning Status", "Non H&Es Scanned?",
        "Slide Location", "Notes",
    ])
    for r in sheet.rows:
        writer.writerow([
            r.accession_number, r.case_status,
            r.all_blocks, r.blocks_available, r.order_id, "Yes" if r.is_consult else "No",
            r.blocks_hes_requested, r.hes_requested, r.non_hes_requested, r.ihc_stains_requested,
            r.block_hes_received, r.hes_received, r.unaccounted_blocks,
            r.non_hes_received, r.fs_received, r.uss_received, r.ihc_received, r.ihc_stains_received,
            r.recut_blocks, r.recut_status,
            r.hes_scanned, r.he_scanning_status, r.non_hes_scanned,
            r.slide_location, r.notes,
        ])

    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={sheet.name.replace(' ', '_')}_requests.csv"},
    )


@app.post("/request-sheets/{sheet_id}/import-csv")
async def import_request_sheet_csv(
    sheet_id: int,
    file: UploadFile = File(...),
    mode: str = Query("skip", pattern="^(skip|upsert)$"),
    db: Session = Depends(get_db),
):
    """
    Import rows from a CSV or XLSX file into a request sheet.
    mode=skip (default): rows whose accession already exists are skipped.
    mode=upsert: existing rows are updated with non-empty CSV values; new rows are inserted.
    Returns {added, updated, skipped, errors}.
    """
    sheet = db.query(RequestSheet).filter_by(id=sheet_id).first()
    if not sheet:
        raise HTTPException(404, "Sheet not found")

    content = await file.read()
    if content.startswith(b'\xef\xbb\xbf'):
        content = content[3:]
    filename = file.filename or ""

    rows: list[dict] = []
    if filename.lower().endswith('.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
            ws = wb.active
            headers = None
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if headers is None:
                    headers = [h.lower() for h in cells]
                else:
                    if any(cells):
                        rows.append(dict(zip(headers, cells)))
        except ImportError:
            raise HTTPException(400, "openpyxl not installed — Excel import unavailable")
    else:
        text = content.decode('utf-8', errors='replace')
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.lower().strip(): (v or "").strip() for k, v in row.items()})

    if not rows:
        raise HTTPException(400, "File is empty or could not be parsed")

    COL_MAP: dict[str, str] = {
        "accession number": "accession_number",
        "case status": "case_status",
        "all blocks": "all_blocks",
        "blocks available": "blocks_available",
        "order id": "order_id",
        "consult?": "is_consult",
        "blocks h&es requested": "blocks_hes_requested",
        "h&es requested": "hes_requested",
        "non h&es requested": "non_hes_requested",
        "ihc stains requested": "ihc_stains_requested",
        "block h&es received": "block_hes_received",
        "h&es received": "hes_received",
        "unaccounted blocks": "unaccounted_blocks",
        "non h&es received": "non_hes_received",
        "fs received": "fs_received",
        "uss received": "uss_received",
        "ihc received": "ihc_received",
        "ihc stains received": "ihc_stains_received",
        "recut blocks": "recut_blocks",
        "recut status": "recut_status",
        "h&es scanned?": "hes_scanned",
        "h&e scanning status": "he_scanning_status",
        "non h&es scanned?": "non_hes_scanned",
        "slide location": "slide_location",
        "notes": "notes",
        "accession": "accession_number",
        "accession #": "accession_number",
        "status": "case_status",
    }
    INT_FIELDS = {"hes_requested", "non_hes_requested", "hes_received", "non_hes_received",
                  "fs_received", "uss_received", "ihc_received"}

    existing_rows: dict[str, RequestRow] = {
        r.accession_number: r
        for r in db.query(RequestRow).filter_by(sheet_id=sheet_id).all()
    }

    added = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    def _coerce(field: str, val: str):
        if field == "is_consult":
            return val.lower() in ("yes", "true", "1")
        if field in INT_FIELDS:
            try:
                return int(float(val)) if val else 0
            except (ValueError, TypeError):
                return 0
        return val

    for i, raw in enumerate(rows, start=2):
        mapped: dict = {}
        for raw_key, val in raw.items():
            field = COL_MAP.get(raw_key.lower().strip())
            if field and val != "":
                mapped[field] = val

        accession = mapped.get("accession_number", "").strip()
        if not accession:
            errors.append(f"Row {i}: missing accession number — skipped")
            continue

        accession = _normalize_accession(accession)

        if accession in existing_rows:
            if mode == "upsert":
                row = existing_rows[accession]
                for field, val in mapped.items():
                    if field == "accession_number":
                        continue
                    setattr(row, field, _coerce(field, val))
                updated += 1
            else:
                skipped += 1
            continue

        kwargs: dict = {"sheet_id": sheet_id, "accession_number": accession}
        for field, val in mapped.items():
            if field == "accession_number":
                continue
            kwargs[field] = _coerce(field, val)

        if "case_status" not in kwargs:
            kwargs["case_status"] = "Not Started"

        new_row = RequestRow(**kwargs)
        db.add(new_row)
        existing_rows[accession] = new_row
        added += 1

    sheet.updated_at = datetime.utcnow()
    db.commit()
    return {"added": added, "updated": updated, "skipped": skipped, "errors": errors}


# ============================================================
# Study endpoints
# ============================================================

class StudyCreate(BaseModel):
    name: str
    description: str = ''
    folder_name: str  # Will create slides/studies/{folder_name} on disk
    created_by: str = ''

class StudyGroupCreate(BaseModel):
    name: str
    label: str = ''
    color: str = ''
    note: str = ''
    parent_id: int | None = None
    sort_order: int = 0

class StudyGroupUpdate(BaseModel):
    name: str | None = None
    label: str | None = None
    color: str | None = None
    note: str | None = None
    parent_id: int | None = None
    sort_order: int | None = None


def _serialize_study(study, include_groups=False, include_slides=False):
    d = {
        'id': study.id,
        'name': study.name,
        'description': study.description,
        'folder_name': study.folder_name,
        'created_by': study.created_by,
        'created_at': study.created_at.isoformat() if study.created_at else None,
        'updated_at': study.updated_at.isoformat() if study.updated_at else None,
        'slide_count': study.slide_count,
        'group_count': study.group_count,
        'folder_path': str(settings.slides_path / 'studies' / study.folder_name),
    }
    if include_groups:
        d['groups'] = [_serialize_study_group(g) for g in study.groups]
    if include_slides:
        d['slides'] = [_serialize_study_slide(s) for s in study.slides]
    return d


def _serialize_study_slide(slide):
    """Serialize a slide for study context, including resolved accession if clinical."""
    d = {
        'id': slide.id,
        'slide_hash': slide.slide_hash,
        'block_id': slide.block_id,
        'stain_type': slide.stain_type,
        'random_id': slide.random_id,
        'file_size_bytes': slide.file_size_bytes,
        'file_exists': bool(slide.file_exists),
    }
    # Try to resolve accession from indexer
    if indexer:
        fp = indexer.get_filepath(slide.slide_hash)
        if fp:
            parsed = indexer.parser.parse(fp.name)
            if parsed:
                d['accession_number'] = parsed.accession
                d['slide_number'] = parsed.slide_number
                d['year'] = parsed.year
            else:
                # External/research slide — use filename stem as display name
                d['filename'] = fp.stem
            d['file_path'] = str(fp)
    return d


def _serialize_study_group(group):
    return {
        'id': group.id,
        'study_id': group.study_id,
        'parent_id': group.parent_id,
        'name': group.name,
        'label': group.label,
        'color': group.color,
        'note': group.note,
        'sort_order': group.sort_order,
        'slide_count': len(group.slides),
        'slide_hashes': [s.slide_hash for s in group.slides],
        'children': [_serialize_study_group(c) for c in group.children] if group.children else [],
    }


@app.get("/studies")
def list_studies(db: Session = Depends(get_db)):
    studies = db.query(Study).order_by(Study.updated_at.desc()).all()
    return [_serialize_study(s) for s in studies]


@app.post("/studies")
def create_study(data: StudyCreate, db: Session = Depends(get_db)):
    # Sanitize folder name
    safe_folder = data.folder_name.strip().replace(' ', '_')
    safe_folder = ''.join(c for c in safe_folder if c.isalnum() or c in '-_')
    if not safe_folder:
        raise HTTPException(status_code=400, detail="Invalid folder name")

    # Check for duplicate folder name
    existing = db.query(Study).filter_by(folder_name=safe_folder).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"Study folder '{safe_folder}' already exists")

    # Create folder on disk
    study_dir = settings.slides_path / 'studies' / safe_folder
    try:
        study_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not create folder: {e}")

    study = Study(
        name=data.name,
        description=data.description,
        folder_name=safe_folder,
        created_by=data.created_by,
    )
    db.add(study)
    db.flush()
    return _serialize_study(study)


@app.get("/studies/{study_id}")
def get_study(study_id: int, db: Session = Depends(get_db)):
    study = db.query(Study).filter_by(id=study_id).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")
    return _serialize_study(study, include_groups=True, include_slides=True)


@app.put("/studies/{study_id}")
def update_study(study_id: int, data: StudyCreate, db: Session = Depends(get_db)):
    study = db.query(Study).filter_by(id=study_id).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")
    study.name = data.name
    study.description = data.description
    db.flush()
    return _serialize_study(study)


@app.delete("/studies/{study_id}")
def delete_study(study_id: int, db: Session = Depends(get_db)):
    study = db.query(Study).filter_by(id=study_id).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")
    db.delete(study)
    return {"ok": True}


@app.post("/studies/{study_id}/slides")
def add_slides_to_study(study_id: int, data: dict, db: Session = Depends(get_db)):
    """Add slides to a study by slide_hash list."""
    study = db.query(Study).filter_by(id=study_id).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    hashes = data.get('slide_hashes', [])
    existing_hashes = {s.slide_hash for s in study.slides}
    added = 0
    for h in hashes:
        if h in existing_hashes:
            continue
        slide = db.query(Slide).filter_by(slide_hash=h).first()
        if slide:
            study.slides.append(slide)
            existing_hashes.add(h)
            added += 1
    db.flush()
    return {"added": added, "total": len(study.slides)}


@app.delete("/studies/{study_id}/slides")
def remove_slides_from_study(study_id: int, data: dict, db: Session = Depends(get_db)):
    """Remove slides from a study by slide_hash list."""
    study = db.query(Study).filter_by(id=study_id).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    hashes = set(data.get('slide_hashes', []))
    study.slides = [s for s in study.slides if s.slide_hash not in hashes]
    db.flush()
    return {"ok": True, "total": len(study.slides)}


# ── Study Group endpoints ─────────────────────────────────────

@app.post("/studies/{study_id}/groups")
def create_study_group(study_id: int, data: StudyGroupCreate, db: Session = Depends(get_db)):
    study = db.query(Study).filter_by(id=study_id).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    group = StudyGroup(
        study_id=study_id,
        name=data.name,
        label=data.label or None,
        color=data.color or None,
        note=data.note or None,
        parent_id=data.parent_id,
        sort_order=data.sort_order,
    )
    db.add(group)
    db.flush()
    return _serialize_study_group(group)


@app.put("/studies/{study_id}/groups/{group_id}")
def update_study_group(study_id: int, group_id: int, data: StudyGroupUpdate, db: Session = Depends(get_db)):
    group = db.query(StudyGroup).filter_by(id=group_id, study_id=study_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    for key, val in data.dict(exclude_unset=True).items():
        setattr(group, key, val)
    db.flush()
    return _serialize_study_group(group)


@app.delete("/studies/{study_id}/groups/{group_id}")
def delete_study_group(study_id: int, group_id: int, db: Session = Depends(get_db)):
    group = db.query(StudyGroup).filter_by(id=group_id, study_id=study_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    db.delete(group)
    return {"ok": True}


@app.post("/studies/{study_id}/groups/{group_id}/slides")
def add_slides_to_group(study_id: int, group_id: int, data: dict, db: Session = Depends(get_db)):
    """Add slides to a group within a study."""
    group = db.query(StudyGroup).filter_by(id=group_id, study_id=study_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    hashes = data.get('slide_hashes', [])
    existing_hashes = {s.slide_hash for s in group.slides}
    added = 0
    for h in hashes:
        if h in existing_hashes:
            continue
        slide = db.query(Slide).filter_by(slide_hash=h).first()
        if slide:
            group.slides.append(slide)
            existing_hashes.add(h)
            added += 1
    db.flush()
    return {"added": added, "total": len(group.slides)}


@app.delete("/studies/{study_id}/groups/{group_id}/slides")
def remove_slides_from_group(study_id: int, group_id: int, data: dict, db: Session = Depends(get_db)):
    group = db.query(StudyGroup).filter_by(id=group_id, study_id=study_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    hashes = set(data.get('slide_hashes', []))
    group.slides = [s for s in group.slides if s.slide_hash not in hashes]
    db.flush()
    return {"ok": True, "total": len(group.slides)}


@app.post("/studies/{study_id}/groups/{group_id}/export")
def export_group_slides(study_id: int, group_id: int, data: dict, db: Session = Depends(get_db)):
    """Copy or symlink a group's slides to a target directory."""
    group = db.query(StudyGroup).filter_by(id=group_id, study_id=study_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    target_dir = data.get('target_dir', '').strip()
    use_symlinks = data.get('symlinks', True)
    if not target_dir:
        raise HTTPException(status_code=400, detail="target_dir is required")

    target_path = Path(target_dir)
    try:
        target_path.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not create directory: {e}")

    exported = 0
    errors = []
    for slide in group.slides:
        src = indexer.get_filepath(slide.slide_hash) if indexer else None
        if not src or not src.exists():
            errors.append(f"No file for {slide.slide_hash[:12]}")
            continue
        dest = target_path / src.name
        if dest.exists():
            exported += 1
            continue
        try:
            if use_symlinks:
                dest.symlink_to(src)
            else:
                import shutil
                shutil.copy2(str(src), str(dest))
            exported += 1
        except Exception as e:
            errors.append(f"Failed to export {src.name}: {e}")

    return {"exported": exported, "errors": errors, "target_dir": str(target_path)}


@app.post("/studies/{study_id}/groups/{group_id}/tag")
def tag_group_slides(study_id: int, group_id: int, data: dict, db: Session = Depends(get_db)):
    """Bulk-add a tag to all slides in a group."""
    group = db.query(StudyGroup).filter_by(id=group_id, study_id=study_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    tag_name = data.get('tag_name', '').strip()
    tag_color = data.get('tag_color')
    if not tag_name:
        raise HTTPException(status_code=400, detail="tag_name is required")

    # Get or create tag
    from sqlalchemy import func
    tag = db.query(Tag).filter(func.lower(Tag.name) == tag_name.lower()).first()
    if not tag:
        tag = Tag(name=tag_name, color=tag_color)
        db.add(tag)
        db.flush()
    elif tag_color and not tag.color:
        tag.color = tag_color

    tagged = 0
    for slide in group.slides:
        if tag not in slide.tags:
            slide.tags.append(tag)
            tagged += 1
    db.flush()
    return {"tagged": tagged, "tag_id": tag.id, "tag_name": tag.name}


@app.get("/studies/{study_id}/unlinked-files")
def get_unlinked_files(study_id: int, db: Session = Depends(get_db)):
    """List slide files in the study folder that aren't linked to this study."""
    study = db.query(Study).filter_by(id=study_id).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    study_dir = settings.slides_path / 'studies' / study.folder_name
    if not study_dir.exists():
        return {"files": [], "folder_path": str(study_dir)}

    linked_hashes = {s.slide_hash for s in study.slides}
    hasher = SlideHasher(settings.salt_path)
    unlinked = []
    valid_exts = ('.svs', '.tif', '.tiff', '.ndpi', '.mrxs')

    # Walk study dir recursively (including subdirectories/folders)
    for f in study_dir.rglob('*'):
        if not f.is_file() or f.suffix.lower() not in valid_exts:
            continue
        stem = f.stem
        slide_hash = hasher.hash_slide_stem(stem)
        # Skip if already linked to study
        if slide_hash in linked_hashes:
            continue
        # Check if it already exists in DB (from clinical index)
        existing = db.query(Slide).filter_by(slide_hash=slide_hash).first()
        rel_path = str(f.relative_to(study_dir))
        subfolder = str(f.parent.relative_to(study_dir)) if f.parent != study_dir else None
        unlinked.append({
            'filename': f.name,
            'relative_path': rel_path,
            'subfolder': subfolder,
            'file_size_bytes': f.stat().st_size,
            'extension': f.suffix.lower(),
            'slide_hash': slide_hash,
            'in_database': existing is not None,
        })

    return {"files": unlinked, "folder_path": str(study_dir)}


@app.post("/studies/{study_id}/import-files")
def import_study_files(study_id: int, data: dict, db: Session = Depends(get_db)):
    """
    Import unlinked files from the study folder into the database and link to study.

    For files already in the DB (clinical slides), just links them to the study.
    For new files (outside hospital, hand-labeled), creates a Case + Slide record.

    Body: { "filenames": ["slide1.svs", "slide2.svs"] }
    Optional: { "filenames": [...], "group_id": 123 } to also add to a group.
    """
    study = db.query(Study).filter_by(id=study_id).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    filenames = data.get('filenames', [])
    group_id = data.get('group_id')
    if not filenames:
        raise HTTPException(status_code=400, detail="No filenames provided")

    study_dir = settings.slides_path / 'studies' / study.folder_name
    hasher = SlideHasher(settings.salt_path)
    valid_exts = ('.svs', '.tif', '.tiff', '.ndpi', '.mrxs')

    linked_hashes = {s.slide_hash for s in study.slides}
    group = None
    if group_id:
        group = db.query(StudyGroup).filter_by(id=group_id, study_id=study_id).first()

    imported = 0
    linked = 0
    errors = []

    for filename in filenames:
        # Find the file in the study directory (could be in a subfolder)
        matches = list(study_dir.rglob(filename))
        if not matches:
            errors.append(f"File not found: {filename}")
            continue
        filepath = matches[0]
        if filepath.suffix.lower() not in valid_exts:
            errors.append(f"Invalid file type: {filename}")
            continue

        stem = filepath.stem
        slide_hash = hasher.hash_slide_stem(stem)

        # Check if already linked to study
        if slide_hash in linked_hashes:
            continue

        # Check if slide exists in DB
        slide = db.query(Slide).filter_by(slide_hash=slide_hash).first()

        if not slide:
            # New external slide — create a Case + Slide record
            # Use filename stem as accession hash (these don't follow standard naming)
            case_hash = hasher.hash_accession(stem)
            case = db.query(Case).filter_by(accession_hash=case_hash).first()
            if not case:
                case = Case(accession_hash=case_hash, year=0)
                db.add(case)
                db.flush()

            try:
                file_size = filepath.stat().st_size
            except OSError:
                file_size = None

            slide = Slide(
                slide_hash=slide_hash,
                case_id=case.id,
                file_size_bytes=file_size,
                file_exists=True,
            )
            db.add(slide)
            db.flush()
            imported += 1

            # Register in indexer path cache so analysis submission can find the file
            if indexer:
                indexer.slide_hash_to_path[slide_hash] = filepath
        else:
            linked += 1

        # Link to study
        study.slides.append(slide)
        linked_hashes.add(slide_hash)

        # Also add to group if specified
        if group and slide not in group.slides:
            group.slides.append(slide)

    db.flush()
    return {
        "imported": imported,
        "linked": linked,
        "errors": errors,
        "total": len(study.slides),
    }


# ============================================================
# Patient Management & SlideCap ID System
# ============================================================


class PatientCreateRequest(BaseModel):
    note: Optional[str] = None


class PatientUpdateRequest(BaseModel):
    note: Optional[str] = None


class AssignCasesRequest(BaseModel):
    case_hashes: List[str]  # accession_hashes to assign to this patient


class ExternalMappingRequest(BaseModel):
    external_system: str       # "redcap", "epic", etc.
    external_project: Optional[str] = None  # REDCap project name
    external_id: str           # The trial/subject ID


class BulkPatientImportRow(BaseModel):
    accession_number: str      # Raw accession (will be hashed for lookup)
    patient_label: Optional[str] = None  # If provided, group by this label
    external_system: Optional[str] = None
    external_project: Optional[str] = None
    external_id: Optional[str] = None


class BulkPatientImportRequest(BaseModel):
    rows: List[BulkPatientImportRow]


# --- Patient CRUD ---

@app.get("/patients")
def list_patients(db: Session = Depends(get_db)):
    """List all patients with case/slide counts and external mappings."""
    patients = db.query(Patient).options(
        joinedload(Patient.cases).joinedload(Case.slides),
        joinedload(Patient.external_mappings),
    ).all()

    return [{
        "id": p.id,
        "slidecap_id": p.slidecap_id,
        "note": p.note,
        "case_count": p.case_count,
        "slide_count": p.slide_count,
        "cases": [{
            "slidecap_id": c.slidecap_id,
            "accession_hash": c.accession_hash[:12] + "...",
            "year": c.year,
            "slide_count": len(c.slides),
        } for c in p.cases],
        "external_mappings": [{
            "id": m.id,
            "system": m.external_system,
            "project": m.external_project,
            "external_id": m.external_id,
        } for m in p.external_mappings],
        "created_at": p.created_at.isoformat() if p.created_at else None,
    } for p in patients]


@app.post("/patients")
def create_patient(data: PatientCreateRequest, db: Session = Depends(get_db)):
    """Create a new patient. Auto-assigns PT ID."""
    patient = Patient(
        slidecap_id=generate_slidecap_id(db, "PT"),
        note=data.note,
    )
    db.add(patient)
    db.flush()
    return {
        "id": patient.id,
        "slidecap_id": patient.slidecap_id,
        "note": patient.note,
    }


@app.get("/patients/{patient_sid}")
def get_patient(patient_sid: str, db: Session = Depends(get_db)):
    """
    Get full patient hierarchy: patient → cases → slides → analysis results.
    Accepts SlideCap ID (PT00001) or numeric DB id.
    """
    patient = _resolve_patient(db, patient_sid)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    cases_out = []
    for case in patient.cases:
        slides_out = []
        for slide in case.slides:
            # Get analysis results for this slide
            analyses_out = []
            for js in slide.job_slides:
                if js.job:
                    analyses_out.append({
                        "job_slidecap_id": js.job.slidecap_id,
                        "model_name": js.job.model_name,
                        "status": js.status,
                        "completed_at": js.completed_at.isoformat() if js.completed_at else None,
                    })
            slides_out.append({
                "slidecap_id": slide.slidecap_id,
                "slide_hash": slide.slide_hash,
                "block_id": slide.block_id,
                "stain_type": slide.stain_type,
                "file_size_bytes": slide.file_size_bytes,
                "tags": [t.name for t in slide.tags],
                "analyses": analyses_out,
            })
        cases_out.append({
            "slidecap_id": case.slidecap_id,
            "accession_hash": case.accession_hash[:12] + "...",
            "year": case.year,
            "tags": [t.name for t in case.tags],
            "slides": slides_out,
        })

    return {
        "id": patient.id,
        "slidecap_id": patient.slidecap_id,
        "note": patient.note,
        "external_mappings": [{
            "id": m.id,
            "system": m.external_system,
            "project": m.external_project,
            "external_id": m.external_id,
        } for m in patient.external_mappings],
        "cases": cases_out,
        "created_at": patient.created_at.isoformat() if patient.created_at else None,
    }


@app.patch("/patients/{patient_sid}")
def update_patient(patient_sid: str, data: PatientUpdateRequest, db: Session = Depends(get_db)):
    """Update patient note."""
    patient = _resolve_patient(db, patient_sid)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    if data.note is not None:
        patient.note = data.note
    return {"slidecap_id": patient.slidecap_id, "note": patient.note}


@app.delete("/patients/{patient_sid}")
def delete_patient(patient_sid: str, db: Session = Depends(get_db)):
    """Delete a patient. Cases are unlinked (not deleted)."""
    patient = _resolve_patient(db, patient_sid)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    # Unlink cases before deleting patient
    for case in patient.cases:
        case.patient_id = None
    db.delete(patient)
    return {"deleted": patient.slidecap_id}


# --- Case ↔ Patient Assignment ---

@app.post("/patients/{patient_sid}/cases")
def assign_cases_to_patient(patient_sid: str, data: AssignCasesRequest, db: Session = Depends(get_db)):
    """
    Assign cases (by accession_hash) to a patient.
    Cases already assigned to another patient will be reassigned.
    """
    patient = _resolve_patient(db, patient_sid)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    assigned = []
    not_found = []
    for case_hash in data.case_hashes:
        case = db.query(Case).filter_by(accession_hash=case_hash).first()
        if not case:
            not_found.append(case_hash[:12])
            continue
        case.patient_id = patient.id
        assigned.append(case.slidecap_id)

    return {
        "patient": patient.slidecap_id,
        "assigned": assigned,
        "not_found": not_found,
    }


@app.delete("/patients/{patient_sid}/cases/{case_sid}")
def unassign_case_from_patient(patient_sid: str, case_sid: str, db: Session = Depends(get_db)):
    """Remove a case from a patient (unlink, not delete)."""
    patient = _resolve_patient(db, patient_sid)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    case = _resolve_case(db, case_sid)
    if not case or case.patient_id != patient.id:
        raise HTTPException(status_code=404, detail="Case not found on this patient")
    case.patient_id = None
    return {"unassigned": case.slidecap_id, "from_patient": patient.slidecap_id}


# --- External Mappings (REDCap Integration) ---

@app.post("/patients/{patient_sid}/mappings")
def add_external_mapping(patient_sid: str, data: ExternalMappingRequest, db: Session = Depends(get_db)):
    """Add an external system mapping (e.g., REDCap trial ID) to a patient."""
    patient = _resolve_patient(db, patient_sid)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    # Check for duplicate
    existing = db.query(ExternalMapping).filter_by(
        external_system=data.external_system,
        external_project=data.external_project,
        external_id=data.external_id,
    ).first()
    if existing:
        if existing.patient_id == patient.id:
            return {"message": "Mapping already exists", "id": existing.id}
        raise HTTPException(status_code=409,
                            detail=f"External ID already mapped to {existing.patient.slidecap_id}")

    mapping = ExternalMapping(
        patient_id=patient.id,
        external_system=data.external_system,
        external_project=data.external_project,
        external_id=data.external_id,
    )
    db.add(mapping)
    db.flush()
    return {
        "id": mapping.id,
        "patient": patient.slidecap_id,
        "system": mapping.external_system,
        "project": mapping.external_project,
        "external_id": mapping.external_id,
    }


@app.delete("/patients/{patient_sid}/mappings/{mapping_id}")
def remove_external_mapping(patient_sid: str, mapping_id: int, db: Session = Depends(get_db)):
    """Remove an external mapping from a patient."""
    patient = _resolve_patient(db, patient_sid)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    mapping = db.query(ExternalMapping).filter_by(id=mapping_id, patient_id=patient.id).first()
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")
    db.delete(mapping)
    return {"deleted": mapping_id}


@app.get("/mappings/lookup")
def lookup_by_external_id(
    system: str,
    external_id: str,
    project: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Reverse lookup: given a REDCap trial ID, find the SlideCap patient.
    Returns full patient hierarchy.
    """
    q = db.query(ExternalMapping).filter_by(
        external_system=system,
        external_id=external_id,
    )
    if project:
        q = q.filter_by(external_project=project)
    mapping = q.first()
    if not mapping:
        raise HTTPException(status_code=404, detail="No mapping found for this external ID")

    # Redirect to full patient view
    return get_patient(mapping.patient.slidecap_id, db)


# --- Bulk Import (CSV-friendly) ---

@app.post("/patients/import")
def bulk_import_patients(data: BulkPatientImportRequest, db: Session = Depends(get_db)):
    """
    Bulk import patient-case assignments and optional external mappings.

    Workflow for scanning:
    1. Scan slides → auto-creates CS/SL IDs
    2. Upload CSV with columns: accession_number, patient_label, [external_system, external_id]
    3. This endpoint groups cases by patient_label, creates Patient records, assigns cases,
       and optionally creates external mappings.

    If patient_label is omitted, each case gets its own new patient.
    """
    if not hasher:
        raise HTTPException(status_code=503, detail="Hasher not initialized")

    # Group rows by patient_label
    label_groups: dict[str, list[BulkPatientImportRow]] = {}
    for row in data.rows:
        label = row.patient_label or f"_auto_{row.accession_number}"
        label_groups.setdefault(label, []).append(row)

    results = {
        "patients_created": 0,
        "cases_assigned": 0,
        "mappings_created": 0,
        "errors": [],
    }

    for label, rows in label_groups.items():
        # Create patient
        patient = Patient(
            slidecap_id=generate_slidecap_id(db, "PT"),
            note=f"Imported: {label}" if not label.startswith("_auto_") else None,
        )
        db.add(patient)
        db.flush()
        results["patients_created"] += 1

        for row in rows:
            # Find case by accession hash
            acc_hash = hasher.hash_accession(row.accession_number)
            case = db.query(Case).filter_by(accession_hash=acc_hash).first()
            if not case:
                results["errors"].append(f"Case not found for accession: {row.accession_number}")
                continue
            case.patient_id = patient.id
            results["cases_assigned"] += 1

            # Create external mapping if provided
            if row.external_system and row.external_id:
                existing = db.query(ExternalMapping).filter_by(
                    external_system=row.external_system,
                    external_project=row.external_project,
                    external_id=row.external_id,
                ).first()
                if not existing:
                    mapping = ExternalMapping(
                        patient_id=patient.id,
                        external_system=row.external_system,
                        external_project=row.external_project,
                        external_id=row.external_id,
                    )
                    db.add(mapping)
                    results["mappings_created"] += 1

    return results


@app.post("/patients/import-csv")
def import_patients_csv(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Import patient assignments from a CSV file.

    Expected columns:
    - accession_number (required): The raw accession number (e.g., "S24-12345")
    - patient_label (optional): Group rows with the same label into one patient
    - external_system (optional): e.g., "redcap"
    - external_project (optional): e.g., "Melanoma Trial 2024"
    - external_id (optional): e.g., "REC-0045"

    Example CSV:
        accession_number,patient_label,external_system,external_id
        S24-12345,Patient_A,redcap,REC-0045
        S24-12346,Patient_A,redcap,REC-0045
        S24-99999,Patient_B,redcap,REC-0101
    """
    content = file.file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))

    rows = []
    for row in reader:
        if "accession_number" not in row:
            raise HTTPException(status_code=400, detail="CSV must have 'accession_number' column")
        rows.append(BulkPatientImportRow(
            accession_number=row["accession_number"].strip(),
            patient_label=row.get("patient_label", "").strip() or None,
            external_system=row.get("external_system", "").strip() or None,
            external_project=row.get("external_project", "").strip() or None,
            external_id=row.get("external_id", "").strip() or None,
        ))

    return bulk_import_patients(BulkPatientImportRequest(rows=rows), db)


# --- Unassigned Cases View ---

@app.get("/cases/unassigned")
def list_unassigned_cases(db: Session = Depends(get_db)):
    """List cases that have no patient assigned. Useful after scanning new slides."""
    cases = db.query(Case).filter(
        Case.patient_id.is_(None)
    ).options(
        joinedload(Case.slides),
    ).order_by(Case.year.desc(), Case.id.desc()).all()

    return [{
        "slidecap_id": c.slidecap_id,
        "accession_hash": c.accession_hash[:12] + "...",
        "year": c.year,
        "slide_count": len(c.slides),
        "slides": [{
            "slidecap_id": s.slidecap_id,
            "block_id": s.block_id,
            "stain_type": s.stain_type,
        } for s in c.slides],
    } for c in cases]


# --- SlideCap ID Lookup ---

@app.get("/lookup/{slidecap_id}")
def lookup_slidecap_id(slidecap_id: str, db: Session = Depends(get_db)):
    """
    Universal lookup: given any SlideCap ID (PT/CS/SL/JB), return the object
    and its connections.
    """
    prefix = slidecap_id[:2].upper()
    if prefix == "PT":
        return get_patient(slidecap_id, db)
    elif prefix == "CS":
        case = db.query(Case).filter_by(slidecap_id=slidecap_id).options(
            joinedload(Case.patient),
            joinedload(Case.slides),
            joinedload(Case.tags),
        ).first()
        if not case:
            raise HTTPException(status_code=404, detail="Case not found")
        return {
            "type": "case",
            "slidecap_id": case.slidecap_id,
            "patient_id": case.patient.slidecap_id if case.patient else None,
            "accession_hash": case.accession_hash[:12] + "...",
            "year": case.year,
            "tags": [t.name for t in case.tags],
            "slides": [{
                "slidecap_id": s.slidecap_id,
                "block_id": s.block_id,
                "stain_type": s.stain_type,
            } for s in case.slides],
        }
    elif prefix == "SL":
        slide = db.query(Slide).filter_by(slidecap_id=slidecap_id).options(
            joinedload(Slide.case).joinedload(Case.patient),
            joinedload(Slide.tags),
            joinedload(Slide.job_slides).joinedload(JobSlide.job),
        ).first()
        if not slide:
            raise HTTPException(status_code=404, detail="Slide not found")
        return {
            "type": "slide",
            "slidecap_id": slide.slidecap_id,
            "case_id": slide.case.slidecap_id if slide.case else None,
            "patient_id": slide.case.patient.slidecap_id if slide.case and slide.case.patient else None,
            "slide_hash": slide.slide_hash,
            "block_id": slide.block_id,
            "stain_type": slide.stain_type,
            "tags": [t.name for t in slide.tags],
            "analyses": [{
                "job_slidecap_id": js.job.slidecap_id if js.job else None,
                "model_name": js.job.model_name if js.job else None,
                "status": js.status,
            } for js in slide.job_slides],
        }
    elif prefix == "JB":
        job = db.query(AnalysisJob).filter_by(slidecap_id=slidecap_id).options(
            joinedload(AnalysisJob.slides).joinedload(JobSlide.slide),
        ).first()
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        return {
            "type": "job",
            "slidecap_id": job.slidecap_id,
            "model_name": job.model_name,
            "status": job.status,
            "slides": [{
                "slide_slidecap_id": js.slide.slidecap_id if js.slide else None,
                "status": js.status,
            } for js in job.slides],
        }
    else:
        raise HTTPException(status_code=400, detail=f"Unknown SlideCap ID prefix: {prefix}")


# --- Helper functions ---

def _resolve_patient(db: Session, sid: str) -> Optional[Patient]:
    """Resolve a patient by SlideCap ID (PT00001) or numeric DB id."""
    if sid.upper().startswith("PT"):
        return db.query(Patient).filter_by(slidecap_id=sid.upper()).options(
            joinedload(Patient.cases).joinedload(Case.slides).joinedload(Slide.tags),
            joinedload(Patient.cases).joinedload(Case.slides).joinedload(Slide.job_slides).joinedload(JobSlide.job),
            joinedload(Patient.cases).joinedload(Case.tags),
            joinedload(Patient.external_mappings),
        ).first()
    try:
        return db.query(Patient).filter_by(id=int(sid)).options(
            joinedload(Patient.cases).joinedload(Case.slides),
            joinedload(Patient.external_mappings),
        ).first()
    except ValueError:
        return None


def _resolve_case(db: Session, sid: str) -> Optional[Case]:
    """Resolve a case by SlideCap ID (CS00001) or accession_hash."""
    if sid.upper().startswith("CS"):
        return db.query(Case).filter_by(slidecap_id=sid.upper()).first()
    return db.query(Case).filter_by(accession_hash=sid).first()


# ============================================================
# Run with uvicorn
# ============================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "app.main:app",
        host=settings.HOST,
        port=settings.PORT,
        reload=True  # Auto-reload on code changes
    )
